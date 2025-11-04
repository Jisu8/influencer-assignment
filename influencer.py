import streamlit as st
import pandas as pd
import numpy as np
import os
import time
import io
import subprocess
from datetime import datetime, timedelta
import requests
import json
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# Snowflake 연결 (선택적)
try:
    import snowflake.connector
    SNOWFLAKE_AVAILABLE = True
except ImportError:
    SNOWFLAKE_AVAILABLE = False

# 환경 감지 함수
def is_running_on_streamlit_cloud():
    """Streamlit Cloud에서 실행 중인지 확인"""
    cloud_indicators = [
        'STREAMLIT_SERVER_HEADLESS',
        'STREAMLIT_SERVER_PORT',
        'STREAMLIT_SERVER_ADDRESS',
        'STREAMLIT_CLOUD_ENVIRONMENT',
        'STREAMLIT_SERVER_RUN_ON_SAVE',
        'STREAMLIT_SERVER_FILE_WATCHER_TYPE'
    ]
    
    cloud_path_indicators = [
        '/app',
        '/home/appuser',
        '/opt/streamlit'
    ]
    
    env_check = any(os.environ.get(indicator) for indicator in cloud_indicators)
    path_check = any(os.path.exists(path) for path in cloud_path_indicators)
    
    return env_check or path_check

# =============================================================================
# 파일 경로 설정
# =============================================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "data")

# 데이터 파일 경로
ASSIGNMENT_FILE = os.path.join(DATA_DIR, "assignment_history.csv")
EXECUTION_FILE = os.path.join(DATA_DIR, "execution_data.csv")
INFLUENCER_FILE = os.path.join(DATA_DIR, "influencer.csv")
SALES_FILE = os.path.join(DATA_DIR, "sales_data.csv")
SEARCH_FILE = os.path.join(DATA_DIR, "search_data.csv")
MONTHLY_TARGETS_FILE = os.path.join(DATA_DIR, "monthly_assignment_targets.csv")
SEARCH_QUERY_FILE = os.path.join(DATA_DIR, "search_query.sql")
SALES_QUERY_FILE = os.path.join(DATA_DIR, "sales_query.sql")

# 데이터 디렉토리 생성
os.makedirs(DATA_DIR, exist_ok=True)

# 상수 정의
BRANDS = ["MLB", "DX", "DV", "ST"]
BRAND_OPTIONS = ["전체"] + BRANDS
MONTHS = ["1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]

# 시즌별 월 매핑
SEASON_MONTHS = {
    "25FW": ["9월", "10월", "11월", "12월", "1월", "2월"],
    "25SS": ["3월", "4월", "5월", "6월", "7월", "8월"],
    "26SS": ["3월", "4월", "5월", "6월", "7월", "8월"]
}

# =============================================================================
# 데이터 로드 및 저장 함수들
# =============================================================================

@st.cache_data
def load_influencer_data():
    """인플루언서 데이터 로드"""
    if os.path.exists(INFLUENCER_FILE):
        return pd.read_csv(INFLUENCER_FILE)
    return pd.DataFrame()

@st.cache_data
def load_assignment_history():
    """배정 이력 데이터 로드"""
    if os.path.exists(ASSIGNMENT_FILE):
        return pd.read_csv(ASSIGNMENT_FILE)
    return pd.DataFrame()

@st.cache_data
def load_execution_data():
    """집행 데이터 로드"""
    if os.path.exists(EXECUTION_FILE):
        return pd.read_csv(EXECUTION_FILE)
    return pd.DataFrame()

@st.cache_data
def load_sales_data():
    """매출 데이터 로드"""
    if os.path.exists(SALES_FILE):
        df = pd.read_csv(SALES_FILE)
        
        # 잘못된 날짜 데이터 필터링
        if 'DT' in df.columns:
            df['DT'] = pd.to_datetime(df['DT'])
            current_year = pd.Timestamp.now().year
            
            # 현실적인 날짜만 유지 (현재 연도 + 1년까지만)
            df = df[df['DT'].dt.year <= current_year + 1]
            
            # 1900년 이전의 데이터도 제거
            df = df[df['DT'].dt.year >= 1900]
        
        return df
    return pd.DataFrame()

def load_influencer_data():
    """인플루언서 데이터 로드"""
    if os.path.exists(INFLUENCER_FILE):
        return pd.read_csv(INFLUENCER_FILE)
    return pd.DataFrame()

def load_assignment_history():
    """배정 이력 데이터 로드"""
    if os.path.exists(ASSIGNMENT_FILE):
        return pd.read_csv(ASSIGNMENT_FILE)
    return pd.DataFrame()

def save_assignment_history(df):
    """배정 이력 데이터 저장"""
    df.to_csv(ASSIGNMENT_FILE, index=False, encoding='utf-8-sig')

def load_monthly_targets():
    """월별 배정 목표 데이터 로드"""
    if os.path.exists(MONTHLY_TARGETS_FILE):
        return pd.read_csv(MONTHLY_TARGETS_FILE)
    return pd.DataFrame()

def save_monthly_targets(df):
    """월별 배정 목표 데이터 저장"""
    df.to_csv(MONTHLY_TARGETS_FILE, index=False, encoding='utf-8-sig')


def save_assignment_history(df):
    """배정 이력 데이터 저장"""
    df.to_csv(ASSIGNMENT_FILE, index=False, encoding='utf-8-sig')
    st.cache_data.clear()

def save_execution_data(df):
    """집행 데이터 저장"""
    df.to_csv(EXECUTION_FILE, index=False, encoding='utf-8-sig')
    st.cache_data.clear()

def save_sales_data(df):
    """매출 데이터 저장"""
    df.to_csv(SALES_FILE, index=False, encoding='utf-8-sig')
    st.cache_data.clear()

def save_monthly_targets(df):
    """월별 배정 목표 데이터 저장"""
    df.to_csv(MONTHLY_TARGETS_FILE, index=False, encoding='utf-8-sig')
    st.cache_data.clear()

# =============================================================================
# 대시보드 관련 함수들
# =============================================================================

def render_dashboard_tab():
    """대시보드 탭 렌더링"""
    st.markdown("# 📊 대시보드")
    
    # 데이터 로드
    execution_df = load_execution_data()
    sales_df = load_sales_data()
    
    if execution_df.empty:
        st.warning("집행 데이터가 없습니다.")
        return
    
    # 인플루언서 노출량 데이터와 연계 분석
    st.markdown("## 📊 매출 연계 분석")
    
    # 브랜드, 아이템, 시즌, 기간 필터 추가
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("#### 🏷️ 브랜드")
        # 브랜드 매핑 설정
        brand_mapping = {
            'MLB': 'M',
            'DX': 'X', 
            'DV': 'V',
            'ST': 'ST'
        }
        
        # 집행 데이터에서 브랜드 목록 가져오기 (MLB 우선순위)
        if '브랜드' in execution_df.columns:
            available_brands = execution_df['브랜드'].unique().tolist()
            # MLB를 첫 번째로 정렬
            if 'MLB' in available_brands:
                available_brands.remove('MLB')
                available_brands = ['MLB'] + available_brands
        else:
            available_brands = []
            
        selected_brand = st.selectbox(
            "분석할 브랜드를 선택하세요", 
            options=available_brands,
            index=0,
            key="brand_filter_trend"
        )
    
    with col2:
        st.markdown("#### 📦 아이템")
        # 매출 데이터에서 아이템 목록 가져오기
        if 'ITEM' in sales_df.columns:
            available_items = sales_df['ITEM'].unique().tolist()
        else:
            available_items = []
            
        selected_item = st.selectbox(
            "분석할 아이템을 선택하세요", 
            options=["전체"] + available_items,
            index=0,
            key="item_filter_trend"
        )
    
    with col3:
        st.markdown("#### 🌟 시즌")
        # 시즌 필터 옵션
        season_options = ["전체", "24FW", "25SS", "25FW"]
        selected_season = st.selectbox(
            "분석할 시즌을 선택하세요",
            options=season_options,
            index=0,
            key="season_filter_trend"
        )
    
    with col4:
        st.markdown("#### 📅 기간 선택")
        # 날짜 범위 설정 (매출 데이터와 집행 데이터 모두 고려)
        all_dates = []
        
        # 매출 데이터에서 날짜 수집
        if not sales_df.empty and 'DT' in sales_df.columns:
            try:
                sales_df['DT'] = pd.to_datetime(sales_df['DT'], errors='coerce')
                sales_dates = sales_df['DT'].dropna()
                all_dates.extend(sales_dates.tolist())
            except Exception as e:
                st.warning(f"매출 데이터 날짜 처리 중 오류: {e}")
        
        # 집행 데이터에서 날짜 수집
        execution_df = load_execution_data()
        if not execution_df.empty:
            # 집행 데이터의 날짜 컬럼 찾기
            date_columns = ['업로드일', 'DT', '날짜', 'date']
            for date_col in date_columns:
                if date_col in execution_df.columns:
                    try:
                        execution_df[date_col] = pd.to_datetime(execution_df[date_col], errors='coerce')
                        exec_dates = execution_df[date_col].dropna()
                        all_dates.extend(exec_dates.tolist())
                        break
                    except Exception as e:
                        continue
        
        if all_dates:
            try:
                all_dates = pd.to_datetime(all_dates, errors='coerce').dropna()
                
                # 유효한 날짜 범위 필터링 (1900년 이전과 2030년 이후 제외)
                valid_dates = all_dates[
                    (all_dates >= pd.to_datetime('2020-01-01')) & 
                    (all_dates <= pd.to_datetime('2030-12-31'))
                ]
                
                if len(valid_dates) > 0:
                    # 시즌 선택에 따른 날짜 범위 설정
                    if selected_season == "24FW":
                        season_min = pd.to_datetime('2024-09-01')
                        season_max = pd.to_datetime('2025-02-28')
                    elif selected_season == "25SS":
                        season_min = pd.to_datetime('2025-03-01')
                        season_max = pd.to_datetime('2025-08-31')
                    elif selected_season == "25FW":
                        season_min = pd.to_datetime('2025-09-01')
                        season_max = pd.to_datetime('2026-02-28')
                    else:
                        # 전체 선택 시 모든 데이터 범위 사용
                        season_min = valid_dates.min()
                        season_max = valid_dates.max()
                    
                    # 시즌 범위와 실제 데이터 범위의 교집합
                    min_date = max(season_min, valid_dates.min())
                    max_date = min(season_max, valid_dates.max())
                    
                    # 날짜 슬라이더
                    date_range = st.slider(
                        "분석할 기간을 선택하세요",
                        min_value=min_date.date(),
                        max_value=max_date.date(),
                        value=(min_date.date(), max_date.date()),
                        format="YYYY-MM-DD",
                        key="date_range_slider_trend"
                    )
                    
                else:
                    st.warning("유효한 날짜 데이터를 찾을 수 없습니다.")
                    date_range = None
            except Exception as e:
                st.error(f"날짜 처리 중 오류가 발생했습니다: {e}")
                date_range = None
        else:
            st.warning("날짜 데이터가 없습니다.")
            date_range = None
    
    if not sales_df.empty and 'ITEM' in sales_df.columns and not execution_df.empty:
        # 마케팅 데이터도 로드
        marketing_df = load_marketing_data()
        
        # 인플루언서 데이터에서 노출수 컬럼 찾기
        influencer_exposure_col = None
        influencer_date_col = None
        
        for col in execution_df.columns:
            if '노출수' in col:
                influencer_exposure_col = col
            if '업로드일' in col or '날짜' in col or 'DT' in col:
                influencer_date_col = col
        
        # 마케팅 데이터에서 노출수 컬럼 찾기
        marketing_exposure_col = None
        marketing_date_col = None
        
        if not marketing_df.empty:
            for col in marketing_df.columns:
                if '노출수' in col:
                    marketing_exposure_col = col
                if '업로드일' in col or '날짜' in col or 'DT' in col:
                    marketing_date_col = col
        
        if influencer_exposure_col and influencer_date_col:
            # 집행 데이터 일자별 노출수 집계
            execution_df_copy = execution_df.copy()
            
            # 집행 데이터와 마케팅 데이터의 노출수를 합치기
            all_exposure_data = []
            
            # 인플루언서 데이터 처리
            try:
                influencer_df_copy = execution_df.copy()
                influencer_df_copy[influencer_date_col] = pd.to_datetime(influencer_df_copy[influencer_date_col], errors='coerce')
                influencer_df_copy = influencer_df_copy.dropna(subset=[influencer_date_col])
                
                if not influencer_df_copy.empty:
                    influencer_exposure = influencer_df_copy.groupby(influencer_date_col)[influencer_exposure_col].sum().reset_index()
                    influencer_exposure.columns = ['DT', '노출수']
                    all_exposure_data.append(influencer_exposure)
            except Exception as e:
                st.warning(f"인플루언서 데이터 처리 중 오류: {e}")
            
            # 마케팅 데이터 처리
            if marketing_exposure_col and marketing_date_col and not marketing_df.empty:
                try:
                    marketing_df_copy = marketing_df.copy()
                    marketing_df_copy[marketing_date_col] = pd.to_datetime(marketing_df_copy[marketing_date_col], errors='coerce')
                    marketing_df_copy = marketing_df_copy.dropna(subset=[marketing_date_col])
                    
                    # 마케팅 데이터에도 시즌 필터 적용
                    if selected_season != "전체":
                        if selected_season == "24FW":
                            season_start = pd.to_datetime('2024-09-01')
                            season_end = pd.to_datetime('2025-02-28')
                        elif selected_season == "25SS":
                            season_start = pd.to_datetime('2025-03-01')
                            season_end = pd.to_datetime('2025-08-31')
                        elif selected_season == "25FW":
                            season_start = pd.to_datetime('2025-09-01')
                            season_end = pd.to_datetime('2026-02-28')
                        else:
                            season_start = None
                            season_end = None
                        
                        if season_start and season_end:
                            marketing_df_copy = marketing_df_copy[
                                (marketing_df_copy[marketing_date_col] >= season_start) & 
                                (marketing_df_copy[marketing_date_col] <= season_end)
                            ]
                    
                    # 기간 필터 적용 (마케팅 데이터)
                    if date_range:
                        start_date = pd.to_datetime(date_range[0])
                        end_date = pd.to_datetime(date_range[1])
                        marketing_df_copy = marketing_df_copy[
                            (marketing_df_copy[marketing_date_col] >= start_date) & 
                            (marketing_df_copy[marketing_date_col] <= end_date)
                        ]
                    
                    if not marketing_df_copy.empty:
                        marketing_exposure = marketing_df_copy.groupby(marketing_date_col)[marketing_exposure_col].sum().reset_index()
                        marketing_exposure.columns = ['DT', '노출수']
                        all_exposure_data.append(marketing_exposure)
                except Exception as e:
                    st.warning(f"마케팅 데이터 처리 중 오류: {e}")
            
            # 유형별 노출수 데이터 처리
            if all_exposure_data:
                # 모든 데이터를 합치고 유형별로 분리
                all_data = []
                
                # 인플루언서 데이터 처리
                if not execution_df.empty and '유형' in execution_df.columns:
                    execution_df_copy = execution_df.copy()
                    execution_df_copy[influencer_date_col] = pd.to_datetime(execution_df_copy[influencer_date_col], errors='coerce')
                    execution_df_copy = execution_df_copy.dropna(subset=[influencer_date_col])
                    
                    # 시즌 필터 적용
                    if selected_season != "전체":
                        if selected_season == "24FW":
                            season_start = pd.to_datetime('2024-09-01')
                            season_end = pd.to_datetime('2025-02-28')
                        elif selected_season == "25SS":
                            season_start = pd.to_datetime('2025-03-01')
                            season_end = pd.to_datetime('2025-08-31')
                        elif selected_season == "25FW":
                            season_start = pd.to_datetime('2025-09-01')
                            season_end = pd.to_datetime('2026-02-28')
                        else:
                            season_start = None
                            season_end = None
                        
                        if season_start and season_end:
                            execution_df_copy = execution_df_copy[
                                (execution_df_copy[influencer_date_col] >= season_start) & 
                                (execution_df_copy[influencer_date_col] <= season_end)
                            ]
                    
                    # 기간 필터 적용
                    if date_range:
                        start_date = pd.to_datetime(date_range[0])
                        end_date = pd.to_datetime(date_range[1])
                        execution_df_copy = execution_df_copy[
                            (execution_df_copy[influencer_date_col] >= start_date) & 
                            (execution_df_copy[influencer_date_col] <= end_date)
                        ]
                    
                    if not execution_df_copy.empty:
                        all_data.append(execution_df_copy)
                
                # 마케팅 데이터 처리
                if not marketing_df.empty and '유형' in marketing_df.columns:
                    marketing_df_copy = marketing_df.copy()
                    marketing_df_copy[marketing_date_col] = pd.to_datetime(marketing_df_copy[marketing_date_col], errors='coerce')
                    marketing_df_copy = marketing_df_copy.dropna(subset=[marketing_date_col])
                    
                    # 시즌 필터 적용
                    if selected_season != "전체":
                        if selected_season == "24FW":
                            season_start = pd.to_datetime('2024-09-01')
                            season_end = pd.to_datetime('2025-02-28')
                        elif selected_season == "25SS":
                            season_start = pd.to_datetime('2025-03-01')
                            season_end = pd.to_datetime('2025-08-31')
                        elif selected_season == "25FW":
                            season_start = pd.to_datetime('2025-09-01')
                            season_end = pd.to_datetime('2026-02-28')
                        else:
                            season_start = None
                            season_end = None
                        
                        if season_start and season_end:
                            marketing_df_copy = marketing_df_copy[
                                (marketing_df_copy[marketing_date_col] >= season_start) & 
                                (marketing_df_copy[marketing_date_col] <= season_end)
                            ]
                    
                    # 기간 필터 적용
                    if date_range:
                        start_date = pd.to_datetime(date_range[0])
                        end_date = pd.to_datetime(date_range[1])
                        marketing_df_copy = marketing_df_copy[
                            (marketing_df_copy[marketing_date_col] >= start_date) & 
                            (marketing_df_copy[marketing_date_col] <= end_date)
                        ]
                    
                    if not marketing_df_copy.empty:
                        all_data.append(marketing_df_copy)
                
                if all_data:
                    # 모든 데이터를 합치고 유형별로 노출수 집계
                    combined_data = pd.concat(all_data, ignore_index=True)
                    
                    # 유형별로 노출수 집계
                    exposure_by_type = {}
                    for _, row in combined_data.iterrows():
                        date_val = row.get(influencer_date_col) if influencer_date_col in row else row.get(marketing_date_col)
                        exposure_val = row.get(influencer_exposure_col) if influencer_exposure_col in row else row.get(marketing_exposure_col)
                        type_val = row.get('유형', '기타')
                        
                        if pd.notna(date_val) and pd.notna(exposure_val):
                            date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
                            if date_str not in exposure_by_type:
                                exposure_by_type[date_str] = {}
                            if type_val not in exposure_by_type[date_str]:
                                exposure_by_type[date_str][type_val] = 0
                            exposure_by_type[date_str][type_val] += float(exposure_val)
                    
                    # 유형별 노출수 데이터프레임 생성
                    type_dataframes = {}
                    for date_str, types in exposure_by_type.items():
                        for type_name, exposure_val in types.items():
                            if type_name not in type_dataframes:
                                type_dataframes[type_name] = []
                            type_dataframes[type_name].append({
                                'DT': pd.to_datetime(date_str),
                                f'{type_name}_노출수': exposure_val
                            })
                    
                    # 각 유형별 데이터프레임 생성
                    daily_exposure = None
                    for type_name, data_list in type_dataframes.items():
                        type_df = pd.DataFrame(data_list)
                        if daily_exposure is None:
                            daily_exposure = type_df
                        else:
                            daily_exposure = pd.merge(daily_exposure, type_df, on='DT', how='outer').fillna(0)
                else:
                    st.warning("노출수 데이터를 찾을 수 없습니다.")
                    return
            else:
                st.warning("노출수 데이터를 찾을 수 없습니다.")
                return
            
            # 필터 적용된 매출 데이터 준비
            filtered_sales_df = sales_df.copy()
            
            # 브랜드 필터 적용
            if selected_brand != "전체":
                brand_code = brand_mapping.get(selected_brand, selected_brand)
                if 'BRD_CD' in filtered_sales_df.columns:
                    filtered_sales_df = filtered_sales_df[filtered_sales_df['BRD_CD'] == brand_code]
            
            # 아이템 필터 적용
            if selected_item != "전체":
                filtered_sales_df = filtered_sales_df[filtered_sales_df['ITEM'] == selected_item]
            
            # 시즌 필터 적용
            if selected_season != "전체":
                if selected_season == "24FW":
                    season_start = pd.to_datetime('2024-09-01')
                    season_end = pd.to_datetime('2025-02-28')
                elif selected_season == "25SS":
                    season_start = pd.to_datetime('2025-03-01')
                    season_end = pd.to_datetime('2025-08-31')
                elif selected_season == "25FW":
                    season_start = pd.to_datetime('2025-09-01')
                    season_end = pd.to_datetime('2026-02-28')
                else:
                    season_start = None
                    season_end = None
                
                if season_start and season_end:
                    filtered_sales_df = filtered_sales_df[
                        (filtered_sales_df['DT'] >= season_start) & 
                        (filtered_sales_df['DT'] <= season_end)
                    ]
            
            # 기간 필터 적용 (시즌 필터와 함께 적용)
            if date_range:
                start_date = pd.to_datetime(date_range[0])
                end_date = pd.to_datetime(date_range[1])
                filtered_sales_df = filtered_sales_df[
                    (filtered_sales_df['DT'] >= start_date) & 
                    (filtered_sales_df['DT'] <= end_date)
                ]
            
            # 필터링된 매출 데이터 일자별 집계
            if not filtered_sales_df.empty:
                daily_sales = filtered_sales_df.groupby('DT').agg({
                    'SALE_AMT_TY': 'sum'
                }).reset_index()
                daily_sales['DT'] = pd.to_datetime(daily_sales['DT'])
            else:
                daily_sales = pd.DataFrame(columns=['DT', 'SALE_AMT_TY'])
            
            # 집행 데이터도 브랜드 필터 적용
            filtered_execution_df = execution_df_copy.copy()
            if selected_brand != "전체":
                filtered_execution_df = filtered_execution_df[filtered_execution_df['브랜드'] == selected_brand]
            
            # 시즌 필터 적용 (집행 데이터)
            if selected_season != "전체":
                if selected_season == "24FW":
                    season_start = pd.to_datetime('2024-09-01')
                    season_end = pd.to_datetime('2025-02-28')
                elif selected_season == "25SS":
                    season_start = pd.to_datetime('2025-03-01')
                    season_end = pd.to_datetime('2025-08-31')
                elif selected_season == "25FW":
                    season_start = pd.to_datetime('2025-09-01')
                    season_end = pd.to_datetime('2026-02-28')
                else:
                    season_start = None
                    season_end = None
                
                if season_start and season_end:
                    filtered_execution_df = filtered_execution_df[
                        (filtered_execution_df[date_col] >= season_start) & 
                        (filtered_execution_df[date_col] <= season_end)
                    ]
            
            # 기간 필터 적용 (집행 데이터)
            if date_range:
                filtered_execution_df = filtered_execution_df[
                    (filtered_execution_df[date_col] >= start_date) & 
                    (filtered_execution_df[date_col] <= end_date)
                ]
            
            # 매출 데이터와 노출량 데이터 병합
            combined_df = pd.merge(daily_sales, daily_exposure, on='DT', how='outer').fillna(0)
            combined_df = combined_df.sort_values('DT')
            
            if not combined_df.empty:
                # 데이터 유효성 검사
                if len(combined_df) < 2:
                    st.info("차트를 그리기에는 데이터가 부족합니다. (최소 2개 이상의 데이터 포인트 필요)")
                    return
                
                # NaN 값 처리
                combined_df = combined_df.fillna(0)
                
                
                
                # 전체 아이템 사용 (브랜드 필터가 적용된 데이터에서)
                if not filtered_sales_df.empty and 'ITEM' in filtered_sales_df.columns:
                    available_items = filtered_sales_df['ITEM'].unique()
                    selected_items = list(available_items)
                else:
                    selected_items = []
                    # 브랜드 필터로 인해 데이터가 없는 경우
                    st.warning(f"선택된 브랜드 '{selected_brand}'에 대한 데이터가 없습니다.")
                    return
                
                if selected_items:
                    # 선택된 아이템으로 매출 데이터 필터링 (브랜드 필터가 이미 적용된 데이터 사용)
                    filtered_sales_by_item = filtered_sales_df[filtered_sales_df['ITEM'].isin(selected_items)]
                    
                    if not filtered_sales_by_item.empty:
                        # 선택된 아이템의 일자별 매출 데이터 집계
                        daily_sales_by_item = filtered_sales_by_item.groupby('DT').agg({
                            'SALE_AMT_TY': 'sum',
                            'SALE_AMT_LY': 'sum'  # 전년 데이터도 포함
                        }).reset_index()
                        daily_sales_by_item['DT'] = pd.to_datetime(daily_sales_by_item['DT'])
                        
                        # 전년 데이터 처리 (YoY 비교용)
                        daily_sales_ly = None
                        if 'SALE_AMT_LY' in daily_sales_by_item.columns and not daily_sales_by_item['SALE_AMT_LY'].isna().all():
                            # 전년 데이터가 있는 경우 - 현재 날짜에 전년 데이터를 매칭
                            daily_sales_ly = daily_sales_by_item[['DT', 'SALE_AMT_LY']].copy()
                            # 전년 데이터는 현재 날짜에 그대로 표시 (1년 전 데이터를 현재 날짜에 표시)
                            # 날짜는 그대로 두고 전년 매출액만 사용
                        else:
                            st.warning("전년 데이터(SALE_AMT_LY)가 없습니다.")
                        
                        # 기간 필터링 제거 (전체 데이터 사용)
                        filtered_sales_by_item_period = daily_sales_by_item.copy()
                        filtered_exposure_period = daily_exposure.copy()
                        
                        # 매출 데이터와 노출량 데이터 병합 (선택된 아이템 기준)
                        combined_df_item = pd.merge(filtered_sales_by_item_period, filtered_exposure_period, on='DT', how='outer').fillna(0)
                        combined_df_item = combined_df_item.sort_values('DT')
                        
                        if not combined_df_item.empty:
                            # NaN 값 처리
                            combined_df_item = combined_df_item.fillna(0)
                            
                            # 꺾은선그래프 생성
                            try:
                                fig_trend = go.Figure()
                                
                                # 매출액 라인 (좌측 Y축)
                                fig_trend.add_trace(go.Scatter(
                                    x=combined_df_item['DT'],
                                    y=combined_df_item['SALE_AMT_TY'],
                                    mode='lines+markers',
                                    name='당해매출액',
                                    line=dict(color='blue', width=3),
                                    yaxis='y1',
                                    zorder=3  # 막대그래프보다 앞에 표시
                                ))
                                
                                # 전년 매출액 라인 (YoY 비교용)
                                if daily_sales_ly is not None and not daily_sales_ly.empty:
                                    fig_trend.add_trace(go.Scatter(
                                        x=daily_sales_ly['DT'],
                                        y=daily_sales_ly['SALE_AMT_LY'],
                                        mode='lines+markers',
                                        name='전년매출액',
                                        line=dict(color='gray', width=2, dash='dash'),
                                        yaxis='y1',
                                        zorder=2  # 막대그래프보다 앞에 표시
                                    ))
                                
                                # 유형별 노출수 스택형 막대그래프 (우측 Y축)
                                # 유형별 색상 매핑
                                type_colors = {
                                    '인플루언서': '#1f77b4',  # 파란색
                                    '마케팅': '#ff7f0e',      # 주황색
                                    '매체SNS': '#2ca02c',    # 초록색
                                    'SEO': '#9467bd',        # 보라색
                                    '자사IG': '#8c564b',     # 갈색
                                    '셀범': '#e377c2',       # 분홍색
                                    '기타': '#d62728'        # 빨간색
                                }
                                
                                # 유형별 노출수 컬럼 찾기
                                exposure_columns = [col for col in combined_df_item.columns if col.endswith('_노출수')]
                                
                                if exposure_columns:
                                    # 스택형 막대그래프를 위해 각 유형별로 별도 trace 생성
                                    for col in exposure_columns:
                                        type_name = col.replace('_노출수', '')
                                        color = type_colors.get(type_name, '#808080')  # 기본 회색
                                        
                                        fig_trend.add_trace(go.Bar(
                                            x=combined_df_item['DT'],
                                            y=combined_df_item[col],
                                            name=f'{type_name}',
                                            marker=dict(color=color, opacity=1.0),
                                            yaxis='y2',
                                            zorder=1  # 막대그래프는 뒤에 표시
                                        ))
                                else:
                                    # 기존 노출수 컬럼이 있으면 사용 (유형별 데이터가 없는 경우)
                                    if '노출수' in combined_df_item.columns:
                                        fig_trend.add_trace(go.Bar(
                                            x=combined_df_item['DT'],
                                            y=combined_df_item['노출수'],
                                            name='노출수',
                                            marker=dict(color='gray', opacity=0.7),
                                            yaxis='y2',
                                            zorder=1  # 막대그래프는 뒤에 표시
                                        ))
                                
                                # X축 범위 설정 (실제 데이터가 있는 기간만 표시)
                                x_range = None
                                if selected_season != "전체":
                                    # 시즌 선택 시 실제 데이터 범위로 X축 설정
                                    if not combined_df_item.empty:
                                        # 실제 데이터가 있는 날짜 범위 계산
                                        data_dates = pd.to_datetime(combined_df_item['DT'])
                                        min_data_date = data_dates.min()
                                        max_data_date = data_dates.max()
                                        
                                        # 시즌 범위와 데이터 범위의 교집합
                                        if selected_season == "24FW":
                                            season_start = pd.to_datetime('2024-09-01')
                                            season_end = pd.to_datetime('2025-02-28')
                                        elif selected_season == "25SS":
                                            season_start = pd.to_datetime('2025-03-01')
                                            season_end = pd.to_datetime('2025-08-31')
                                        elif selected_season == "25FW":
                                            season_start = pd.to_datetime('2025-09-01')
                                            season_end = pd.to_datetime('2026-02-28')
                                        else:
                                            season_start = min_data_date
                                            season_end = max_data_date
                                        
                                        # 실제 데이터 범위와 시즌 범위의 교집합
                                        actual_start = max(min_data_date, season_start)
                                        actual_end = min(max_data_date, season_end)
                                        
                                        # 데이터가 있는 경우에만 X축 범위 설정
                                        if actual_start <= actual_end:
                                            x_range = [actual_start.strftime('%Y-%m-%d'), actual_end.strftime('%Y-%m-%d')]
                                elif date_range:
                                    # 기간 선택 시 선택된 기간으로 X축 설정
                                    x_range = [date_range[0], date_range[1]]
                                
                                # 레이아웃 설정
                                fig_trend.update_layout(
                                    title="일자별 노출수 및 매출액 트렌드",
                                    xaxis_title="날짜",
                                    barmode='stack',  # 막대그래프 스택 모드
                                    xaxis=dict(
                                        range=x_range,  # X축 범위를 선택된 기간으로 제한
                                        type='date',
                                        showgrid=False  # X축 눈금선 제거
                                    ),
                                    yaxis=dict(
                                        title=dict(
                                            text="매출액",
                                            font=dict(color='blue')
                                        ),
                                        tickfont=dict(color='blue'),
                                        showgrid=False  # Y축 눈금선 제거
                                    ),
                                    yaxis2=dict(
                                        title=dict(
                                            text="노출수",
                                            font=dict(color='red')
                                        ),
                                        tickfont=dict(color='red'),
                                        overlaying='y',
                                        side='right',
                                        showgrid=False  # Y2축 눈금선 제거
                                    ),
                                    hovermode='x unified',
                                    legend=dict(x=0.01, y=0.99, bgcolor='rgba(255,255,255,0.8)'),
                                    height=500
                                )
                                
                                # 비용 트렌드 추가 (매출 연계 분석에 속함)
                                
                                # 노출수 트렌드와 동일한 로직으로 비용 트렌드 생성
                                cost_col = None
                                cost_date_col = None
                                
                                # 비용 컬럼 찾기
                                for col in execution_df.columns:
                                    if '전체비용' in col:
                                        cost_col = col
                                    if '업로드일' in col or '날짜' in col or 'DT' in col:
                                        cost_date_col = col
                                
                                if cost_col and cost_date_col:
                                    # 집행 데이터 일자별 비용 집계 (노출수 트렌드와 동일한 로직)
                                    execution_df_cost = execution_df.copy()
                                    
                                    # 필터 적용 (노출수 트렌드와 동일)
                                    # 브랜드 필터 적용
                                    if selected_brand != "전체":
                                        if '브랜드' in execution_df_cost.columns:
                                            execution_df_cost = execution_df_cost[execution_df_cost['브랜드'] == selected_brand]
                                    
                                    # 아이템 필터 적용
                                    if selected_item != "전체":
                                        if '아이템' in execution_df_cost.columns:
                                            # 아이템 컬럼에서 쉼표로 구분된 값들을 처리
                                            execution_df_cost = execution_df_cost[
                                                execution_df_cost['아이템'].str.contains(selected_item, na=False)
                                            ]
                                    
                                    # 시즌 필터 적용
                                    if selected_season != "전체":
                                        if '시즌' in execution_df_cost.columns:
                                            execution_df_cost = execution_df_cost[execution_df_cost['시즌'] == selected_season]
                                    
                                    # 날짜 변환을 안전하게 처리 (노출수 트렌드와 동일)
                                    try:
                                        execution_df_cost[cost_date_col] = pd.to_datetime(execution_df_cost[cost_date_col], errors='coerce')
                                        # 변환 실패한 행 제거
                                        execution_df_cost = execution_df_cost.dropna(subset=[cost_date_col])
                                        
                                        # 날짜 범위 필터 적용
                                        if date_range:
                                            execution_df_cost = execution_df_cost[
                                                (execution_df_cost[cost_date_col] >= pd.to_datetime(date_range[0])) &
                                                (execution_df_cost[cost_date_col] <= pd.to_datetime(date_range[1]))
                                            ]
                                        
                                        if execution_df_cost.empty:
                                            st.warning("필터링 후 비용 데이터가 없습니다.")
                                            # 비용 데이터가 없어도 매출액 그래프는 표시
                                            daily_cost = None
                                            
                                            # 매출 데이터 준비 (비용 데이터가 없을 때도 매출액 그래프 표시)
                                            filtered_sales_df_cost = sales_df.copy()
                                            
                                            # 브랜드 필터 적용
                                            if selected_brand != "전체":
                                                brand_code = brand_mapping.get(selected_brand, selected_brand)
                                                if 'BRD_CD' in filtered_sales_df_cost.columns:
                                                    filtered_sales_df_cost = filtered_sales_df_cost[filtered_sales_df_cost['BRD_CD'] == brand_code]
                                            
                                            # 아이템 필터 적용
                                            if selected_item != "전체":
                                                filtered_sales_df_cost = filtered_sales_df_cost[filtered_sales_df_cost['ITEM'] == selected_item]
                                            
                                            # 시즌 필터 적용
                                            if selected_season != "전체":
                                                if '시즌' in filtered_sales_df_cost.columns:
                                                    filtered_sales_df_cost = filtered_sales_df_cost[filtered_sales_df_cost['시즌'] == selected_season]
                                            
                                            # 날짜 필터 적용
                                            if date_range:
                                                filtered_sales_df_cost['DT'] = pd.to_datetime(filtered_sales_df_cost['DT'])
                                                filtered_sales_df_cost = filtered_sales_df_cost[
                                                    (filtered_sales_df_cost['DT'] >= pd.to_datetime(date_range[0])) &
                                                    (filtered_sales_df_cost['DT'] <= pd.to_datetime(date_range[1]))
                                                ]
                                            
                                            # 전체 매출 데이터 일자별 집계
                                            daily_sales_cost = filtered_sales_df_cost.groupby('DT').agg({
                                                'SALE_AMT_TY': 'sum',
                                                'SALE_AMT_LY': 'sum'  # 전년 데이터도 포함
                                            }).reset_index()
                                            daily_sales_cost['DT'] = pd.to_datetime(daily_sales_cost['DT'])
                                            
                                            # 비용 데이터가 없을 때도 매출액만으로 그래프 표시
                                            if not daily_sales_cost.empty:
                                                # 매출 데이터만으로 그래프 생성
                                                fig_cost = go.Figure()
                                                
                                                # 매출액 라인 (좌측 Y축)
                                                fig_cost.add_trace(go.Scatter(
                                                    x=daily_sales_cost['DT'],
                                                    y=daily_sales_cost['SALE_AMT_TY'],
                                                    mode='lines+markers',
                                                    name='당해매출액',
                                                    line=dict(color='blue', width=3),
                                                    yaxis='y1',
                                                    zorder=3
                                                ))
                                                
                                                # 전년 매출액 라인 (YoY 비교용)
                                                if 'SALE_AMT_LY' in daily_sales_cost.columns and not daily_sales_cost['SALE_AMT_LY'].isna().all():
                                                    fig_cost.add_trace(go.Scatter(
                                                        x=daily_sales_cost['DT'],
                                                        y=daily_sales_cost['SALE_AMT_LY'],
                                                        mode='lines+markers',
                                                        name='전년매출액',
                                                        line=dict(color='gray', width=2, dash='dash'),
                                                        yaxis='y1',
                                                        zorder=2
                                                    ))
                                                
                                                # 레이아웃 설정
                                                fig_cost.update_layout(
                                                    title="일자별 비용 및 매출액 트렌드",
                                                    xaxis_title="날짜",
                                                    xaxis=dict(
                                                        type='date',
                                                        showgrid=False
                                                    ),
                                                    yaxis=dict(
                                                        title=dict(
                                                            text="매출액",
                                                            font=dict(color='blue')
                                                        ),
                                                        tickfont=dict(color='blue'),
                                                        showgrid=False
                                                    ),
                                                    hovermode='x unified',
                                                    legend=dict(x=0.01, y=0.99, bgcolor='rgba(255,255,255,0.8)'),
                                                    height=500
                                                )
                                                
                                                st.plotly_chart(fig_cost, use_container_width=True)
                                            return
                                        
                                        # 마케팅 데이터도 포함하여 유형별 비용 데이터 처리
                                        all_cost_data = []
                                        
                                        # 인플루언서 데이터 처리
                                        if '유형' in execution_df_cost.columns:
                                            all_cost_data.append(execution_df_cost)
                                        
                                        # 마케팅 데이터 처리
                                        marketing_df = load_marketing_data()
                                        if not marketing_df.empty and '유형' in marketing_df.columns:
                                            # 마케팅 데이터 필터 적용
                                            marketing_df_cost = marketing_df.copy()
                                            
                                            # 브랜드 필터 적용
                                            if selected_brand != "전체":
                                                if '브랜드' in marketing_df_cost.columns:
                                                    marketing_df_cost = marketing_df_cost[marketing_df_cost['브랜드'] == selected_brand]
                                            
                                            # 아이템 필터 적용
                                            if selected_item != "전체":
                                                if '아이템' in marketing_df_cost.columns:
                                                    marketing_df_cost = marketing_df_cost[
                                                        marketing_df_cost['아이템'].str.contains(selected_item, na=False)
                                                    ]
                                            
                                            # 시즌 필터 적용
                                            if selected_season != "전체":
                                                if '시즌' in marketing_df_cost.columns:
                                                    marketing_df_cost = marketing_df_cost[marketing_df_cost['시즌'] == selected_season]
                                            
                                            # 마케팅 데이터의 날짜 및 비용 컬럼 찾기
                                            marketing_cost_col = None
                                            marketing_cost_date_col = None
                                            for col in marketing_df_cost.columns:
                                                if '비용' in col or '전체비용' in col:
                                                    marketing_cost_col = col
                                                if '업로드일' in col or '날짜' in col or 'DT' in col:
                                                    marketing_cost_date_col = col
                                            
                                            if marketing_cost_col and marketing_cost_date_col:
                                                # 마케팅 데이터 날짜 변환
                                                marketing_df_cost[marketing_cost_date_col] = pd.to_datetime(marketing_df_cost[marketing_cost_date_col], errors='coerce')
                                                marketing_df_cost = marketing_df_cost.dropna(subset=[marketing_cost_date_col])
                                                
                                                # 날짜 범위 필터 적용
                                                if date_range:
                                                    marketing_df_cost = marketing_df_cost[
                                                        (marketing_df_cost[marketing_cost_date_col] >= pd.to_datetime(date_range[0])) &
                                                        (marketing_df_cost[marketing_cost_date_col] <= pd.to_datetime(date_range[1]))
                                                    ]
                                                
                                                if not marketing_df_cost.empty:
                                                    all_cost_data.append(marketing_df_cost)
                                        
                                        if all_cost_data:
                                            # 모든 데이터를 합치고 유형별로 비용 집계
                                            combined_cost_data = pd.concat(all_cost_data, ignore_index=True)
                                            
                                            # 유형별로 비용 집계
                                            cost_by_type = {}
                                            for _, row in combined_cost_data.iterrows():
                                                # 날짜 값 찾기
                                                date_val = None
                                                for col in [cost_date_col, '업로드일', '날짜', 'DT']:
                                                    if col in row and pd.notna(row[col]):
                                                        date_val = row[col]
                                                        break
                                                
                                                # 비용 값 찾기
                                                cost_val = None
                                                for col in [cost_col, '비용', '전체비용']:
                                                    if col in row and pd.notna(row[col]):
                                                        cost_val = row[col]
                                                        break
                                                
                                                type_val = row.get('유형', '기타')
                                                
                                                if pd.notna(date_val) and pd.notna(cost_val):
                                                    # 비용 값이 유효한 숫자인지 확인
                                                    try:
                                                        cost_float = float(cost_val)
                                                        if cost_float >= 0:  # 음수 제거
                                                            date_str = pd.to_datetime(date_val).strftime('%Y-%m-%d')
                                                            if date_str not in cost_by_type:
                                                                cost_by_type[date_str] = {}
                                                            if type_val not in cost_by_type[date_str]:
                                                                cost_by_type[date_str][type_val] = 0
                                                            cost_by_type[date_str][type_val] += cost_float
                                                    except (ValueError, TypeError):
                                                        # 유효하지 않은 숫자는 무시
                                                        continue
                                            
                                            # 유형별 비용 데이터프레임 생성
                                            type_cost_dataframes = {}
                                            for date_str, types in cost_by_type.items():
                                                for type_name, cost_val in types.items():
                                                    if type_name not in type_cost_dataframes:
                                                        type_cost_dataframes[type_name] = []
                                                    type_cost_dataframes[type_name].append({
                                                        'DT': pd.to_datetime(date_str),
                                                        f'{type_name}_비용': cost_val
                                                    })
                                            
                                            # 각 유형별 데이터프레임 생성
                                            daily_cost = None
                                            for type_name, data_list in type_cost_dataframes.items():
                                                type_df = pd.DataFrame(data_list)
                                                if daily_cost is None:
                                                    daily_cost = type_df
                                                else:
                                                    daily_cost = pd.merge(daily_cost, type_df, on='DT', how='outer').fillna(0)
                                        else:
                                            # 유형별 데이터가 없는 경우 기본 처리
                                            daily_cost = execution_df_cost.groupby(cost_date_col)[cost_col].sum().reset_index()
                                            daily_cost.columns = ['DT', '비용']
                                    except Exception as e:
                                        st.error(f"날짜 변환 중 오류가 발생했습니다: {e}")
                                        st.info("집행 데이터의 날짜 형식을 확인해주세요.")
                                        return
                                    
                                    # 필터 적용된 매출 데이터 준비
                                    filtered_sales_df_cost = sales_df.copy()
                                    
                                    # 브랜드 필터 적용
                                    if selected_brand != "전체":
                                        brand_code = brand_mapping.get(selected_brand, selected_brand)
                                        if 'BRD_CD' in filtered_sales_df_cost.columns:
                                            filtered_sales_df_cost = filtered_sales_df_cost[filtered_sales_df_cost['BRD_CD'] == brand_code]
                                    
                                    # 아이템 필터 적용
                                    if selected_item != "전체":
                                        filtered_sales_df_cost = filtered_sales_df_cost[filtered_sales_df_cost['ITEM'] == selected_item]
                                    
                                    # 시즌 필터 적용
                                    if selected_season != "전체":
                                        if '시즌' in filtered_sales_df_cost.columns:
                                            filtered_sales_df_cost = filtered_sales_df_cost[filtered_sales_df_cost['시즌'] == selected_season]
                                    
                                    # 날짜 필터 적용
                                    if date_range:
                                        filtered_sales_df_cost['DT'] = pd.to_datetime(filtered_sales_df_cost['DT'])
                                        filtered_sales_df_cost = filtered_sales_df_cost[
                                            (filtered_sales_df_cost['DT'] >= pd.to_datetime(date_range[0])) &
                                            (filtered_sales_df_cost['DT'] <= pd.to_datetime(date_range[1]))
                                        ]
                                    
                                    # 전체 매출 데이터 일자별 집계
                                    daily_sales_cost = filtered_sales_df_cost.groupby('DT').agg({
                                        'SALE_AMT_TY': 'sum',
                                        'SALE_AMT_LY': 'sum'  # 전년 데이터도 포함
                                    }).reset_index()
                                    daily_sales_cost['DT'] = pd.to_datetime(daily_sales_cost['DT'])
                                    
                                    # 매출 데이터와 비용 데이터 병합
                                    combined_df_cost = pd.merge(daily_sales_cost, daily_cost, on='DT', how='outer').fillna(0)
                                    combined_df_cost = combined_df_cost.sort_values('DT')
                                    
                                    if not combined_df_cost.empty:
                                        # 데이터 유효성 검사
                                        if len(combined_df_cost) < 2:
                                            st.info("차트를 그리기에는 데이터가 부족합니다. (최소 2개 이상의 데이터 포인트 필요)")
                                        else:
                                            # NaN 값 처리
                                            combined_df_cost = combined_df_cost.fillna(0)
                                            
                                            # 꺾은선그래프 생성 (노출수 트렌드와 동일한 구조)
                                            try:
                                                fig_cost = go.Figure()
                                                
                                                # 매출액 라인 (좌측 Y축)
                                                fig_cost.add_trace(go.Scatter(
                                                    x=combined_df_cost['DT'],
                                                    y=combined_df_cost['SALE_AMT_TY'],
                                                    mode='lines+markers',
                                                    name='당해매출액',
                                                    line=dict(color='blue', width=3),
                                                    yaxis='y1',
                                                    zorder=3  # 막대그래프보다 앞에 표시
                                                ))
                                                
                                                # 전년 매출액 라인 (YoY 비교용)
                                                if 'SALE_AMT_LY' in combined_df_cost.columns and not combined_df_cost['SALE_AMT_LY'].isna().all():
                                                    fig_cost.add_trace(go.Scatter(
                                                        x=combined_df_cost['DT'],
                                                        y=combined_df_cost['SALE_AMT_LY'],
                                                        mode='lines+markers',
                                                        name='전년매출액',
                                                        line=dict(color='gray', width=2, dash='dash'),
                                                        yaxis='y1',
                                                        zorder=2  # 막대그래프보다 앞에 표시
                                                    ))
                                                
                                                # 유형별 비용 스택형 막대그래프 (우측 Y축)
                                                # 유형별 색상 매핑 (노출수 트렌드와 동일)
                                                type_colors = {
                                                    '인플루언서': '#1f77b4',  # 파란색
                                                    '마케팅': '#ff7f0e',      # 주황색
                                                    '매체SNS': '#2ca02c',    # 초록색
                                                    'SEO': '#9467bd',        # 보라색
                                                    '자사IG': '#8c564b',     # 갈색
                                                    '셀범': '#e377c2',       # 분홍색
                                                    '기타': '#d62728'        # 빨간색
                                                }
                                                
                                                # 유형별 비용 컬럼 찾기
                                                cost_columns = [col for col in combined_df_cost.columns if col.endswith('_비용')]
                                                
                                                if cost_columns:
                                                    # 스택형 막대그래프를 위해 각 유형별로 별도 trace 생성
                                                    for i, col in enumerate(cost_columns):
                                                        type_name = col.replace('_비용', '')
                                                        color = type_colors.get(type_name, '#808080')  # 기본 회색
                                                        
                                                        # 각 유형별 호버 템플릿
                                                        hover_template = f'<b>{type_name}: %{{y:,.0f}}원</b><extra></extra>'
                                                        
                                                        fig_cost.add_trace(go.Bar(
                                                            x=combined_df_cost['DT'],
                                                            y=combined_df_cost[col],
                                                            name=f'{type_name}',
                                                            marker=dict(color=color, opacity=1.0),
                                                            yaxis='y2',
                                                            zorder=1,  # 막대그래프는 뒤에 표시
                                                            hovertemplate=hover_template
                                                        ))
                                                else:
                                                    # 기존 비용 컬럼이 있으면 사용 (유형별 데이터가 없는 경우)
                                                    if '비용' in combined_df_cost.columns:
                                                        fig_cost.add_trace(go.Bar(
                                                            x=combined_df_cost['DT'],
                                                            y=combined_df_cost['비용'],
                                                            name='비용',
                                                            marker=dict(color='red', opacity=0.7),
                                                            yaxis='y2',
                                                            zorder=1,  # 막대그래프는 뒤에 표시
                                                            hovertemplate='<b>비용: %{y:,.0f}원</b><extra></extra>'
                                                        ))
                                                
                                                # 레이아웃 설정 (노출수 트렌드와 동일)
                                                fig_cost.update_layout(
                                                    title="일자별 비용 및 매출액 트렌드",
                                                    xaxis_title="날짜",
                                                    barmode='stack',  # 막대그래프 스택 모드
                                                    xaxis=dict(
                                                        type='date',
                                                        showgrid=False  # X축 눈금선 제거
                                                    ),
                                                    yaxis=dict(
                                                        title=dict(
                                                            text="매출액",
                                                            font=dict(color='blue')
                                                        ),
                                                        tickfont=dict(color='blue'),
                                                        showgrid=False  # Y축 눈금선 제거
                                                    ),
                                                    yaxis2=dict(
                                                        title=dict(
                                                            text="비용",
                                                            font=dict(color='red')
                                                        ),
                                                        tickfont=dict(color='red'),
                                                        overlaying='y',
                                                        side='right',
                                                        showgrid=False  # Y2축 눈금선 제거
                                                    ),
                                                    hovermode='x unified',
                                                    legend=dict(x=0.01, y=0.99, bgcolor='rgba(255,255,255,0.8)'),
                                                    height=500
                                                )
                                                
                                                st.plotly_chart(fig_cost, use_container_width=True)
                                                
                                            except Exception as e:
                                                st.error(f"차트 생성 중 오류가 발생했습니다: {e}")
                                                st.info("데이터 형식을 확인해주세요.")
                                    else:
                                        st.info("매출과 비용 데이터를 연계할 수 없습니다.")
                                else:
                                    st.info("비용 데이터를 찾을 수 없습니다.")
                                
                            except Exception as e:
                                st.error(f"차트 생성 중 오류가 발생했습니다: {e}")
                                st.info("데이터 형식을 확인해주세요.")
                                return
                            
                            # 노출수 트렌드 그래프 표시
                            st.plotly_chart(fig_trend, use_container_width=True)
                            
                            # 상관계수 계산 (순서 변경)
                            if len(combined_df_item) > 1:
                                # 비용-매출액 상관계수 계산 (먼저 표시)
                                if 'daily_cost' in locals() and daily_cost is not None and not daily_cost.empty:
                                    # 비용 데이터와 매출 데이터 병합
                                    cost_sales_df = pd.merge(combined_df_item[['DT', 'SALE_AMT_TY']], daily_cost, on='DT', how='inner')
                                    
                                    if len(cost_sales_df) > 1:
                                        # 유형별 비용 컬럼이 있는지 확인
                                        cost_columns = [col for col in cost_sales_df.columns if col.endswith('_비용')]
                                        
                                        if cost_columns:
                                            # 유형별 비용의 합계 계산
                                            total_cost = cost_sales_df[cost_columns].sum(axis=1)
                                            cost_correlation = cost_sales_df['SALE_AMT_TY'].corr(total_cost)
                                            st.metric("💰 비용-매출액 상관계수", f"{cost_correlation:.3f}")
                                        elif '비용' in cost_sales_df.columns:
                                            # 기존 비용 컬럼이 있는 경우
                                            cost_correlation = cost_sales_df['SALE_AMT_TY'].corr(cost_sales_df['비용'])
                                            st.metric("💰 비용-매출액 상관계수", f"{cost_correlation:.3f}")
                                        else:
                                            st.warning("비용 데이터를 찾을 수 없습니다.")
                                    else:
                                        st.warning("비용과 매출 데이터를 연계할 수 없습니다.")
                                else:
                                    st.warning("비용 데이터가 없습니다.")
                                
                                # 노출수-매출액 상관계수 계산 (나중에 표시)
                                # 유형별 노출수 컬럼이 있는지 확인
                                exposure_columns = [col for col in combined_df_item.columns if col.endswith('_노출수')]
                                
                                if exposure_columns:
                                    # 유형별 노출수의 합계 계산
                                    total_exposure = combined_df_item[exposure_columns].sum(axis=1)
                                    correlation = combined_df_item['SALE_AMT_TY'].corr(total_exposure)
                                    st.metric("📊 노출수-매출액 상관계수", f"{correlation:.3f}")
                                elif '노출수' in combined_df_item.columns:
                                    # 기존 노출수 컬럼이 있는 경우
                                    correlation = combined_df_item['SALE_AMT_TY'].corr(combined_df_item['노출수'])
                                    st.metric("📊 노출수-매출액 상관계수", f"{correlation:.3f}")
                                else:
                                    st.warning("노출수 데이터를 찾을 수 없습니다.")
                            
                            # YoY 성장률 계산 (전년 데이터가 있는 경우)
                            if 'SALE_AMT_LY' in combined_df_item.columns and not combined_df_item['SALE_AMT_LY'].isna().all():
                                # 필터링된 데이터에서 직접 YoY 성장률 계산
                                valid_data = combined_df_item[
                                    (combined_df_item['SALE_AMT_TY'] > 0) & 
                                    (combined_df_item['SALE_AMT_LY'] > 0)
                                ]
                                
                                if not valid_data.empty:
                                    # YoY 성장률 계산
                                    yoy_growth = ((valid_data['SALE_AMT_TY'] - valid_data['SALE_AMT_LY']) / valid_data['SALE_AMT_LY'] * 100).mean()
                                    st.metric("📈 YoY 매출 비교", f"{yoy_growth:.1f}%")
                                else:
                                    st.warning("YoY 성장률 계산을 위한 유효한 데이터가 없습니다.")
                            else:
                                st.warning("전년 데이터(SALE_AMT_LY)가 없어 YoY 성장률을 계산할 수 없습니다.")
                            
                            
                            # 상세 데이터 테이블
                            st.markdown("#### 📋 일자별 상세 데이터")
                            display_df = combined_df_item.copy()
                            display_df['DT'] = display_df['DT'].dt.strftime('%Y-%m-%d')
                            
                            # 유형별 비용 데이터 추가
                            if 'daily_cost' in locals() and daily_cost is not None:
                                # DT 컬럼 데이터 타입 통일 (datetime으로 변환)
                                display_df['DT'] = pd.to_datetime(display_df['DT'], errors='coerce')
                                daily_cost['DT'] = pd.to_datetime(daily_cost['DT'], errors='coerce')
                                
                                # 비용 데이터와 매출 데이터 병합
                                display_df = pd.merge(display_df, daily_cost, on='DT', how='left')
                            
                            # 컬럼명 변경 (유형별 노출수 컬럼 처리)
                            column_mapping = {'DT': '날짜', 'SALE_AMT_TY': '당해매출액'}
                            
                            # 전년 데이터가 있는 경우 추가
                            if daily_sales_ly is not None and not daily_sales_ly.empty:
                                # 전년 데이터와 올해 데이터를 같은 날짜로 매칭하여 표시
                                current_sales = combined_df_item[['DT', 'SALE_AMT_TY']].copy()
                                current_sales['DT'] = current_sales['DT'] - pd.DateOffset(years=1)
                                yoy_comparison = pd.merge(current_sales, daily_sales_ly, on='DT', how='inner')
                                if not yoy_comparison.empty:
                                    # YoY 비교 데이터를 원래 날짜로 복원
                                    yoy_comparison['DT'] = yoy_comparison['DT'] + pd.DateOffset(years=1)
                                    
                                    # DT 컬럼 데이터 타입 통일 (datetime으로 변환)
                                    display_df['DT'] = pd.to_datetime(display_df['DT'], errors='coerce')
                                    yoy_comparison['DT'] = pd.to_datetime(yoy_comparison['DT'], errors='coerce')
                                    
                                    display_df = pd.merge(display_df, yoy_comparison[['DT', 'SALE_AMT_LY']], on='DT', how='left')
                                    column_mapping['SALE_AMT_LY'] = '전년매출액'
                            
                            
                            # 유형별 노출수 컬럼이 있는 경우
                            exposure_columns = [col for col in display_df.columns if col.endswith('_노출수')]
                            if exposure_columns:
                                # 유형별 노출수 컬럼명을 한국어로 변경
                                for col in exposure_columns:
                                    type_name = col.replace('_노출수', '')
                                    column_mapping[col] = f'{type_name} 노출수'
                            elif '노출수' in display_df.columns:
                                column_mapping['노출수'] = '노출수'
                            
                            # 유형별 비용 컬럼이 있는 경우
                            cost_columns = [col for col in display_df.columns if col.endswith('_비용')]
                            if cost_columns:
                                # 유형별 비용 컬럼명을 한국어로 변경
                                for col in cost_columns:
                                    type_name = col.replace('_비용', '')
                                    column_mapping[col] = f'{type_name} 비용'
                            elif '비용' in display_df.columns:
                                column_mapping['비용'] = '비용'
                            
                            # 컬럼명 변경 적용
                            display_df = display_df.rename(columns=column_mapping)
                            display_df = display_df.sort_values('날짜', ascending=False)
                            
                            st.dataframe(display_df, use_container_width=True, hide_index=True)
                        else:
                            st.info("매출과 노출량 데이터를 연계할 수 없습니다.")
                    else:
                        st.info("선택된 아이템에 대한 매출 데이터가 없습니다.")
                else:
                    st.info("아이템을 선택해주세요.")
            else:
                st.error(f"집행 데이터에서 필요한 컬럼을 찾을 수 없습니다. 노출수: {exposure_col}, 날짜: {date_col}")
        else:
            st.info("집행 데이터가 없어 노출량 연계 분석을 수행할 수 없습니다.")
    else:
        st.info("매출 데이터에 아이템 정보가 없습니다.")
    
    # AI 인사이트 및 넥스트 스텝 제안 (필터링된 데이터 기반)
    st.markdown("---")
    st.markdown("## 🤖 AI 인사이트 & 넥스트 스텝")
    
    # 현재 필터링된 데이터로 분석
    try:
        # 필터링된 데이터 확인
        if 'combined_df_item' in locals() and not combined_df_item.empty:
            # 현재 필터 조건 표시
            filter_info = []
            if selected_brand != "전체":
                filter_info.append(f"브랜드: {selected_brand}")
            if selected_item != "전체":
                filter_info.append(f"아이템: {selected_item}")
            if selected_season != "전체":
                filter_info.append(f"시즌: {selected_season}")
            if date_range:
                filter_info.append(f"기간: {date_range[0]} ~ {date_range[1]}")
            
            if filter_info:
                st.markdown(f"**📊 분석 대상**: {', '.join(filter_info)}")
            
            # 필터링된 데이터의 상관관계 분석
            insights = []
            recommendations = []
            
            # 사용 가능한 모든 컬럼 확인 (내부적으로만 사용)
            available_columns = list(combined_df_item.columns)
            
            # 매출액 컬럼 확인
            sales_columns = [col for col in available_columns if 'SALE_AMT' in col]
            
            # 노출수 관련 컬럼 확인
            exposure_columns = [col for col in available_columns if '노출수' in col or '노출' in col]
            
            # 비용 관련 컬럼 확인
            cost_columns = [col for col in available_columns if '비용' in col or 'COST' in col]
            
            # 1. 노출수-매출액 상관관계 분석 (모든 노출수 컬럼)
            if 'SALE_AMT_TY' in combined_df_item.columns:
                for exposure_col in exposure_columns:
                    if exposure_col in combined_df_item.columns:
                        # 유효한 데이터만 추출
                        valid_data = combined_df_item[[exposure_col, 'SALE_AMT_TY']].dropna()
                        if len(valid_data) > 1:
                            correlation = valid_data[exposure_col].corr(valid_data['SALE_AMT_TY'])
                            
                            if not pd.isna(correlation):
                                # 노출수 유형별 한국어 이름 매핑
                                type_names = {
                                    '인플루언서_노출수': '인플루언서 마케팅',
                                    '자사IG_노출수': '자사 인스타그램',
                                    'SEO_노출수': '검색엔진 최적화',
                                    '셀럽_노출수': '셀럽 마케팅',
                                    '매체SNS_노출수': '매체 SNS'
                                }
                                type_name = type_names.get(exposure_col, exposure_col)
                                
                                if correlation > 0.7:
                                    insights.append(f"🎯 **{type_name}이 매출에 강력한 영향을 미치고 있습니다!** 노출이 늘어날수록 매출이 확실히 증가하는 패턴을 보입니다. 이 채널에 더 집중하세요.")
                                elif correlation > 0.3:
                                    insights.append(f"📈 **{type_name}이 매출 증가에 도움이 되고 있습니다.** 어느 정도 효과가 있지만, 더 강한 연관성을 위해 콘텐츠 품질을 높여보세요.")
                                elif correlation > -0.3:
                                    insights.append(f"⚠️ **{type_name}의 매출 기여도가 제한적입니다.** 노출수가 늘어나도 매출에 큰 변화가 없습니다. 타겟팅과 메시지를 재검토해야 합니다.")
                                else:
                                    insights.append(f"📉 **{type_name}이 오히려 매출에 부정적 영향을 미치고 있습니다.** 노출이 늘어날수록 매출이 감소하는 패턴입니다. 즉시 전략을 바꿔야 합니다.")
            
            # 2. 비용-매출액 상관관계 분석 (모든 비용 컬럼)
            if 'SALE_AMT_TY' in combined_df_item.columns:
                for cost_col in cost_columns:
                    if cost_col in combined_df_item.columns:
                        # 유효한 데이터만 추출
                        valid_data = combined_df_item[[cost_col, 'SALE_AMT_TY']].dropna()
                        if len(valid_data) > 1:
                            correlation = valid_data[cost_col].corr(valid_data['SALE_AMT_TY'])
                            
                            if not pd.isna(correlation):
                                if correlation > 0.5:
                                    insights.append(f"💰 **{cost_col} 투자가 매출에 큰 도움이 되고 있습니다!** 비용을 늘릴수록 매출이 확실히 증가하는 패턴입니다. 이 채널에 더 투자하세요.")
                                elif correlation > 0:
                                    insights.append(f"💸 **{cost_col} 투자의 효과가 제한적입니다.** 비용을 늘려도 매출 증가가 미미합니다. ROI를 높이기 위해 전략을 개선하세요.")
                                else:
                                    insights.append(f"⚠️ **{cost_col} 투자가 비효율적입니다.** 비용을 늘릴수록 오히려 매출이 감소하는 패턴입니다. 즉시 투자 전략을 바꿔야 합니다.")
            
            # 3. 노출수-비용 상관관계 분석 (노출수와 비용 간의 관계)
            for exposure_col in exposure_columns:
                for cost_col in cost_columns:
                    if exposure_col in combined_df_item.columns and cost_col in combined_df_item.columns:
                        valid_data = combined_df_item[[exposure_col, cost_col]].dropna()
                        if len(valid_data) > 1:
                            correlation = valid_data[exposure_col].corr(valid_data[cost_col])
                            
                            if not pd.isna(correlation):
                                # 노출수 유형별 한국어 이름 매핑
                                type_names = {
                                    '인플루언서_노출수': '인플루언서 마케팅',
                                    '자사IG_노출수': '자사 인스타그램',
                                    'SEO_노출수': '검색엔진 최적화',
                                    '셀럽_노출수': '셀럽 마케팅',
                                    '매체SNS_노출수': '매체 SNS'
                                }
                                exposure_name = type_names.get(exposure_col, exposure_col)
                                
                                if correlation > 0.7:
                                    insights.append(f"🎯 **{exposure_name}에 투자할수록 노출이 확실히 늘어납니다!** 비용 대비 노출 효과가 매우 좋습니다. 이 채널에 더 집중하세요.")
                                elif correlation > 0.3:
                                    insights.append(f"📊 **{exposure_name} 투자가 노출 증가에 도움이 됩니다.** 어느 정도 효과가 있지만, 더 효율적인 방법을 찾아보세요.")
                                elif correlation > -0.3:
                                    insights.append(f"⚠️ **{exposure_name} 투자의 노출 효과가 제한적입니다.** 비용을 늘려도 노출이 크게 늘지 않습니다. 전략을 재검토하세요.")
                                else:
                                    insights.append(f"📉 **{exposure_name} 투자가 오히려 노출을 줄이고 있습니다.** 비용을 늘릴수록 노출이 감소하는 패턴입니다. 즉시 접근법을 바꿔야 합니다.")
            
            # 3. YoY 성장률 기반 인사이트
            if 'SALE_AMT_LY' in combined_df_item.columns:
                valid_yoy_data = combined_df_item[
                    (combined_df_item['SALE_AMT_TY'] > 0) & 
                    (combined_df_item['SALE_AMT_LY'] > 0)
                ]
                if not valid_yoy_data.empty:
                    yoy_growth = ((valid_yoy_data['SALE_AMT_TY'] - valid_yoy_data['SALE_AMT_LY']) / valid_yoy_data['SALE_AMT_LY'] * 100).mean()
                    
                    if yoy_growth > 20:
                        insights.append(f"🚀 **대박! 전년 대비 {yoy_growth:.1f}% 성장했습니다!** 현재 마케팅 전략이 매우 효과적입니다. 이 성공 요인을 분석해서 더 확장하세요.")
                    elif yoy_growth > 0:
                        insights.append(f"📈 **좋은 성장세입니다! 전년 대비 {yoy_growth:.1f}% 성장했습니다.** 현재 방향이 맞습니다. 더 적극적으로 마케팅을 늘려보세요.")
                    else:
                        insights.append(f"⚠️ **성장이 둔화되고 있습니다. 전년 대비 {yoy_growth:.1f}% 변화입니다.** 현재 전략에 문제가 있을 수 있습니다. 원인을 분석하고 새로운 접근법을 시도해보세요.")
            
            # 4. 유형별 성과 분석 (노출수 기준)
            exposure_columns = [col for col in combined_df_item.columns if col.endswith('_노출수')]
            if exposure_columns:
                type_performance = {}
                for col in exposure_columns:
                    type_name = col.replace('_노출수', '')
                    type_exposure = combined_df_item[col].sum()
                    if type_exposure > 0:
                        type_performance[type_name] = type_exposure
                
                if type_performance:
                    best_type = max(type_performance, key=type_performance.get)
                    worst_type = min(type_performance, key=type_performance.get)
                    
                    # 유형별 한국어 이름 매핑
                    type_names = {
                        '인플루언서': '인플루언서 마케팅',
                        '자사IG': '자사 인스타그램',
                        'SEO': '검색엔진 최적화',
                        '셀럽': '셀럽 마케팅',
                        '매체SNS': '매체 SNS'
                    }
                    best_name = type_names.get(best_type, best_type)
                    worst_name = type_names.get(worst_type, worst_type)
                    
                    insights.append(f"🏆 **{best_name}이 가장 효과적입니다!** {type_performance[best_type]:,.0f}회의 노출을 기록했습니다. 이 채널의 성공 비법을 다른 채널에도 적용해보세요.")
                    
                    if type_performance[best_type] > type_performance[worst_type] * 3:
                        insights.append(f"💡 **{best_name}의 성공 요인을 {worst_name}에 적용하세요.** 성과 차이가 3배 이상 납니다. 성공한 전략을 복사해보세요.")
                    else:
                        insights.append(f"⚖️ **각 채널별로 차별화된 전략이 필요합니다.** 모든 채널이 비슷한 성과를 보이므로, 각각의 특성에 맞는 맞춤형 접근이 필요합니다.")
            
            # 5. 아이템별 분석 추가 (다양한 컬럼명 확인)
            item_column = None
            if 'ITEM' in combined_df_item.columns:
                item_column = 'ITEM'
            elif '아이템' in combined_df_item.columns:
                item_column = '아이템'
            elif 'item' in combined_df_item.columns:
                item_column = 'item'
            
            if item_column:
                # 아이템별 노출수 분석
                if '노출수' in combined_df_item.columns:
                    item_exposure = combined_df_item.groupby(item_column)['노출수'].sum()
                    if not item_exposure.empty and item_exposure.sum() > 0:
                        best_item = item_exposure.idxmax()
                        best_exposure = item_exposure.max()
                        insights.append(f"📦 **{best_item} 아이템이 가장 많은 관심을 받고 있습니다!** {best_exposure:,.0f}회의 노출을 기록했습니다. 이 아이템의 마케팅 전략을 다른 아이템에도 적용해보세요.")

                # 아이템별 매출액 분석
                if 'SALE_AMT_TY' in combined_df_item.columns:
                    item_sales = combined_df_item.groupby(item_column)['SALE_AMT_TY'].sum()
                    if not item_sales.empty and item_sales.sum() > 0:
                        best_sales_item = item_sales.idxmax()
                        best_sales = item_sales.max()
                        insights.append(f"💰 **{best_sales_item} 아이템이 매출의 주력군입니다!** {best_sales:,.0f}원의 매출을 기록했습니다. 이 아이템에 더 집중하세요.")

                # 아이템별 효율성 분석 (노출수 대비 매출액)
                if '노출수' in combined_df_item.columns and 'SALE_AMT_TY' in combined_df_item.columns:
                    item_exposure = combined_df_item.groupby(item_column)['노출수'].sum()
                    item_sales = combined_df_item.groupby(item_column)['SALE_AMT_TY'].sum()

                    if not item_exposure.empty and not item_sales.empty:
                        item_efficiency = {}
                        for item in item_exposure.index:
                            if item in item_sales.index and item_exposure[item] > 0:
                                efficiency = item_sales[item] / item_exposure[item]
                                item_efficiency[item] = efficiency

                        if item_efficiency:
                            best_efficiency_item = max(item_efficiency, key=item_efficiency.get)
                            best_efficiency = item_efficiency[best_efficiency_item]
                            insights.append(f"⚡ **{best_efficiency_item} 아이템이 가장 효율적입니다!** 노출 1회당 {best_efficiency:,.0f}원의 매출을 만들어냅니다. 이 아이템의 성공 공식을 분석해보세요.")
            else:
                # 아이템 컬럼이 없는 경우
                insights.append(f"📊 **아이템별 분석을 위한 데이터가 부족합니다.** 현재 데이터로는 아이템별 성과를 비교할 수 없습니다.")
            
            # 6. 통합적 비즈니스 인사이트 생성
            if insights:
                st.markdown("### 🧠 AI 비즈니스 인사이트")
                
                # 핵심 성과 지표 추출
                best_channel = None
                worst_channel = None
                growth_rate = None
                total_exposure = 0
                total_sales = 0
                
                # 데이터에서 핵심 지표 추출
                for insight in insights:
                    if "가장 효과적" in insight and "인플루언서" in insight:
                        best_channel = "인플루언서 마케팅"
                    elif "가장 효과적" in insight:
                        best_channel = insight.split("이 가장 효과적")[0].split("**")[-1]
                    elif "대박" in insight and "성장" in insight:
                        growth_rate = insight.split("대박! 전년 대비 ")[1].split("%")[0] if "대박! 전년 대비 " in insight else None
                    elif "성장" in insight and "%" in insight:
                        growth_rate = insight.split("전년 대비 ")[1].split("%")[0] if "전년 대비 " in insight else None
                
                # 노출수 합계 계산
                if '노출수' in combined_df_item.columns:
                    total_exposure = combined_df_item['노출수'].sum()
                elif any('_노출수' in col for col in combined_df_item.columns):
                    exposure_cols = [col for col in combined_df_item.columns if '_노출수' in col]
                    total_exposure = combined_df_item[exposure_cols].sum().sum()
                
                # 매출액 합계 계산
                if 'SALE_AMT_TY' in combined_df_item.columns:
                    total_sales = combined_df_item['SALE_AMT_TY'].sum()
                
                # 통합 인사이트 생성
                st.markdown("#### 📊 핵심 성과 지표")
                
                # 1. 전체 성과 요약
                if total_exposure > 0 and total_sales > 0:
                    efficiency = total_sales / total_exposure
                    st.markdown(f"**💰 매출 성과**: {total_sales:,.0f}원 (노출당 {efficiency:,.0f}원)")
                    st.markdown(f"**📈 노출 성과**: {total_exposure:,.0f}회 노출")
                
                # 2. 성장률 분석
                if growth_rate:
                    if float(growth_rate) > 50:
                        st.markdown(f"**📈 YoY 성장률**: {growth_rate}% (50% 이상 고성장)")
                    elif float(growth_rate) > 20:
                        st.markdown(f"**📈 YoY 성장률**: {growth_rate}% (20% 이상 양호한 성장)")
                    else:
                        st.markdown(f"**📈 YoY 성장률**: {growth_rate}% (20% 미만 저성장)")
                
                # 3. 최고 성과 채널
                if best_channel:
                    st.markdown(f"**🏆 최고 성과 채널**: {best_channel}")
                
                # 4. 구체적 채널별 상관관계 분석
                st.markdown("#### 🎯 채널별 매출 기여도")
                
                # 상관관계가 강한 채널들만 필터링
                strong_correlations = []
                moderate_correlations = []
                
                for insight in insights:
                    if "강력한 영향을 미치고 있습니다" in insight:
                        strong_correlations.append(insight)
                    elif "매출 증가에 도움이 되고 있습니다" in insight:
                        moderate_correlations.append(insight)
                
                if strong_correlations:
                    st.markdown("**📊 높은 매출 기여도 (상관계수 0.7+):**")
                    for insight in strong_correlations:
                        # 구체적인 상관계수와 해석 추가
                        if "인플루언서" in insight:
                            st.markdown(f"- **인플루언서 마케팅**: 노출수와 매출액 간 강한 양의 상관관계 확인")
                        elif "SEO" in insight:
                            st.markdown(f"- **검색엔진 최적화**: 검색 노출과 매출 간 높은 상관관계 확인")
                        elif "자사" in insight:
                            st.markdown(f"- **자사 인스타그램**: 브랜드 계정 노출과 매출 간 강한 연관성 확인")
                        else:
                            st.markdown(f"- {insight}")
                
                if moderate_correlations:
                    st.markdown("**📈 중간 매출 기여도 (상관계수 0.3-0.7):**")
                    for insight in moderate_correlations:
                        if "SEO" in insight:
                            st.markdown(f"- **검색엔진 최적화**: 노출수와 매출 간 중간 수준의 양의 상관관계 확인")
                        elif "자사" in insight:
                            st.markdown(f"- **자사 인스타그램**: 노출수와 매출 간 중간 수준의 양의 상관관계 확인")
                        else:
                            st.markdown(f"- {insight}")
                
                # 5. 비용 효율성 분석
                cost_insights = [insight for insight in insights if "투자가 매출에 큰 도움이 되고 있습니다" in insight]
                if cost_insights:
                    st.markdown("**💰 높은 비용 효율성:**")
                    for insight in cost_insights:
                        st.markdown(f"- {insight}")
                
            # 6. AI 알고리즘 기반 지능형 분석
            st.markdown("#### 🤖 AI 알고리즘 분석")
            
            # 다차원 데이터 분석을 위한 AI 인사이트 생성
            ai_insights = []
            
            # 1. 패턴 인식 알고리즘 - 시계열 트렌드 분석
            if 'DT' in combined_df_item.columns and 'SALE_AMT_TY' in combined_df_item.columns:
                # 날짜별 매출 트렌드 분석
                daily_sales = combined_df_item.groupby('DT')['SALE_AMT_TY'].sum().reset_index()
                daily_sales['DT'] = pd.to_datetime(daily_sales['DT'])
                daily_sales = daily_sales.sort_values('DT')
                
                if len(daily_sales) > 1:
                    # 매출 증가/감소 패턴 분석
                    daily_sales['매출_변화율'] = daily_sales['SALE_AMT_TY'].pct_change() * 100
                    avg_growth = daily_sales['매출_변화율'].mean()
                    volatility = daily_sales['매출_변화율'].std()
                    
                    if avg_growth > 5:
                        ai_insights.append(f"📈 **상승 트렌드**: 일평균 {avg_growth:.1f}% 매출 증가 패턴 감지")
                    elif avg_growth < -5:
                        ai_insights.append(f"📉 **하락 트렌드**: 일평균 {abs(avg_growth):.1f}% 매출 감소 패턴 감지")
                    
                    if volatility > 50:
                        ai_insights.append(f"⚡ **높은 변동성**: 매출 변동폭이 {volatility:.0f}%로 불안정한 패턴")
                    elif volatility < 10:
                        ai_insights.append(f"📊 **안정적 패턴**: 매출 변동폭이 {volatility:.0f}%로 예측 가능한 패턴")
            
            # 2. 클러스터링 알고리즘 - 성과 그룹 분석
            performance_metrics = []
            
            # 캠페인별 성과 클러스터링
            if '캠페인명' in combined_df_item.columns or '캠페인' in combined_df_item.columns:
                campaign_col = '캠페인명' if '캠페인명' in combined_df_item.columns else '캠페인'
                if '노출수' in combined_df_item.columns and 'SALE_AMT_TY' in combined_df_item.columns:
                    campaign_data = combined_df_item.groupby(campaign_col).agg({
                        '노출수': 'sum',
                        'SALE_AMT_TY': 'sum'
                    }).reset_index()
                    campaign_data['효율성'] = campaign_data['SALE_AMT_TY'] / campaign_data['노출수']
                    
                    # K-means 클러스터링 시뮬레이션 (간단한 분위수 기반)
                    campaign_data['성과_등급'] = pd.qcut(campaign_data['효율성'], q=3, labels=['저성과', '중성과', '고성과'])
                    
                    high_performers = campaign_data[campaign_data['성과_등급'] == '고성과']
                    if not high_performers.empty:
                        ai_insights.append(f"🎯 **고성과 캠페인 그룹**: {len(high_performers)}개 캠페인이 평균 효율성 {high_performers['효율성'].mean():,.0f}원/노출 달성")
            
            # 3. 이상치 탐지 알고리즘 - 비정상 패턴 감지
            if '노출수' in combined_df_item.columns and 'SALE_AMT_TY' in combined_df_item.columns:
                # Z-score 기반 이상치 탐지
                exposure_z = (combined_df_item['노출수'] - combined_df_item['노출수'].mean()) / combined_df_item['노출수'].std()
                sales_z = (combined_df_item['SALE_AMT_TY'] - combined_df_item['SALE_AMT_TY'].mean()) / combined_df_item['SALE_AMT_TY'].std()
                
                outliers = combined_df_item[(abs(exposure_z) > 2) | (abs(sales_z) > 2)]
                if not outliers.empty:
                    ai_insights.append(f"🔍 **이상치 감지**: {len(outliers)}개 데이터 포인트에서 비정상적 패턴 발견")
            
            # 4. 상관관계 네트워크 분석
            correlation_matrix = combined_df_item.select_dtypes(include=[np.number]).corr()
            
            # 강한 상관관계 쌍 찾기
            strong_pairs = []
            for i in range(len(correlation_matrix.columns)):
                for j in range(i+1, len(correlation_matrix.columns)):
                    corr_val = correlation_matrix.iloc[i, j]
                    if abs(corr_val) > 0.7 and not pd.isna(corr_val):
                        strong_pairs.append((correlation_matrix.columns[i], correlation_matrix.columns[j], corr_val))
            
            if strong_pairs:
                ai_insights.append(f"🔗 **강한 상관관계 네트워크**: {len(strong_pairs)}개 변수 간 강한 연관성 발견")
                for var1, var2, corr in strong_pairs[:3]:  # 상위 3개만 표시
                    ai_insights.append(f"   - {var1} ↔ {var2}: {corr:.3f}")
            
            # 5. 고급 머신러닝 알고리즘 분석
            if len(combined_df_item) > 10:  # 충분한 데이터가 있을 때만
                try:
                    from sklearn.linear_model import LinearRegression, Ridge, Lasso
                    from sklearn.ensemble import RandomForestRegressor
                    from sklearn.metrics import r2_score, mean_absolute_error
                    from sklearn.preprocessing import StandardScaler
                    from sklearn.model_selection import cross_val_score
                    
                    # 다중 변수 예측 모델
                    numeric_cols = combined_df_item.select_dtypes(include=[np.number]).columns
                    feature_cols = [col for col in numeric_cols if col != 'SALE_AMT_TY' and col != 'SALE_AMT_LY']
                    
                    if len(feature_cols) > 0 and 'SALE_AMT_TY' in combined_df_item.columns:
                        X = combined_df_item[feature_cols].fillna(0)
                        y = combined_df_item['SALE_AMT_TY'].fillna(0)
                        
                        # 데이터 정규화
                        scaler = StandardScaler()
                        X_scaled = scaler.fit_transform(X)
                        
                        # 여러 모델 비교
                        models = {
                            'Linear Regression': LinearRegression(),
                            'Ridge Regression': Ridge(alpha=1.0),
                            'Lasso Regression': Lasso(alpha=1.0),
                            'Random Forest': RandomForestRegressor(n_estimators=100, random_state=42)
                        }
                        
                        best_model = None
                        best_score = -np.inf
                        model_scores = {}
                        
                        for name, model in models.items():
                            try:
                                # 교차 검증으로 모델 성능 평가
                                scores = cross_val_score(model, X_scaled, y, cv=3, scoring='r2')
                                avg_score = scores.mean()
                                model_scores[name] = avg_score
                                
                                if avg_score > best_score:
                                    best_score = avg_score
                                    best_model = name
                            except:
                                continue
                        
                        # 최고 성능 모델 결과
                        if best_model and best_score > 0:
                            ai_insights.append(f"🤖 **최고 예측 모델**: {best_model} (R² = {best_score:.3f})")
                            
                            # 모델별 성능 비교
                            if len(model_scores) > 1:
                                sorted_models = sorted(model_scores.items(), key=lambda x: x[1], reverse=True)
                                ai_insights.append(f"📊 **모델 성능 순위**:")
                                for i, (model_name, score) in enumerate(sorted_models[:3], 1):
                                    ai_insights.append(f"   {i}. {model_name}: {score:.3f}")
                            
                            # 예측력 해석
                            if best_score > 0.8:
                                ai_insights.append(f"🎯 **매우 높은 예측력**: {best_score:.1%}로 매우 정확한 예측 가능")
                            elif best_score > 0.6:
                                ai_insights.append(f"📈 **높은 예측력**: {best_score:.1%}로 상당히 정확한 예측 가능")
                            elif best_score > 0.3:
                                ai_insights.append(f"📊 **중간 예측력**: {best_score:.1%}로 어느 정도 예측 가능")
                            else:
                                ai_insights.append(f"⚠️ **낮은 예측력**: {best_score:.1%}로 예측이 어려움")
                        
                        # 특성 중요도 분석 (Random Forest)
                        if 'Random Forest' in models:
                            try:
                                rf_model = RandomForestRegressor(n_estimators=100, random_state=42)
                                rf_model.fit(X_scaled, y)
                                feature_importance = rf_model.feature_importances_
                                
                                # 상위 3개 중요 특성
                                importance_pairs = list(zip(feature_cols, feature_importance))
                                importance_pairs.sort(key=lambda x: x[1], reverse=True)
                                
                                if importance_pairs:
                                    ai_insights.append(f"🔍 **특성 중요도 분석**:")
                                    for i, (feature, importance) in enumerate(importance_pairs[:3], 1):
                                        ai_insights.append(f"   {i}. {feature}: {importance:.3f}")
                            except:
                                pass
                                
                except ImportError:
                    # sklearn이 없는 경우 기본 분석
                    if '노출수' in combined_df_item.columns and 'SALE_AMT_TY' in combined_df_item.columns:
                        correlation = combined_df_item['노출수'].corr(combined_df_item['SALE_AMT_TY'])
                        if not pd.isna(correlation):
                            ai_insights.append(f"📊 **기본 상관관계**: 노출수-매출액 상관계수 {correlation:.3f}")
            
            # 6. 동적 인사이트 생성
            if ai_insights:
                st.markdown("**🧠 AI 알고리즘 분석 결과:**")
                for insight in ai_insights:
                    st.markdown(f"- {insight}")
            else:
                st.markdown("**📊 데이터 분석**: 현재 데이터로는 AI 알고리즘 분석이 제한적입니다.")
            
            # 7. 전략적 제안
            st.markdown("#### 📋 전략적 제안")
            
            if strong_correlations:
                st.markdown("**📈 확장 권장 채널:**")
                if any("인플루언서" in insight for insight in strong_correlations):
                    st.markdown("- 인플루언서 마케팅: 예산 증액 및 파트너십 확대 검토")
                if any("SEO" in insight for insight in strong_correlations):
                    st.markdown("- SEO: 콘텐츠 제작 및 키워드 최적화 강화")
            
            if moderate_correlations:
                st.markdown("**🔧 개선 권장 채널:**")
                if any("SEO" in insight for insight in moderate_correlations):
                    st.markdown("- SEO: 콘텐츠 품질 향상 및 백링크 구축")
                if any("자사" in insight for insight in moderate_correlations):
                    st.markdown("- 자사 인스타그램: 타겟팅 정교화 및 콘텐츠 전략 개선")
                
            else:
                st.info("현재 필터 조건에서 분석할 수 있는 충분한 데이터가 없습니다.")
                
        else:
            st.info("분석을 위한 데이터가 준비되지 않았습니다.")
            
    except Exception as e:
        st.warning(f"인사이트 생성 중 오류가 발생했습니다: {e}")


# =============================================================================
# F&F CREW LIST 관련 함수들
# =============================================================================

def prepare_influencer_summary(df, brand_filter=None, season_filter=None):
    """인플루언서 요약 데이터 준비"""
    if df.empty:
        return pd.DataFrame()
    
    # 필터링
    filtered_df = df.copy()
    if brand_filter and brand_filter != "전체":
        filtered_df = filtered_df[filtered_df['브랜드'] == brand_filter]
    if season_filter and season_filter != "전체":
        filtered_df = filtered_df[filtered_df['시즌'] == season_filter]
    
    # 필요한 컬럼 확인 및 선택
    required_columns = ["sns_id", "name", "follower", "unit_fee", "sec_usage", "sec_period"]
    available_columns = [col for col in required_columns if col in filtered_df.columns]
    
    if not available_columns:
        st.error("필요한 컬럼이 데이터에 없습니다.")
        return pd.DataFrame()
    
    influencer_summary = filtered_df[available_columns].copy()
    
    # 중복 제거
    influencer_summary = influencer_summary.drop_duplicates()
    
    return influencer_summary

def render_influencer_tab(df):
    """F&F CREW LIST 탭 렌더링"""
    st.markdown("# 👥 F&F CREW LIST")
    
    if df.empty:
        st.warning("인플루언서 데이터가 없습니다.")
        return
    
    # 필터 섹션
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**🏷️ 브랜드**")
        brands = df['브랜드'].unique() if '브랜드' in df.columns else []
        selected_brand = st.selectbox("브랜드를 선택하세요", ["전체"] + list(brands), key="crew_brand_filter", label_visibility="collapsed")
    
    with col2:
        st.markdown("**📅 시즌**")
        seasons = df['시즌'].unique() if '시즌' in df.columns else []
        selected_season = st.selectbox("시즌을 선택하세요", ["전체"] + list(seasons), key="crew_season_filter", label_visibility="collapsed")
    
    # 인플루언서 데이터 준비
    influencer_summary = prepare_influencer_summary(df, selected_brand, selected_season)
    
    if influencer_summary.empty:
        st.warning("선택된 조건에 맞는 인플루언서 데이터가 없습니다.")
        return
    
    # 데이터 표시
    st.dataframe(influencer_summary, use_container_width=True)
    
    # 통계 정보
    st.markdown("## 📊 통계 정보")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("총 인플루언서 수", len(influencer_summary))
    with col2:
        avg_followers = influencer_summary['follower'].mean() if 'follower' in influencer_summary.columns else 0
        st.metric("평균 팔로워 수", f"{avg_followers:,.0f}")
    with col3:
        avg_fee = influencer_summary['unit_fee'].mean() if 'unit_fee' in influencer_summary.columns else 0
        st.metric("평균 계약단가", f"{avg_fee:,.0f}원")

# =============================================================================
# 자동 배정 시스템 관련 함수들
# =============================================================================

def execute_optimal_assignment(month, targets_df, influencer_df):
    """최적 배정 알고리즘 실행"""
    try:
        import pulp
        
        # 문제 생성
        prob = pulp.LpProblem("OptimalAssignment", pulp.LpMaximize)
        
        # 변수 생성
        assignments = {}
        for _, target in targets_df.iterrows():
            brand = target['브랜드']
            for _, influencer in influencer_df.iterrows():
                influencer_id = influencer['sns_id']
                var_name = f"assign_{brand}_{influencer_id}"
                assignments[(brand, influencer_id)] = pulp.LpVariable(var_name, cat='Binary')
        
        # 목적함수: 총 계약수량 최대화
        prob += pulp.lpSum([assignments[(brand, influencer_id)] * target['계약수량'] 
                      for (brand, influencer_id), var in assignments.items()
                      for _, target in targets_df.iterrows()
                      if target['브랜드'] == brand])
        
        # 제약조건
        # 1. 각 인플루언서는 한 브랜드에만 배정
        for influencer_id in influencer_df['sns_id']:
            prob += pulp.lpSum([assignments[(brand, influencer_id)] for brand in targets_df['브랜드']]) <= 1
        
        # 2. 각 브랜드의 요청수량 충족
        for _, target in targets_df.iterrows():
            brand = target['브랜드']
            prob += pulp.lpSum([assignments[(brand, influencer_id)] * target['계약수량'] 
                          for influencer_id in influencer_df['sns_id']]) >= target['요청수량']
        
        # 문제 해결
        prob.solve()
        
        # 결과 추출
        results = []
        for (brand, influencer_id), var in assignments.items():
            if var.varValue == 1:
                influencer_info = influencer_df[influencer_df['sns_id'] == influencer_id].iloc[0]
                results.append({
                    'sns_id': influencer_id,
                    '브랜드': brand,
                    '배정월': month,
                    '이름': influencer_info['name'],
                    'FLW': influencer_info['follower'],
                    '1회계약단가': influencer_info['unit_fee'],
                    '2차활용': influencer_info['sec_usage'],
                    '2차기간': influencer_info['sec_period'],
                    '계약수량': target['계약수량'],
                    '배정여부': '배정',
                    '집행상태': '미집행',
                    '집행수량': 0,
                    '최종상태': '배정'
                })
        
        return results
        
    except ImportError:
        st.error("PuLP 라이브러리가 설치되지 않았습니다. pip install pulp를 실행해주세요.")
        return []
    except Exception as e:
        st.error(f"최적 배정 실행 중 오류가 발생했습니다: {str(e)}")
        return []

def render_assignment_tab():
    """자동 배정 탭 렌더링"""
    st.markdown("# 🎯 자동 배정 시스템")
    
    # 월별 배정 목표 데이터 로드
    targets_df = load_monthly_targets()
    influencer_df = load_influencer_data()
    
    if targets_df.empty:
        st.warning("월별 배정 목표 데이터가 없습니다.")
        return
    
    if influencer_df.empty:
        st.warning("인플루언서 데이터가 없습니다.")
        return
    
    # 월 선택
    if 'month' in targets_df.columns:
        months = targets_df['month'].unique()
        selected_month = st.selectbox("배정할 월을 선택하세요", months)
        
        # 해당 월의 목표 데이터
        month_targets = targets_df[targets_df['month'] == selected_month]
    else:
        st.warning("배정 목표 데이터가 없습니다. 먼저 배정 목표를 설정해주세요.")
        return
    
    # 목표 데이터 표시
    st.markdown("## 📋 배정 목표")
    st.dataframe(month_targets, use_container_width=True)
    
    # 자동 배정 실행
    if st.button("🎯 자동 배정 실행", type="primary"):
        with st.spinner("최적 배정을 실행 중입니다..."):
            results = execute_optimal_assignment(selected_month, month_targets, influencer_df)
            
            if results:
                st.success(f"✅ {len(results)}명의 인플루언서가 배정되었습니다.")
                
                # 결과 저장
                results_df = pd.DataFrame(results)
                existing_assignments = load_assignment_history()
                updated_assignments = pd.concat([existing_assignments, results_df], ignore_index=True)
                save_assignment_history(updated_assignments)
                
                # 결과 표시
                st.markdown("## 📊 배정 결과")
                st.dataframe(results_df, use_container_width=True)
            else:
                st.warning("배정 가능한 인플루언서가 없습니다.")

# =============================================================================
# 전체 집행 데이터 관리 관련 함수들
# =============================================================================

# 마케팅 데이터 파일 경로
MARKETING_FILE = 'data/marketing_data.csv'

def load_marketing_data():
    """마케팅 데이터 로드"""
    if os.path.exists(MARKETING_FILE):
        try:
            return pd.read_csv(MARKETING_FILE, encoding='utf-8')
        except:
            return pd.read_csv(MARKETING_FILE, encoding='cp949')
    return pd.DataFrame()

def save_marketing_data(df):
    """마케팅 데이터 저장"""
    os.makedirs('data', exist_ok=True)
    df.to_csv(MARKETING_FILE, index=False, encoding='utf-8')

def render_execution_data_management_tab():
    """데이터 업로드 관리 탭 렌더링"""
    st.markdown("# 📈 데이터 업로드 관리")
    
    # 인플루언서 데이터 표시
    execution_df = load_execution_data()
    if not execution_df.empty:
        st.markdown("## 📊 인플루언서 데이터")
        st.success(f"총 {len(execution_df)}건의 인플루언서 데이터가 있습니다.")
        
        # 필터
        col1, col2, col3 = st.columns(3)
        
        with col1:
            seasons = execution_df['시즌'].unique() if '시즌' in execution_df.columns else []
            selected_season = st.selectbox("📅 시즌", ["전체"] + list(seasons), key="execution_season_filter")
        
        with col2:
            brands = execution_df['브랜드'].unique() if '브랜드' in execution_df.columns else []
            selected_brand = st.selectbox("🏷️ 브랜드", ["전체"] + list(brands), key="execution_brand_filter")
        
        with col3:
            # 아이템 필터 (쉼표로 분리된 값들을 개별적으로 처리)
            if '아이템' in execution_df.columns:
                # 모든 아이템 값을 쉼표로 분리하여 개별 아이템 목록 생성
                all_items = []
                for items_str in execution_df['아이템'].dropna():
                    if isinstance(items_str, str):
                        # 쉼표로 분리하고 공백 제거
                        items = [item.strip() for item in items_str.split(',')]
                        all_items.extend(items)
                
                # 중복 제거 및 정렬
                unique_items = sorted(list(set(all_items)))
                selected_execution_items = st.multiselect(
                    "📦 아이템", 
                    unique_items, 
                    key="execution_item_filter",
                    placeholder="아이템을 선택하세요"
                )
            else:
                selected_execution_items = []
        
        # 데이터 필터링
        filtered_execution_df = execution_df.copy()
        if selected_season != "전체":
            filtered_execution_df = filtered_execution_df[filtered_execution_df['시즌'] == selected_season]
        if selected_brand != "전체":
            filtered_execution_df = filtered_execution_df[filtered_execution_df['브랜드'] == selected_brand]
        
        # 아이템 필터링 (쉼표로 분리된 값들 중 하나라도 선택된 아이템과 일치하면 포함)
        if selected_execution_items and '아이템' in filtered_execution_df.columns:
            def contains_selected_item(items_str):
                if pd.isna(items_str) or not isinstance(items_str, str):
                    return False
                items = [item.strip() for item in items_str.split(',')]
                return any(item in selected_execution_items for item in items)
            
            filtered_execution_df = filtered_execution_df[filtered_execution_df['아이템'].apply(contains_selected_item)]
        
        # 시트명 컬럼 제거 및 아이템 컬럼 추가 (내부 처리용이므로 표시하지 않음)
        display_df = filtered_execution_df.copy()
        if '시트명' in display_df.columns:
            display_df = display_df.drop(columns=['시트명'])
        
        # 아이템 컬럼이 없으면 빈 컬럼으로 추가 (메인제품 앞에 위치)
        if '아이템' not in display_df.columns:
            # 메인제품 컬럼의 위치를 찾아서 그 앞에 아이템 컬럼 추가
            if '메인제품' in display_df.columns:
                main_product_idx = display_df.columns.get_loc('메인제품')
                display_df.insert(main_product_idx, '아이템', None)
            else:
                display_df['아이템'] = None
        
        st.dataframe(display_df, use_container_width=True)
        
        # 관리 옵션
        col1, col2 = st.columns(2)
        
        with col1:
            # 엑셀 파일 생성 (시트명 컬럼 제거 및 컬럼 순서 정리)
            download_df = filtered_execution_df.copy()
            if '시트명' in download_df.columns:
                download_df = download_df.drop(columns=['시트명'])
            
            # 표준 컬럼 순서 정의 (아이템이 메인제품 앞에 오도록)
            standard_columns = [
                '유형', '브랜드', '시즌', '연도', '캠페인월', '업로드월', '업로드일',
                '채널', '이름', 'sns_id', '캠페인명', '아이템', '메인제품',
                '컨텐츠URL', '컨텐츠유형', '팔로워', '노출수', '좋아요', '댓글수', '조회수', '전체비용'
            ]
            
            # 존재하는 컬럼만 선택하고 순서 정리
            existing_columns = [col for col in standard_columns if col in download_df.columns]
            other_columns = [col for col in download_df.columns if col not in standard_columns]
            final_columns = existing_columns + other_columns
            
            download_df = download_df[final_columns]
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                download_df.to_excel(writer, index=False, sheet_name='집행데이터')
            output.seek(0)
            
            st.download_button(
                label="📥 엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"집행데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            if st.button("🗑️ 기존 데이터 삭제", use_container_width=True):
                if os.path.exists(EXECUTION_FILE):
                    os.remove(EXECUTION_FILE)
                    st.success("집행 데이터가 삭제되었습니다.")
                    st.rerun()
    
    # 인플루언서 데이터 업로드
    st.markdown("---")
    st.markdown("### 📤 인플루언서 데이터 업로드")
    
    # 2단 레이아웃
    col1, col2 = st.columns([1, 1])
    
    with col1:
        # 템플릿 다운로드 (엑셀 다운로드와 동일한 컬럼 구조)
        template_data = {
            '유형': ['인플루언서', '인플루언서', '인플루언서', '인플루언서'],
            '브랜드': ['MLB', 'DX', 'DV', 'ST'],
            '시즌': ['25FW', '25FW', '25FW', '25FW'],
            '연도': [2025, 2025, 2025, 2025],
            '캠페인월': ['5월', '5월', '6월', '6월'],
            '업로드월': ['5월', '5월', '6월', '6월'],
            '업로드일': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04'],
            '채널': ['인스타그램', '인스타그램', '인스타그램', '인스타그램'],
            '이름': ['인플루언서1', '인플루언서2', '인플루언서3', '인플루언서4'],
            'sns_id': ['influencer1', 'influencer2', 'influencer3', 'influencer4'],
            '캠페인유형': ['5월인플루언서', '5월인플루언서', '6월인플루언서', '6월인플루언서'],
            '캠페인명': ['캠페인1', '캠페인2', '캠페인3', '캠페인4'],
            '아이템': ['아이템1', '아이템2', '아이템3', '아이템4'],
            '메인제품': ['제품1', '제품2', '제품3', '제품4'],
            '서브제품1': ['서브제품1-1', '서브제품1-2', '서브제품1-3', '서브제품1-4'],
            '서브제품2': ['서브제품2-1', '서브제품2-2', '서브제품2-3', '서브제품2-4'],
            '컨텐츠URL': ['https://instagram.com/1', 'https://instagram.com/2', 'https://instagram.com/3', 'https://instagram.com/4'],
            '컨텐츠유형': ['PHOTO', 'PHOTO', 'VIDEO', 'VIDEO'],
            '팔로워': [10000, 20000, 15000, 30000],
            '노출수': [1000, 2000, 1500, 3000],
            '좋아요': [100, 200, 150, 300],
            '댓글수': [50, 100, 75, 150],
            '조회수': [5000, 10000, 7500, 15000],
            '전체비용': [100000, 200000, 150000, 300000]
        }
        
        template_df = pd.DataFrame(template_data)
        template_output = io.BytesIO()
        with pd.ExcelWriter(template_output, engine='openpyxl') as writer:
            template_df.to_excel(writer, index=False, sheet_name='집행데이터템플릿')
        template_output.seek(0)
        
        st.download_button(
            label="📥 인플루언서 데이터 템플릿 다운로드",
            data=template_output.getvalue(),
            file_name="인플루언서데이터_템플릿.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        # 사용 가이드
        st.markdown("**필수 컬럼:**")
        st.markdown("""
        - **유형**: 데이터 유형 (인플루언서, 마케팅 등)
        - **시즌**: 시즌 정보 (25FW, 26SS 등)
        - **업로드일**: 업로드 날짜 (YYYY-MM-DD 형식)
        - **sns_id**: SNS 아이디
        """)
    
    with col2:
        # 업로드 모드 선택 및 파일 업로드
        upload_mode_exec = st.radio("업로드 모드", ["추가", "교체"], horizontal=True, key="exec_upload_mode")
        uploaded_file = st.file_uploader(
            "인플루언서 데이터 파일을 업로드하세요",
            type=['xlsx', 'csv'],
            help="Excel 또는 CSV 파일을 업로드할 수 있습니다.",
            key="execution_upload"
        )
    
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                new_execution_df = pd.read_csv(uploaded_file)
            else:
                # 엑셀 파일의 모든 시트 읽기
                excel_file = pd.ExcelFile(uploaded_file)
                sheet_names = excel_file.sheet_names
                
                # 모든 시트의 데이터를 통합
                all_sheets_data = []
                for sheet_name in sheet_names:
                    try:
                        sheet_df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                        if not sheet_df.empty:
                            # 시트명을 컬럼으로 추가 (선택사항)
                            sheet_df['시트명'] = sheet_name
                            all_sheets_data.append(sheet_df)
                    except Exception as e:
                        st.warning(f"시트 '{sheet_name}' 읽기 실패: {str(e)}")
                
                if all_sheets_data:
                    new_execution_df = pd.concat(all_sheets_data, ignore_index=True)
                    st.info(f"📊 총 {len(sheet_names)}개 시트에서 데이터를 읽었습니다: {', '.join(sheet_names)}")
                else:
                    st.error("읽을 수 있는 시트가 없습니다.")
                    return
            
            # 데이터 검증
            required_columns = ['유형', '시즌', '업로드일', 'sns_id']
            missing_columns = [col for col in required_columns if col not in new_execution_df.columns]
            
            if missing_columns:
                st.error(f"❌ 필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}")
                st.info(f"📋 현재 업로드된 파일의 컬럼: {list(new_execution_df.columns)}")
                st.info(f"📋 필요한 필수 컬럼: {required_columns}")
                st.warning("💡 인플루언서 데이터 템플릿을 다운로드하여 올바른 형식으로 데이터를 작성해주세요.")
            else:
                # 업로드일에서 월 추출하여 캠페인월, 업로드월 자동 채우기
                if '업로드일' in new_execution_df.columns:
                    # 업로드일을 datetime으로 변환
                    new_execution_df['업로드일'] = pd.to_datetime(new_execution_df['업로드일'], errors='coerce')
                    
                    # 캠페인월이 비어있으면 업로드일의 월로 채우기
                    if '캠페인월' in new_execution_df.columns:
                        mask_campaign = new_execution_df['캠페인월'].isna() | (new_execution_df['캠페인월'] == '') | (new_execution_df['캠페인월'] == 0)
                        new_execution_df.loc[mask_campaign, '캠페인월'] = new_execution_df.loc[mask_campaign, '업로드일'].dt.month
                    
                    # 업로드월이 비어있으면 업로드일의 월로 채우기
                    if '업로드월' in new_execution_df.columns:
                        mask_upload = new_execution_df['업로드월'].isna() | (new_execution_df['업로드월'] == '') | (new_execution_df['업로드월'] == 0)
                        new_execution_df.loc[mask_upload, '업로드월'] = new_execution_df.loc[mask_upload, '업로드일'].dt.month
                
                # 업로드 모드 처리 (추가/교체)
                if upload_mode_exec == "추가":
                    existing = load_execution_data()
                    combined = pd.concat([existing, new_execution_df], ignore_index=True)
                    # 중복 제거 기준: ['유형','브랜드','시즌','연도','업로드일','sns_id','캠페인명','컨텐츠URL'] 존재하는 컬럼만 사용
                    dedup_keys = [c for c in ['유형','브랜드','시즌','연도','업로드일','sns_id','캠페인명','컨텐츠URL'] if c in combined.columns]
                    if dedup_keys:
                        combined = combined.drop_duplicates(subset=dedup_keys, keep='last')
                    save_execution_data(combined)
                else:
                    save_execution_data(new_execution_df)
                st.success("집행 데이터가 성공적으로 업로드되었습니다.")
                st.rerun()
                
        except Exception as e:
            st.error(f"파일 업로드 중 오류가 발생했습니다: {str(e)}")
    
    # 마케팅 데이터 섹션
    st.markdown("---")
    st.markdown("## 📊 마케팅 데이터")
    
    # 마케팅 데이터 표시
    marketing_df = load_marketing_data()
    if not marketing_df.empty:
        st.success(f"총 {len(marketing_df)}건의 마케팅 데이터가 있습니다.")
        
        # 필터
        col1, col2, col3 = st.columns(3)
        
        with col1:
            seasons = marketing_df['시즌'].unique() if '시즌' in marketing_df.columns else []
            selected_marketing_season = st.selectbox("📅 시즌", ["전체"] + list(seasons), key="marketing_season_filter")
        
        with col2:
            brands = marketing_df['브랜드'].unique() if '브랜드' in marketing_df.columns else []
            selected_marketing_brand = st.selectbox("🏷️ 브랜드", ["전체"] + list(brands), key="marketing_brand_filter")
        
        with col3:
            # 아이템 필터 (쉼표로 분리된 값들을 개별적으로 처리)
            if '아이템' in marketing_df.columns:
                # 모든 아이템 값을 쉼표로 분리하여 개별 아이템 목록 생성
                all_items = []
                for items_str in marketing_df['아이템'].dropna():
                    if isinstance(items_str, str):
                        # 쉼표로 분리하고 공백 제거
                        items = [item.strip() for item in items_str.split(',')]
                        all_items.extend(items)
                
                # 중복 제거 및 정렬
                unique_items = sorted(list(set(all_items)))
                selected_marketing_items = st.multiselect(
                    "📦 아이템", 
                    unique_items, 
                    key="marketing_item_filter",
                    placeholder="아이템을 선택하세요"
                )
            else:
                selected_marketing_items = []
        
        # 데이터 필터링
        filtered_marketing_df = marketing_df.copy()
        if selected_marketing_season != "전체":
            filtered_marketing_df = filtered_marketing_df[filtered_marketing_df['시즌'] == selected_marketing_season]
        if selected_marketing_brand != "전체":
            filtered_marketing_df = filtered_marketing_df[filtered_marketing_df['브랜드'] == selected_marketing_brand]
        
        # 아이템 필터링 (쉼표로 분리된 값들 중 하나라도 선택된 아이템과 일치하면 포함)
        if selected_marketing_items and '아이템' in filtered_marketing_df.columns:
            def contains_selected_item(items_str):
                if pd.isna(items_str) or not isinstance(items_str, str):
                    return False
                items = [item.strip() for item in items_str.split(',')]
                return any(item in selected_marketing_items for item in items)
            
            filtered_marketing_df = filtered_marketing_df[filtered_marketing_df['아이템'].apply(contains_selected_item)]
        
        # 데이터 표시
        # 시트명 컬럼 제거 및 아이템 컬럼 추가 (내부 처리용이므로 표시하지 않음)
        display_marketing_df = filtered_marketing_df.copy()
        if '시트명' in display_marketing_df.columns:
            display_marketing_df = display_marketing_df.drop(columns=['시트명'])
        
        # 아이템 컬럼이 없으면 빈 컬럼으로 추가 (메인제품 앞에 위치)
        if '아이템' not in display_marketing_df.columns:
            # 메인제품 컬럼의 위치를 찾아서 그 앞에 아이템 컬럼 추가
            if '메인제품' in display_marketing_df.columns:
                main_product_idx = display_marketing_df.columns.get_loc('메인제품')
                display_marketing_df.insert(main_product_idx, '아이템', None)
            else:
                display_marketing_df['아이템'] = None
        
        st.dataframe(display_marketing_df, use_container_width=True)
        
        # 엑셀 다운로드 및 데이터 삭제 버튼 (인플루언서 데이터와 동일한 형식)
        col1, col2 = st.columns(2)
        
        with col1:
            # 엑셀 파일 생성 (컬럼 순서 정리)
            download_df = display_marketing_df.copy()
            
            # 표준 컬럼 순서 정의 (아이템이 메인제품 앞에 오도록)
            standard_columns = [
                '유형', '브랜드', '시즌', '연도', '캠페인월', '업로드월', '업로드일',
                '채널', '이름', 'sns_id', '캠페인명', '아이템', '메인제품',
                '컨텐츠URL', '컨텐츠유형', '팔로워', '노출수', '좋아요', '댓글수', '조회수', '전체비용'
            ]
            
            # 존재하는 컬럼만 선택하고 순서 정리
            existing_columns = [col for col in standard_columns if col in download_df.columns]
            other_columns = [col for col in download_df.columns if col not in standard_columns]
            final_columns = existing_columns + other_columns
            
            download_df = download_df[final_columns]
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                download_df.to_excel(writer, index=False, sheet_name='마케팅데이터')
            output.seek(0)
            
            st.download_button(
                label="📥 엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"마케팅데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="marketing_excel_download",
                use_container_width=True
            )
        
        with col2:
            if st.button("🗑️ 기존 데이터 삭제", key="marketing_delete", use_container_width=True):
                if os.path.exists(MARKETING_FILE):
                    os.remove(MARKETING_FILE)
                    st.success("마케팅 데이터가 삭제되었습니다.")
                    st.rerun()
    else:
        st.info("마케팅 데이터가 없습니다. 아래에서 업로드해주세요.")
    
    # 마케팅 데이터 업로드 섹션
    st.markdown("---")
    st.markdown("### 📤 마케팅 데이터 업로드")
    
    # 2단 레이아웃
    col1, col2 = st.columns([1, 1])
    
    with col1:
        # 템플릿 다운로드 (마케팅 데이터 컬럼 구조)
        marketing_template_data = {
            '유형': ['마케팅', '마케팅', '마케팅', '마케팅'],
            '브랜드': ['MLB', 'DX', 'DV', 'ST'],
            '시즌': ['25FW', '25FW', '25FW', '25FW'],
            '연도': [2025, 2025, 2025, 2025],
            '캠페인월': ['5월', '5월', '6월', '6월'],
            '업로드월': ['5월', '5월', '6월', '6월'],
            '업로드일': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04'],
            '채널': ['인스타그램', '인스타그램', '인스타그램', '인스타그램'],
            '이름': ['마케터1', '마케터2', '마케터3', '마케터4'],
            'sns_id': ['marketer1', 'marketer2', 'marketer3', 'marketer4'],
            '캠페인명': ['마케팅캠페인1', '마케팅캠페인2', '마케팅캠페인3', '마케팅캠페인4'],
            '아이템': ['아이템1', '아이템2', '아이템3', '아이템4'],
            '메인제품': ['제품1', '제품2', '제품3', '제품4'],
            '컨텐츠URL': ['https://instagram.com/1', 'https://instagram.com/2', 'https://instagram.com/3', 'https://instagram.com/4'],
            '컨텐츠유형': ['PHOTO', 'PHOTO', 'VIDEO', 'VIDEO'],
            '팔로워': [10000, 20000, 15000, 30000],
            '노출수': [1000, 2000, 1500, 3000],
            '좋아요': [100, 200, 150, 300],
            '댓글수': [50, 100, 75, 150],
            '조회수': [5000, 10000, 7500, 15000],
            '전체비용': [100000, 200000, 150000, 300000]
        }
        
        marketing_template_df = pd.DataFrame(marketing_template_data)
        marketing_template_output = io.BytesIO()
        with pd.ExcelWriter(marketing_template_output, engine='openpyxl') as writer:
            marketing_template_df.to_excel(writer, index=False, sheet_name='마케팅데이터템플릿')
        marketing_template_output.seek(0)
        
        st.download_button(
            label="📥 마케팅 데이터 템플릿 다운로드",
            data=marketing_template_output.getvalue(),
            file_name="마케팅데이터_템플릿.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        # 사용 가이드
        st.markdown("**필수 컬럼:**")
        st.markdown("""
        - **유형**: 데이터 유형 (인플루언서, 마케팅 등)
        - **시즌**: 시즌 정보 (25FW, 26SS 등)
        - **업로드일**: 업로드 날짜 (YYYY-MM-DD 형식)
        - **sns_id**: SNS 아이디
        """)
    
    with col2:
        # 업로드 모드 선택 및 파일 업로드
        upload_mode_mkt = st.radio("업로드 모드", ["추가", "교체"], horizontal=True, key="mkt_upload_mode")
        marketing_uploaded_file = st.file_uploader(
            "마케팅 데이터 파일을 업로드하세요",
            type=['xlsx', 'csv'],
            help="Excel 또는 CSV 파일을 업로드할 수 있습니다.",
            key="marketing_upload"
        )
    
    if marketing_uploaded_file is not None:
        try:
            if marketing_uploaded_file.name.endswith('.csv'):
                new_marketing_df = pd.read_csv(marketing_uploaded_file)
            else:
                # 엑셀 파일의 모든 시트 읽기
                excel_file = pd.ExcelFile(marketing_uploaded_file)
                sheet_names = excel_file.sheet_names
                
                # 모든 시트의 데이터를 통합
                all_sheets_data = []
                for sheet_name in sheet_names:
                    try:
                        sheet_df = pd.read_excel(marketing_uploaded_file, sheet_name=sheet_name)
                        if not sheet_df.empty:
                            # 시트명을 컬럼으로 추가 (선택사항)
                            sheet_df['시트명'] = sheet_name
                            all_sheets_data.append(sheet_df)
                    except Exception as e:
                        st.warning(f"시트 '{sheet_name}' 읽기 실패: {str(e)}")
                
                if all_sheets_data:
                    new_marketing_df = pd.concat(all_sheets_data, ignore_index=True)
                    st.info(f"📊 총 {len(sheet_names)}개 시트에서 데이터를 읽었습니다: {', '.join(sheet_names)}")
                else:
                    st.error("읽을 수 있는 시트가 없습니다.")
                    return
            
            # 데이터 검증
            required_columns = ['유형', '시즌', '업로드일', 'sns_id']
            missing_columns = [col for col in required_columns if col not in new_marketing_df.columns]
            
            if missing_columns:
                st.error(f"❌ 필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}")
                st.info(f"📋 현재 업로드된 파일의 컬럼: {list(new_marketing_df.columns)}")
                st.info(f"📋 필요한 필수 컬럼: {required_columns}")
                st.warning("💡 마케팅 데이터 템플릿을 다운로드하여 올바른 형식으로 데이터를 작성해주세요.")
            else:
                # 업로드일에서 월 추출하여 캠페인월, 업로드월 자동 채우기
                if '업로드일' in new_marketing_df.columns:
                    # 업로드일을 datetime으로 변환
                    new_marketing_df['업로드일'] = pd.to_datetime(new_marketing_df['업로드일'], errors='coerce')
                    
                    # 캠페인월이 비어있으면 업로드일의 월로 채우기
                    if '캠페인월' in new_marketing_df.columns:
                        mask_campaign = new_marketing_df['캠페인월'].isna() | (new_marketing_df['캠페인월'] == '') | (new_marketing_df['캠페인월'] == 0)
                        new_marketing_df.loc[mask_campaign, '캠페인월'] = new_marketing_df.loc[mask_campaign, '업로드일'].dt.month
                    
                    # 업로드월이 비어있으면 업로드일의 월로 채우기
                    if '업로드월' in new_marketing_df.columns:
                        mask_upload = new_marketing_df['업로드월'].isna() | (new_marketing_df['업로드월'] == '') | (new_marketing_df['업로드월'] == 0)
                        new_marketing_df.loc[mask_upload, '업로드월'] = new_marketing_df.loc[mask_upload, '업로드일'].dt.month
                
                # 업로드 모드 처리 (추가/교체)
                if upload_mode_mkt == "추가":
                    existing = load_marketing_data()
                    combined = pd.concat([existing, new_marketing_df], ignore_index=True)
                    dedup_keys = [c for c in ['유형','브랜드','시즌','연도','업로드일','sns_id','캠페인명','컨텐츠URL'] if c in combined.columns]
                    if dedup_keys:
                        combined = combined.drop_duplicates(subset=dedup_keys, keep='last')
                    save_marketing_data(combined)
                else:
                    save_marketing_data(new_marketing_df)
                st.success("마케팅 데이터가 성공적으로 업로드되었습니다.")
                st.rerun()
                
        except Exception as e:
            st.error(f"파일 업로드 중 오류가 발생했습니다: {str(e)}")

    # 매출 데이터 관리 섹션
    st.markdown("---")
    st.markdown("## 💰 매출 데이터 관리")
    
    
    # 매출 데이터 표시
    sales_df = load_sales_data()
    if not sales_df.empty:
        st.markdown("### 📊 매출 데이터")
        st.success(f"총 {len(sales_df)}건의 매출 데이터가 있습니다.")
        
        # 필터
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if 'BRD_CD' in sales_df.columns:
                valid_brands = ['M', 'X', 'V', 'ST']
                sales_df_filtered = sales_df[sales_df['BRD_CD'].isin(valid_brands)]
                brand_mapping = {'M': 'MLB', 'X': 'DX', 'V': 'DV', 'ST': 'ST'}
                unique_brands = sales_df_filtered['BRD_CD'].unique()
                brand_names = [brand_mapping.get(brand, brand) for brand in unique_brands]
                brand_order = ['MLB', 'DX', 'DV', 'ST']
                ordered_brands = [brand for brand in brand_order if brand in brand_names]
                selected_brd = st.selectbox("🏷️ 브랜드", ordered_brands, key="execution_sales_brand_filter")
            else:
                selected_brd = "MLB"
        
        with col2:
            if 'DT' in sales_df.columns:
                sales_df['월'] = pd.to_datetime(sales_df['DT']).dt.to_period('M')
                months = sorted(sales_df['월'].unique())
                selected_month = st.selectbox("📅 월", months, key="execution_sales_month_filter")
            else:
                selected_month = None
        
        with col3:
            if 'ITEM_NM' in sales_df.columns:
                categories = sales_df['ITEM_NM'].unique()
                selected_item = st.selectbox("📦 카테고리", ["전체"] + list(categories), key="execution_sales_item_filter")
            else:
                selected_item = "전체"
        
        # 데이터 필터링
        filtered_sales_df = sales_df.copy()
        if 'BRD_CD' in sales_df.columns:
            brand_reverse_mapping = {'MLB': 'M', 'DX': 'X', 'DV': 'V', 'ST': 'ST'}
            brand_code = brand_reverse_mapping.get(selected_brd, selected_brd)
            filtered_sales_df = filtered_sales_df[filtered_sales_df['BRD_CD'] == brand_code]
        
        if selected_month and '월' in sales_df.columns:
            filtered_sales_df = filtered_sales_df[filtered_sales_df['월'] == selected_month]
        
        if selected_item != "전체" and 'ITEM_NM' in sales_df.columns:
            filtered_sales_df = filtered_sales_df[filtered_sales_df['ITEM_NM'] == selected_item]
        
        # 표시용 데이터 준비
        display_sales_df = filtered_sales_df.copy()
        if 'DT' in display_sales_df.columns:
            display_sales_df['DT'] = pd.to_datetime(display_sales_df['DT'], errors='coerce').dt.date
        
        # 컬럼 순서 조정
        column_order = ['BRD_CD', 'DT', 'ITEM', 'ITEM_NM', 'SALE_AMT_TY', 'SALE_QTY_TY', 'SALE_AMT_LY', 'SALE_QTY_LY']
        available_columns = [col for col in column_order if col in display_sales_df.columns]
        display_sales_df = display_sales_df[available_columns]
        
        st.dataframe(display_sales_df, use_container_width=True)
        
        # 매출 데이터 관리 옵션
        col1, col2 = st.columns(2)
        
        with col1:
            # 엑셀 파일 생성
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                display_sales_df.to_excel(writer, index=False, sheet_name='매출데이터')
            output.seek(0)
            
            st.download_button(
                label="📥 매출 데이터 엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"매출데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            if st.button("🗑️ 매출 데이터 삭제", use_container_width=True):
                # 먼저 캐시 초기화
                st.cache_data.clear()
                
                deleted_files = []
                if os.path.exists(SALES_FILE):
                    os.remove(SALES_FILE)
                    deleted_files.append("sales_data.csv")
                
                # 모든 매출 데이터 파일 삭제
                import glob
                sales_files = glob.glob(os.path.join(DATA_DIR, "*sales*.csv"))
                for file in sales_files:
                    if os.path.exists(file):
                        os.remove(file)
                        deleted_files.append(os.path.basename(file))
                
                if deleted_files:
                    st.success(f"✅ 다음 파일들이 삭제되었습니다: {', '.join(deleted_files)}")
                    # 즉시 페이지 새로고침
                    st.rerun()
                else:
                    st.info("삭제할 매출 데이터가 없습니다.")
    
    # 매출 데이터 업로드
    st.markdown("---")
    st.markdown("### 📤 매출 데이터 업로드")
    
    # Snowflake 연동 섹션
    if SNOWFLAKE_AVAILABLE:
        st.markdown("#### ❄️ Snowflake 연동")

        # 2열 레이아웃으로 나란히 배치
        col1, col2 = st.columns([1, 1])
        
        with col1:
            # Snowflake 연결 상태 및 정보
            with st.expander("ℹ️ Snowflake 연결 정보 & 쿼리 관리", expanded=False):
                # 연결 테스트
                conn = get_snowflake_connection()
                if conn:
                    try:
                        # 연결 정보 가져오기
                        account = conn.account
                        user = conn.user
                        database = conn.database
                        schema = conn.schema
                        warehouse = conn.warehouse
                        
                        st.success("✅ Snowflake 연결됨")
                        st.markdown(f"""
                        **연결 정보:**
                        - **Account**: {account}
                        - **User**: {user}
                        - **Database**: {database}
                        - **Schema**: {schema}
                        - **Warehouse**: {warehouse}
                        """)
                        
                        # 테이블 정보 확인
                        try:
                            cursor = conn.cursor()
                            cursor.execute("SHOW TABLES LIKE 'SALES_DATA'")
                            tables = cursor.fetchall()
                            if tables:
                                st.info("✅ sales_data 테이블 확인됨")
                            else:
                                st.warning("⚠️ sales_data 테이블을 찾을 수 없습니다.")
                            cursor.close()
                        except Exception as e:
                            st.warning(f"⚠️ 테이블 정보 확인 실패: {str(e)}")
                        
                        # 연결을 닫지 않고 유지
                        # conn.close() 제거
                        
                    except Exception as e:
                        st.error(f"❌ 연결 정보 조회 실패: {str(e)}")
                        if conn:
                            conn.close()
                else:
                    st.error("❌ Snowflake 연결 실패")
                    st.markdown("""
                    **연결 설정 방법:**
                    1. `.streamlit/secrets.toml` 파일에 다음 정보를 설정하세요:
                    ```toml
                    [snowflake]
                    account = "your_account"
                    user = "your_username"
                    password = "your_password"
                    database = "your_database"
                    schema = "your_schema"
                    warehouse = "your_warehouse"
                    ```
                    2. 필요한 테이블: `sales_data` (BRD_CD, DT, ITEM, ITEM_NM, SALE_AMT_TY, SALE_QTY_TY, SALE_AMT_LY, SALE_QTY_LY 컬럼 포함)
                    """)
                
                # 연결 테스트 버튼
                st.markdown("---")
                if st.button("🔗 Snowflake 연결 테스트", use_container_width=True, key="sales_snowflake_connection_test"):
                    conn = get_snowflake_connection()
                    if conn:
                        st.success("✅ Snowflake 연결 성공!")
                        conn.close()
                    else:
                        st.error("❌ Snowflake 연결 실패")
                
                # 쿼리 관리 섹션
                st.markdown("---")
                st.markdown("#### 🔧 쿼리 관리")
                
                col_query1, col_query2 = st.columns([1, 1])
                
                with col_query1:
                    if st.button("📋 현재 쿼리 확인", use_container_width=True, key="sales_current_query_check"):
                        st.markdown("**현재 사용 중인 Snowflake 쿼리:**")
                        
                        # 사용자 정의 쿼리가 있으면 표시, 없으면 기본 쿼리 표시
                        if hasattr(st.session_state, 'custom_query') and st.session_state.custom_query:
                            st.code(st.session_state.custom_query, language='sql')
                            st.success("✅ 사용자 정의 쿼리가 적용되어 있습니다.")
                        else:
                            # 기본 쿼리 함수에서 가져오기
                            current_query = get_default_sales_query()
                            st.code(current_query, language='sql')
                            st.info("ℹ️ 기본 쿼리가 사용되고 있습니다.")
                
                with col_query2:
                    if st.button("✏️ 쿼리 수정", use_container_width=True, key="sales_query_edit"):
                        st.session_state.show_query_editor = True
                
                # 쿼리 수정 에디터
                if st.session_state.get('show_query_editor', False):
                    st.markdown("**쿼리 수정:**")
                    st.warning("⚠️ 쿼리 수정은 고급 사용자만 사용하세요. 잘못된 쿼리는 데이터 로딩에 실패할 수 있습니다.")
                    
                    # 기본 쿼리 함수에서 가져오기
                    default_query = get_default_sales_query()
                    
                    modified_query = st.text_area(
                        "수정할 쿼리를 입력하세요:",
                        value=default_query,
                        height=400,
                        help="필수 컬럼: BRD_CD, DT, ITEM, ITEM_NM, SALE_AMT_TY, SALE_QTY_TY, SALE_AMT_LY, SALE_QTY_LY"
                    )
                    
                    col_save, col_cancel = st.columns([1, 1])
                    
                    with col_save:
                        if st.button("💾 쿼리 저장", use_container_width=True, key="sales_query_save"):
                            # 쿼리를 세션에 저장 (실제로는 파일에 저장하거나 데이터베이스에 저장)
                            st.session_state.custom_query = modified_query
                            # 기본 쿼리도 업데이트 (다음에 기본 쿼리를 사용할 때 반영됨)
                            st.session_state.default_sales_query = modified_query
                            st.success("쿼리가 저장되었습니다!")
                            st.session_state.show_query_editor = False
                            st.rerun()
                    
                    with col_cancel:
                        if st.button("❌ 취소", use_container_width=True, key="sales_query_cancel"):
                            st.session_state.show_query_editor = False
                            st.rerun()
        
        with col2:
            # Snowflake 작업 버튼
            if st.button("🔄 Snowflake에서 매출 데이터 불러오기", use_container_width=True, key="sales_snowflake_load"):
                with st.spinner("Snowflake에서 매출 데이터를 불러오는 중..."):
                    snowflake_sales_df = load_snowflake_sales_data()
                    if not snowflake_sales_df.empty:
                        # Snowflake 데이터를 로컬 파일로 저장
                        save_sales_data(snowflake_sales_df)
                        st.success(f"✅ Snowflake에서 {len(snowflake_sales_df)}건의 매출 데이터를 성공적으로 불러왔습니다!")
                        st.rerun()
                    else:
                        st.error("❌ Snowflake에서 매출 데이터를 불러올 수 없습니다.")
    
    # 엑셀 업로드 섹션
    st.markdown("#### 📊 엑셀 업로드")
    
    # 2단 레이아웃
    col1, col2 = st.columns([1, 1])
    
    with col1:
        # 매출 데이터 템플릿 다운로드
        sales_template_data = {
            'BRD_CD': ['M', 'X', 'V', 'ST'],
            'DT': ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04'],
            'ITEM': ['A001', 'A002', 'A003', 'A004'],
            'ITEM_NM': ['카테고리1', '카테고리2', '카테고리3', '카테고리4'],
            'SALE_AMT_TY': [1000000, 2000000, 1500000, 3000000],
            'SALE_QTY_TY': [100, 200, 150, 300],
            'SALE_AMT_LY': [800000, 1800000, 1200000, 2500000],
            'SALE_QTY_LY': [80, 180, 120, 250]
        }
        
        sales_template_df = pd.DataFrame(sales_template_data)
        sales_template_output = io.BytesIO()
        with pd.ExcelWriter(sales_template_output, engine='openpyxl') as writer:
            sales_template_df.to_excel(writer, index=False, sheet_name='매출데이터템플릿')
        sales_template_output.seek(0)
        
        st.download_button(
            label="📥 매출 데이터 템플릿 다운로드",
            data=sales_template_output.getvalue(),
            file_name="매출데이터_템플릿.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        # 매출 데이터 사용 가이드
        st.markdown("**필수 컬럼:**")
        st.markdown("""
        - BRD_CD: 브랜드 코드 (M=MLB, X=DX, V=DV, ST=ST)
        - DT: 날짜 (YYYY-MM-DD 형식)
        - ITEM: 카테고리 코드
        - ITEM_NM: 카테고리명
        - SALE_AMT_TY: 올해 매출액
        - SALE_QTY_TY: 올해 판매수량
        - SALE_AMT_LY: 작년 매출액
        - SALE_QTY_LY: 작년 판매수량
        """)
    
    with col2:
        # 매출 데이터 파일 업로드
        sales_uploaded_file = st.file_uploader(
            "매출 데이터 파일을 업로드하세요",
            type=['xlsx', 'csv'],
            help="Excel 또는 CSV 파일을 업로드할 수 있습니다.",
            key="sales_upload"
        )
    
    if sales_uploaded_file is not None:
        try:
            if sales_uploaded_file.name.endswith('.csv'):
                new_sales_df = pd.read_csv(sales_uploaded_file)
            else:
                new_sales_df = pd.read_excel(sales_uploaded_file)
            
            # 데이터 검증
            required_columns = ['BRD_CD', 'DT', 'ITEM', 'ITEM_NM', 'SALE_AMT_TY', 'SALE_QTY_TY', 'SALE_AMT_LY', 'SALE_QTY_LY']
            missing_columns = [col for col in required_columns if col not in new_sales_df.columns]
            
            if missing_columns:
                st.error(f"필수 컬럼이 누락되었습니다: {', '.join(missing_columns)}")
            else:
                # 데이터 저장
                save_sales_data(new_sales_df)
                st.success("매출 데이터가 성공적으로 업로드되었습니다.")
                st.rerun()
                
        except Exception as e:
            st.error(f"파일 업로드 중 오류가 발생했습니다: {str(e)}")
    
    # 검색량 데이터 업로드 섹션 추가
    st.markdown("---")
    st.markdown("### 📤 검색량 데이터 업로드")
    
    # Snowflake 연동 섹션
    if SNOWFLAKE_AVAILABLE:
        st.markdown("#### ❄️ Snowflake 연동")
        
        # 2열 레이아웃으로 나란히 배치
        col1, col2 = st.columns([1, 1])
        
        with col1:
            # Snowflake 연결 상태 및 정보
            with st.expander("ℹ️ Snowflake 연결 정보 & 쿼리 관리", expanded=False):
                # 연결 테스트
                conn = get_snowflake_connection()
                if conn:
                    try:
                        # 연결 정보 가져오기
                        account = conn.account
                        user = conn.user
                        database = conn.database
                        schema = conn.schema
                        warehouse = conn.warehouse
                        
                        st.success("✅ Snowflake 연결됨")
                        st.markdown(f"""
                        **연결 정보:**
                        - **Account**: {account}
                        - **User**: {user}
                        - **Database**: {database}
                        - **Schema**: {schema}
                        - **Warehouse**: {warehouse}
                        """)
                        
                        # 테이블 정보 확인
                        try:
                            cursor = conn.cursor()
                            cursor.execute("SHOW TABLES LIKE 'DB_srch_kwd_naver_w'")
                            tables = cursor.fetchall()
                            if tables:
                                st.info("✅ PRCS.DB_srch_kwd_naver_w 테이블 확인됨")
                            else:
                                st.warning("⚠️ PRCS.DB_srch_kwd_naver_w 테이블을 찾을 수 없습니다.")
                            cursor.close()
                        except Exception as e:
                            st.warning(f"⚠️ 테이블 정보 확인 실패: {str(e)}")
                        
                        # 연결을 닫지 않고 유지
                        # conn.close() 제거
                        
                    except Exception as e:
                        st.error(f"❌ 연결 정보 조회 실패: {str(e)}")
                        if conn:
                            conn.close()
                else:
                    st.error("❌ Snowflake 연결 실패")
                    st.markdown("""
                    **연결 설정 방법:**
                    1. `.streamlit/secrets.toml` 파일에 다음 정보를 설정하세요:
                    ```toml
                    [snowflake]
                    account = "your_account"
                    user = "your_username"
                    password = "your_password"
                    database = "your_database"
                    schema = "your_schema"
                    warehouse = "your_warehouse"
                    ```
                    2. 필요한 테이블: `PRCS.DB_srch_kwd_naver_w` (START_DT, END_DT, A-Z KWD, A-Z DVC, 123 SRCH_CNT 컬럼 포함)
                    """)
                
                # 연결 테스트 버튼
                st.markdown("---")
                if st.button("🔗 Snowflake 연결 테스트", use_container_width=True, key="search_snowflake_connection_test"):
                    conn = get_snowflake_connection()
                    if conn:
                        st.success("✅ Snowflake 연결 성공!")
                        conn.close()
                    else:
                        st.error("❌ Snowflake 연결 실패")
                
                # 쿼리 관리 섹션
                st.markdown("---")
                st.markdown("#### 🔧 쿼리 관리")
                
                col_query1, col_query2 = st.columns([1, 1])
                
                with col_query1:
                    if st.button("📋 현재 쿼리 확인", use_container_width=True, key="search_current_query_check"):
                        st.markdown("**현재 사용 중인 Snowflake 쿼리:**")
                        
                        # 사용자 정의 쿼리가 있으면 표시, 없으면 기본 쿼리 표시
                        if hasattr(st.session_state, 'custom_search_query') and st.session_state.custom_search_query:
                            st.code(st.session_state.custom_search_query, language='sql')
                            st.success("✅ 사용자 정의 쿼리가 적용되어 있습니다.")
                        else:
                            # 기본 쿼리 함수에서 가져오기
                            current_query = get_default_search_query()
                            st.code(current_query, language='sql')
                            st.info("ℹ️ 기본 쿼리가 사용되고 있습니다.")
                
                with col_query2:
                    if st.button("✏️ 쿼리 수정", use_container_width=True, key="search_query_edit"):
                        st.session_state.show_search_query_editor = True
                
                # 쿼리 수정 에디터
                if st.session_state.get('show_search_query_editor', False):
                    st.markdown("**쿼리 수정:**")
                    st.warning("⚠️ 쿼리 수정은 고급 사용자만 사용하세요. 잘못된 쿼리는 데이터 로딩에 실패할 수 있습니다.")
                    
                    # 기본 쿼리 함수에서 가져오기
                    default_query = get_default_search_query()
                    
                    # 쿼리 에디터
                    edited_query = st.text_area(
                        "SQL 쿼리를 수정하세요:",
                        value=st.session_state.get('custom_search_query', default_query),
                        height=400,
                        help="쿼리를 수정한 후 '저장' 버튼을 클릭하세요."
                    )
                    
                    # 저장/취소 버튼
                    col_save, col_cancel = st.columns([1, 1])
                    
                    with col_save:
                        if st.button("💾 저장", use_container_width=True, key="search_query_save"):
                            st.session_state.custom_search_query = edited_query
                            # 기본 쿼리도 업데이트 (다음에 기본 쿼리를 사용할 때 반영됨)
                            st.session_state.default_search_query = edited_query
                            # 파일에도 저장
                            if save_search_query(edited_query):
                                st.success("쿼리가 저장되었습니다! (파일에 영구 저장됨)")
                            st.session_state.show_search_query_editor = False
                            st.rerun()
                    
                    with col_cancel:
                        if st.button("❌ 취소", use_container_width=True, key="search_query_cancel"):
                            st.session_state.show_search_query_editor = False
                            st.rerun()
        
        with col2:
            # Snowflake 작업 버튼
            if st.button("🔄 Snowflake에서 검색량 데이터 불러오기", use_container_width=True, key="search_snowflake_load"):
                with st.spinner("Snowflake에서 검색량 데이터를 불러오는 중..."):
                    # 기존 데이터 삭제
                    if os.path.exists(SEARCH_FILE):
                        os.remove(SEARCH_FILE)
                        st.info("기존 데이터를 삭제했습니다.")
                    
                    # 전체 데이터 새로 불러오기
                    start_date_str = '2024-09-02'
                    end_date_str = datetime.now().strftime('%Y-%m-%d')
                    
                    # 현재 사용 중인 쿼리 확인 (우선순위: custom > default > 파일 > 하드코딩)
                    current_query = None
                    if hasattr(st.session_state, 'custom_search_query') and st.session_state.custom_search_query:
                        current_query = st.session_state.custom_search_query
                        st.info("✅ 사용자 정의 쿼리를 사용합니다.")
                    elif hasattr(st.session_state, 'default_search_query') and st.session_state.default_search_query:
                        current_query = st.session_state.default_search_query
                        st.info("✅ 저장된 쿼리를 사용합니다.")
                    else:
                        # 파일에서 저장된 쿼리 로드 시도
                        saved_query = load_search_query()
                        if saved_query:
                            st.session_state.default_search_query = saved_query
                            current_query = saved_query
                            st.info(f"✅ 파일에서 저장된 쿼리를 불러왔습니다. ({SEARCH_QUERY_FILE})")
                        else:
                            # 최후의 수단: 하드코딩된 기본 쿼리
                            default_query = get_default_search_query()
                            st.session_state.default_search_query = default_query
                            current_query = default_query
                            st.info("ℹ️ 기본 쿼리를 사용합니다. (쿼리 수정하기에서 저장하면 파일에 영구 저장됩니다)")
                    
                    # 쿼리 미리보기 (디버깅용)
                    with st.expander("🔍 사용 중인 쿼리 미리보기", expanded=False):
                        st.code(current_query, language='sql')
                    
                    snowflake_search_df = load_snowflake_search_data(start_date_str, end_date_str)
                    if not snowflake_search_df.empty:
                        # 새 데이터를 로컬 파일로 저장
                        save_search_data(snowflake_search_df)
                        # session_state에도 저장
                        st.session_state.search_data = snowflake_search_df
                        st.success(f"✅ Snowflake에서 {len(snowflake_search_df)}건의 검색량 데이터를 성공적으로 불러왔습니다!")
                        st.rerun()
                    else:
                        st.error("❌ Snowflake에서 검색량 데이터를 불러올 수 없습니다.")
    
    # 검색량 데이터 표시
    st.markdown("---")
    st.markdown("### 📊 검색량 데이터")
    
    # 검색량 데이터 로드
    search_df = load_search_data()
    
    if not search_df.empty:
        st.success(f"총 {len(search_df)}건의 검색량 데이터가 있습니다.")
        
        # 필터
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**🏷️ 브랜드**")
            if 'BRAND_CODE' in search_df.columns:
                brand_col = 'BRAND_CODE'
                # 브랜드 코드를 브랜드명으로 매핑
                brand_mapping = {
                    'M': 'MLB',
                    'X': 'DX', 
                    'V': 'DV',
                    'ST': 'ST'
                }
                unique_brands = search_df[brand_col].unique()
                # M, X, V, ST만 필터링하고 브랜드명으로 변환
                filtered_brands = [b for b in unique_brands if b in brand_mapping]
                brand_options = ['전체'] + [brand_mapping[b] for b in sorted(filtered_brands)]
                selected_brand = st.selectbox("브랜드를 선택하세요", brand_options, key="search_brand_filter", label_visibility="collapsed")
                # 선택된 브랜드명을 다시 코드로 변환
                if selected_brand != "전체":
                    selected_brand = next(k for k, v in brand_mapping.items() if v == selected_brand)
            elif 'BRD_CD' in search_df.columns:
                brand_col = 'BRD_CD'
                brand_mapping = {
                    'M': 'MLB',
                    'X': 'DX', 
                    'V': 'DV',
                    'ST': 'ST'
                }
                unique_brands = search_df[brand_col].unique()
                filtered_brands = [b for b in unique_brands if b in brand_mapping]
                brand_options = ['전체'] + [brand_mapping[b] for b in sorted(filtered_brands)]
                selected_brand = st.selectbox("브랜드를 선택하세요", brand_options, key="search_brand_filter", label_visibility="collapsed")
                if selected_brand != "전체":
                    selected_brand = next(k for k, v in brand_mapping.items() if v == selected_brand)
            elif 'brand_code' in search_df.columns:
                brand_col = 'brand_code'
                brand_mapping = {
                    'M': 'MLB',
                    'X': 'DX', 
                    'V': 'DV',
                    'ST': 'ST'
                }
                unique_brands = search_df[brand_col].unique()
                filtered_brands = [b for b in unique_brands if b in brand_mapping]
                brand_options = ['전체'] + [brand_mapping[b] for b in sorted(filtered_brands)]
                selected_brand = st.selectbox("브랜드를 선택하세요", brand_options, key="search_brand_filter", label_visibility="collapsed")
                if selected_brand != "전체":
                    selected_brand = next(k for k, v in brand_mapping.items() if v == selected_brand)
            else:
                st.error("브랜드 컬럼을 찾을 수 없습니다.")
                selected_brand = "전체"
        
        with col2:
            st.markdown("**📅 시작날짜**")
            if 'START_DT' in search_df.columns:
                unique_dates = sorted(search_df['START_DT'].unique())
                date_options = ['전체'] + [str(date) for date in unique_dates]
                selected_date = st.selectbox("시작날짜를 선택하세요", date_options, key="search_date_filter", label_visibility="collapsed")
            else:
                selected_date = "전체"
        
        # 데이터 필터링
        filtered_search_df = search_df.copy()
        
        # 브랜드 필터링
        if selected_brand != "전체" and brand_col in filtered_search_df.columns:
            filtered_search_df = filtered_search_df[filtered_search_df[brand_col] == selected_brand]
        
        # 날짜 필터링 (타입 변환 처리)
        if selected_date != "전체" and 'START_DT' in filtered_search_df.columns:
            # 날짜 컬럼을 문자열로 변환하여 비교
            filtered_search_df['START_DT_str'] = filtered_search_df['START_DT'].astype(str)
            filtered_search_df = filtered_search_df[filtered_search_df['START_DT_str'] == selected_date]
            filtered_search_df = filtered_search_df.drop('START_DT_str', axis=1)  # 임시 컬럼 제거
        
        # 필터링된 데이터 표시
        if not filtered_search_df.empty:
            st.markdown(f"**필터링된 데이터: {len(filtered_search_df)}건**")
            
            # 컬럼명 한국어로 변경
            display_df = filtered_search_df.copy()
            column_mapping = {
                'rank': '순위',
                'search_keyword': '검색어',
                'yoy_change_pct': '전년대비 증감률(%)',
                'period_search_cnt': '기간 검색량',
                'category': '카테고리',
                'sub_category': '하위 카테고리',
                'keyword_type': '키워드 유형',
                'brand_code': '브랜드 코드',
                'BRD_CD': '브랜드 코드',
                'CAT_NM': '카테고리',
                'SUB_CAT_NM': '하위 카테고리',
                'KWD_TYPE': '키워드 유형',
                'START_DT': '시작날짜',
                'END_DT': '종료날짜'
            }
            
            # 실제 컬럼명에 맞게 매핑
            for old_col, new_col in column_mapping.items():
                if old_col in display_df.columns:
                    display_df = display_df.rename(columns={old_col: new_col})
            
            st.dataframe(display_df, use_container_width=True)
            
            # 다운로드 및 삭제 버튼
            col_download, col_delete = st.columns([1, 1])
            
            with col_download:
                if st.button("📥 검색량 데이터 엑셀 다운로드", use_container_width=True):
                    # 엑셀 파일로 다운로드
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        display_df.to_excel(writer, sheet_name='검색량데이터', index=False)
                    excel_buffer.seek(0)
                    
                    st.download_button(
                        label="📥 엑셀 파일 다운로드",
                        data=excel_buffer.getvalue(),
                        file_name=f"검색량데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            
            with col_delete:
                if st.button("🗑️ 검색량 데이터 삭제", use_container_width=True):
                    if os.path.exists(SEARCH_FILE):
                        os.remove(SEARCH_FILE)
                        st.success("✅ 검색량 데이터가 삭제되었습니다!")
                        st.rerun()
                    else:
                        st.info("삭제할 검색량 데이터가 없습니다.")
        else:
            st.warning("선택된 조건에 맞는 데이터가 없습니다.")
    else:
        st.info("검색량 데이터가 없습니다. 위의 Snowflake 연동을 통해 데이터를 불러와주세요.")

# =============================================================================
# 매출 데이터 관리 관련 함수들
# =============================================================================

def render_sales_management_tab():
    """매출 데이터 관리 탭 렌더링"""
    st.markdown("# 💰 매출 데이터 관리")
    
    # 매출 데이터 표시
    sales_df = load_sales_data()
    if not sales_df.empty:
        st.markdown("## 📊 매출 데이터")
        st.success(f"총 {len(sales_df)}건의 매출 데이터가 있습니다.")
        
        # 필터
        st.markdown("#### 🔍 매출 데이터 필터")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("**🏷️ 브랜드**")
            if 'BRD_CD' in sales_df.columns:
                valid_brands = ['M', 'X', 'V', 'ST']
                sales_df_filtered = sales_df[sales_df['BRD_CD'].isin(valid_brands)]
                brand_mapping = {'M': 'MLB', 'X': 'DX', 'V': 'DV', 'ST': 'ST'}
                unique_brands = sales_df_filtered['BRD_CD'].unique()
                brand_names = [brand_mapping.get(brand, brand) for brand in unique_brands]
                brand_order = ['MLB', 'DX', 'DV', 'ST']
                ordered_brands = [brand for brand in brand_order if brand in brand_names]
                selected_brd = st.selectbox("브랜드를 선택하세요", ordered_brands, key="sales_table_brand_filter", label_visibility="collapsed")
            else:
                selected_brd = "MLB"
        
        with col2:
            st.markdown("**📅 월**")
            if 'DT' in sales_df.columns:
                sales_df['월'] = pd.to_datetime(sales_df['DT']).dt.to_period('M')
                months = sorted(sales_df['월'].unique())
                selected_month = st.selectbox("월을 선택하세요", months, key="sales_table_month_filter", label_visibility="collapsed")
            else:
                selected_month = None
        
        with col3:
            st.markdown("**📦 카테고리**")
            if 'ITEM_NM' in sales_df.columns:
                categories = sales_df['ITEM_NM'].unique()
                selected_item = st.selectbox("카테고리를 선택하세요", ["전체"] + list(categories), key="sales_table_item_filter", label_visibility="collapsed")
            else:
                selected_item = "전체"
        
        # 데이터 필터링
        filtered_sales_df = sales_df.copy()
        if 'BRD_CD' in sales_df.columns:
            brand_reverse_mapping = {'MLB': 'M', 'DX': 'X', 'DV': 'V', 'ST': 'ST'}
            brand_code = brand_reverse_mapping.get(selected_brd, selected_brd)
            filtered_sales_df = filtered_sales_df[filtered_sales_df['BRD_CD'] == brand_code]
        
        if selected_month and '월' in sales_df.columns:
            filtered_sales_df = filtered_sales_df[filtered_sales_df['월'] == selected_month]
        
        if selected_item != "전체" and 'ITEM_NM' in sales_df.columns:
            filtered_sales_df = filtered_sales_df[filtered_sales_df['ITEM_NM'] == selected_item]
        
        # 표시용 데이터 준비
        display_df = filtered_sales_df.copy()
        if 'DT' in display_df.columns:
            display_df['DT'] = pd.to_datetime(display_df['DT'], errors='coerce').dt.date
        
        # 컬럼 순서 조정
        column_order = ['BRD_CD', 'DT', 'ITEM', 'ITEM_NM', 'SALE_AMT_TY', 'SALE_QTY_TY', 'SALE_AMT_LY', 'SALE_QTY_LY']
        available_columns = [col for col in column_order if col in display_df.columns]
        display_df = display_df[available_columns]
        
        st.dataframe(display_df, use_container_width=True)
        
        # 관리 옵션
        st.markdown("## ⚙️ 매출 데이터 관리")
        col1, col2 = st.columns(2)
        
        with col1:
            # 엑셀 파일 생성
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                display_df.to_excel(writer, index=False, sheet_name='매출데이터')
            output.seek(0)
            
            st.download_button(
                label="📥 엑셀 다운로드",
                data=output.getvalue(),
                file_name=f"매출데이터_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            if st.button("🗑️ 기존 데이터 삭제"):
                if os.path.exists(SALES_FILE):
                    os.remove(SALES_FILE)
                    st.success("매출 데이터가 삭제되었습니다.")
                    st.rerun()
    

# =============================================================================
# Snowflake 연결 기능
# =============================================================================

def get_snowflake_connection():
    """Snowflake 연결 설정 - 매번 새로운 연결 생성"""
    if not SNOWFLAKE_AVAILABLE:
        st.error("Snowflake 패키지가 설치되지 않았습니다.")
        return None
    
    try:
        # Streamlit secrets에서 Snowflake 설정 가져오기
        conn = snowflake.connector.connect(
            user=st.secrets["snowflake"]["user"],
            password=st.secrets["snowflake"]["password"],
            account=st.secrets["snowflake"]["account"],
            warehouse=st.secrets["snowflake"]["warehouse"],
            database=st.secrets["snowflake"]["database"],
            schema=st.secrets["snowflake"]["schema"]
        )
        return conn
    except Exception as e:
        st.error(f"Snowflake 연결 실패: {str(e)}")
        return None

def execute_snowflake_query(query):
    """Snowflake 쿼리 실행"""
    if not SNOWFLAKE_AVAILABLE:
        st.error("Snowflake 패키지가 설치되지 않았습니다.")
        return pd.DataFrame()
    
    conn = None
    cursor = None
    try:
        # 새로운 연결 생성
        conn = snowflake.connector.connect(
            user=st.secrets["snowflake"]["user"],
            password=st.secrets["snowflake"]["password"],
            account=st.secrets["snowflake"]["account"],
            warehouse=st.secrets["snowflake"]["warehouse"],
            database=st.secrets["snowflake"]["database"],
            schema=st.secrets["snowflake"]["schema"]
        )
        
        cursor = conn.cursor()
        cursor.execute(query)
        results = cursor.fetchall()
        
        # 컬럼명 가져오기
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(results, columns=columns)
        
        # 날짜 데이터 검증 및 수정
        if 'DT' in df.columns and not df.empty:
            df['DT'] = pd.to_datetime(df['DT'])
            current_year = pd.Timestamp.now().year
            
            # 잘못된 날짜 필터링 (현재 연도 + 1년까지만 허용)
            df = df[df['DT'].dt.year <= current_year + 1]
            df = df[df['DT'].dt.year >= 1900]
            
            # 잘못된 날짜 데이터는 조용히 제거 (경고 메시지 제거)
        
        return df
    except Exception as e:
        st.error(f"쿼리 실행 실패: {str(e)}")
        return pd.DataFrame()
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def load_snowflake_influencer_data():
    """Snowflake에서 인플루언서 데이터 로드"""
    query = """
    SELECT 
        contract_id,
        contract_sesn,
        sns_id,
        name,
        gender,
        follower,
        agency,
        mlb_qty,
        dx_qty,
        dv_qty,
        st_qty,
        total_qty,
        total_amt_incl2nd,
        total_amt_exc2nd,
        sec_usage,
        sec_period,
        sec_commercial,
        sec_sns,
        sec_ads,
        unit_fee
    FROM influencer_data
    ORDER BY contract_id
    """
    return execute_snowflake_query(query)

def load_snowflake_execution_data():
    """Snowflake에서 집행 데이터 로드"""
    query = """
    SELECT 
        날짜,
        인플루언서,
        노출수,
        좋아요,
        댓글수,
        조회수
    FROM execution_data
    ORDER BY 날짜 DESC
    """
    return execute_snowflake_query(query)

def get_default_sales_query():
    """기본 매출 쿼리 반환"""
    # session_state에 저장된 기본 쿼리가 있으면 사용, 없으면 하드코딩된 기본 쿼리 사용
    if hasattr(st.session_state, 'default_sales_query') and st.session_state.default_sales_query:
        return st.session_state.default_sales_query
    
    # 하드코딩된 기본 쿼리
    return """
    WITH TY AS (
        SELECT  
            b.brd_cd,
            a.dt,
            b.item,
            b.item_nm,
            SUM(a.sale_nml_sale_amt_cns + a.sale_ret_sale_amt_cns) AS sale_amt_ty,
            SUM(a.sale_nml_qty_cns + a.sale_ret_qty_cns) AS sale_qty_ty
        FROM prcs.dw_scs_d a
        JOIN prcs.db_prdt b ON a.prdt_cd = b.prdt_cd 
        WHERE 1=1
          AND a.dt >= DATE '2025-09-01'   -- 2025년 9월 이후 최신 데이터
        GROUP BY b.brd_cd, a.dt, b.item, b.item_nm
    ),
    LY AS (
        SELECT  
            b.brd_cd,
            DATEADD(year, 1, a.dt) AS dt,   -- 작년 → 올해 날짜로 이동
            b.item,
            b.item_nm,
            SUM(a.sale_nml_sale_amt_cns + a.sale_ret_sale_amt_cns) AS sale_amt_ly,
            SUM(a.sale_nml_qty_cns + a.sale_ret_qty_cns) AS sale_qty_ly
        FROM prcs.dw_scs_d a
        JOIN prcs.db_prdt b ON a.prdt_cd = b.prdt_cd 
        WHERE 1=1
          AND a.dt >= DATEADD(year, -1, DATE '2025-09-01')  -- 2024-09-01 이후
        GROUP BY b.brd_cd, a.dt, b.item, b.item_nm
    )
    SELECT 
        t.brd_cd,
        t.dt,
        t.item,
        t.item_nm,
        t.sale_amt_ty,
        t.sale_qty_ty,
        l.sale_amt_ly,
        l.sale_qty_ly
    FROM TY t
    LEFT JOIN LY l 
        ON t.dt = l.dt 
       AND t.item = l.item 
       AND t.brd_cd = l.brd_cd
    WHERE 1=1
      AND t.brd_cd IN ('M', 'X', 'V', 'ST')
    ORDER BY t.brd_cd, t.dt DESC, t.item
    """

def load_snowflake_sales_data():
    """Snowflake에서 매출 데이터 로드"""
    # 사용자가 수정한 쿼리가 있으면 사용, 없으면 기본 쿼리 사용
    if hasattr(st.session_state, 'custom_query') and st.session_state.custom_query:
        query = st.session_state.custom_query
    else:
        # 기본 쿼리 함수에서 가져오기
        query = get_default_sales_query()
    return execute_snowflake_query(query)

def load_snowflake_assignment_data():
    """Snowflake에서 배정 데이터 로드"""
    query = """
    SELECT 
        contract_id,
        brand,
        month,
        assigned_qty,
        season,
        assignment_date
    FROM assignment_history
    ORDER BY assignment_date DESC
    """
    return execute_snowflake_query(query)

# =============================================================================
# 데이터 문의 기능
# =============================================================================

def analyze_data_question(question, execution_df, influencer_df):
    """AI 기반 데이터 문의 분석 - 판다스 기반 자연어 처리"""
    question_lower = question.lower()
    
    # 모든 데이터 로드
    sales_df = load_sales_data()
    assignment_df = load_assignment_history()
    
    # 데이터 통합 분석
    try:
        # 1. 브랜드별 성과 분석
        if any(keyword in question_lower for keyword in ["성과", "실적", "결과", "효과"]):
            return analyze_brand_performance(execution_df, sales_df, influencer_df, question_lower)
        
        # 2. 인플루언서별 분석
        if any(keyword in question_lower for keyword in ["인플루언서", "크리에이터", "인플"]):
            return analyze_influencer_performance(execution_df, influencer_df, question_lower)
        
        # 3. 매출 분석
        if any(keyword in question_lower for keyword in ["매출", "판매", "수익", "매출액"]):
            return analyze_sales_data(sales_df, question_lower)
        
        # 4. 집행 데이터 분석
        if any(keyword in question_lower for keyword in ["집행", "실행", "노출", "좋아요", "댓글", "조회"]):
            return analyze_execution_data(execution_df, question_lower)
        
        # 5. 배정 분석
        if any(keyword in question_lower for keyword in ["배정", "할당", "계획"]):
            return analyze_assignment_data(assignment_df, influencer_df, question_lower)
        
        # 6. 트렌드 분석
        if any(keyword in question_lower for keyword in ["트렌드", "추이", "변화", "증가", "감소"]):
            return analyze_trends(execution_df, sales_df, question_lower)
        
        # 7. 비교 분석
        if any(keyword in question_lower for keyword in ["비교", "vs", "대비", "차이"]):
            return analyze_comparison(execution_df, sales_df, influencer_df, question_lower)
        
        # 8. 고급 분석 (새로 추가)
        if any(keyword in question_lower for keyword in ["상관관계", "연관성", "패턴", "분포", "통계"]):
            return analyze_advanced_statistics(execution_df, sales_df, influencer_df, question_lower)
        
        # 9. 예측 분석
        if any(keyword in question_lower for keyword in ["예측", "전망", "향후", "미래"]):
            return analyze_predictions(execution_df, sales_df, question_lower)
        
        # 10. 인사이트 분석
        if any(keyword in question_lower for keyword in ["인사이트", "인사이트", "발견", "특징"]):
            return analyze_insights(execution_df, sales_df, influencer_df, question_lower)
        
    except Exception as e:
        return f"분석 중 오류가 발생했습니다: {str(e)}"
    
    # 기존 단순 매핑 방식 (fallback)
    return analyze_simple_questions(question, execution_df, influencer_df, sales_df)

def analyze_brand_performance(execution_df, sales_df, influencer_df, question_lower):
    """브랜드별 성과 분석"""
    results = []
    
    # 브랜드 매핑
    brand_mapping = {'mlb': 'M', 'dx': 'X', 'dv': 'V', 'st': 'ST'}
    
    for brand_name, brand_code in brand_mapping.items():
        if brand_name in question_lower:
            # 집행 데이터 분석
            if not execution_df.empty and '브랜드' in execution_df.columns:
                brand_execution = execution_df[execution_df['브랜드'] == brand_name.upper()]
                if not brand_execution.empty:
                    total_exposure = brand_execution['노출수'].sum() if '노출수' in brand_execution.columns else 0
                    total_likes = brand_execution['좋아요'].sum() if '좋아요' in brand_execution.columns else 0
                    total_comments = brand_execution['댓글수'].sum() if '댓글수' in brand_execution.columns else 0
                    total_views = brand_execution['조회수'].sum() if '조회수' in brand_execution.columns else 0
                    
                    results.append(f"**{brand_name.upper()} 브랜드 성과:**")
                    results.append(f"• 총 노출수: {total_exposure:,}")
                    results.append(f"• 총 좋아요: {total_likes:,}")
                    results.append(f"• 총 댓글: {total_comments:,}")
                    results.append(f"• 총 조회수: {total_views:,}")
            
            # 매출 데이터 분석
            if not sales_df.empty and 'BRD_CD' in sales_df.columns:
                brand_sales = sales_df[sales_df['BRD_CD'] == brand_code]
                if not brand_sales.empty:
                    total_sales = brand_sales['SALE_AMT_TY'].sum() if 'SALE_AMT_TY' in brand_sales.columns else 0
                    total_qty = brand_sales['SALE_QTY_TY'].sum() if 'SALE_QTY_TY' in brand_sales.columns else 0
                    results.append(f"• 총 매출액: {total_sales:,.0f}원")
                    results.append(f"• 총 판매량: {total_qty:,}개")
            
            # 인플루언서 계약 분석
            if not influencer_df.empty:
                brand_qty_col = f"{brand_name}_qty"
                if brand_qty_col in influencer_df.columns:
                    total_contracts = influencer_df[brand_qty_col].sum()
                    active_influencers = len(influencer_df[influencer_df[brand_qty_col] > 0])
                    results.append(f"• 총 계약수: {total_contracts}건")
                    results.append(f"• 활성 인플루언서: {active_influencers}명")
            
            return "\n".join(results) if results else f"{brand_name.upper()} 브랜드 데이터를 찾을 수 없습니다."
    
    return "브랜드 성과 분석을 위해 구체적인 브랜드명을 포함해주세요."

def analyze_influencer_performance(execution_df, influencer_df, question_lower):
    """인플루언서별 성과 분석"""
    if execution_df.empty:
        return "집행 데이터가 없습니다."
    
    # 최고 성과 인플루언서 찾기
    if "최고" in question_lower or "1위" in question_lower or "최대" in question_lower:
        if "노출수" in question_lower:
            top_influencer = execution_df.loc[execution_df['노출수'].idxmax()]
            return f"최고 노출수: {top_influencer['인플루언서']} ({top_influencer['노출수']:,})"
        elif "좋아요" in question_lower:
            top_influencer = execution_df.loc[execution_df['좋아요'].idxmax()]
            return f"최고 좋아요: {top_influencer['인플루언서']} ({top_influencer['좋아요']:,})"
        elif "댓글" in question_lower:
            top_influencer = execution_df.loc[execution_df['댓글수'].idxmax()]
            return f"최고 댓글수: {top_influencer['인플루언서']} ({top_influencer['댓글수']:,})"
        elif "조회수" in question_lower:
            top_influencer = execution_df.loc[execution_df['조회수'].idxmax()]
            return f"최고 조회수: {top_influencer['인플루언서']} ({top_influencer['조회수']:,})"
    
    # 인플루언서 수
    if "몇명" in question_lower or "수" in question_lower:
        unique_influencers = execution_df['인플루언서'].nunique()
        return f"총 {unique_influencers}명의 인플루언서가 활동했습니다."
    
    return "인플루언서 성과 분석을 위해 구체적인 지표를 포함해주세요."

def analyze_sales_data(sales_df, question_lower):
    """매출 데이터 분석"""
    if sales_df.empty:
        return "매출 데이터가 없습니다."
    
    # 브랜드별 매출 분석
    brand_mapping = {'mlb': 'M', 'dx': 'X', 'dv': 'V', 'st': 'ST'}
    
    for brand_name, brand_code in brand_mapping.items():
        if brand_name in question_lower:
            brand_sales = sales_df[sales_df['BRD_CD'] == brand_code]
            if not brand_sales.empty:
                # 월별 필터링 처리
                target_month = None
                month_mapping = {
                    '1월': 1, '2월': 2, '3월': 3, '4월': 4, '5월': 5, '6월': 6,
                    '7월': 7, '8월': 8, '9월': 9, '10월': 10, '11월': 11, '12월': 12
                }
                
                for month_name, month_num in month_mapping.items():
                    if month_name in question_lower:
                        target_month = month_num
                        break
                
                # 날짜 컬럼을 datetime으로 변환
                brand_sales['DT'] = pd.to_datetime(brand_sales['DT'])
                
                # 잘못된 날짜 필터링 (2025년 이후의 현실적인 날짜만)
                current_year = pd.Timestamp.now().year
                brand_sales = brand_sales[brand_sales['DT'].dt.year <= current_year + 1]  # 내년까지만 허용
                
                # 특정 월 필터링
                if target_month:
                    brand_sales = brand_sales[brand_sales['DT'].dt.month == target_month]
                
                if not brand_sales.empty:
                    # 마지막 일자 처리
                    if "마지막" in question_lower or "최신" in question_lower or "최근" in question_lower:
                        latest_date = brand_sales['DT'].max()
                        latest_sales = brand_sales[brand_sales['DT'] == latest_date]
                    else:
                        # 월별 전체 데이터 분석 (마지막 일자가 아닌 경우)
                        total_sales = brand_sales['SALE_AMT_TY'].sum()
                        total_qty = brand_sales['SALE_QTY_TY'].sum()
                        avg_sales = brand_sales['SALE_AMT_TY'].mean()
                        
                        month_name = f"{target_month}월" if target_month else "전체"
                        result = f"**{brand_name.upper()} {month_name} 매출 분석:**\n"
                        result += f"• 총 매출액: {total_sales:,.0f}원\n"
                        result += f"• 총 판매량: {total_qty:,}개\n"
                        result += f"• 평균 매출액: {avg_sales:,.0f}원\n"
                        
                        # 카테고리별 분석
                        if "카테고리" in question_lower or "상품" in question_lower:
                            category_sales = brand_sales.groupby('ITEM_NM').agg({
                                'SALE_AMT_TY': 'sum',
                                'SALE_QTY_TY': 'sum'
                            }).sort_values('SALE_AMT_TY', ascending=False)
                            
                            result += "\n**카테고리별 매출:**\n"
                            for item, row in category_sales.head(5).iterrows():
                                result += f"• {item}: {row['SALE_AMT_TY']:,.0f}원 ({row['SALE_QTY_TY']:,}개)\n"
                        
                        return result
                    
                    if not latest_sales.empty:
                        total_sales = latest_sales['SALE_AMT_TY'].sum()
                        total_qty = latest_sales['SALE_QTY_TY'].sum()
                        avg_sales = latest_sales['SALE_AMT_TY'].mean()
                        
                        result = f"**{brand_name.upper()} {latest_date.strftime('%Y-%m-%d')} 매출 분석:**\n"
                        result += f"• 총 매출액: {total_sales:,.0f}원\n"
                        result += f"• 총 판매량: {total_qty:,}개\n"
                        result += f"• 평균 매출액: {avg_sales:,.0f}원\n"
                        
                        # 카테고리별 분석
                        if "카테고리" in question_lower or "상품" in question_lower:
                            category_sales = latest_sales.groupby('ITEM_NM').agg({
                                'SALE_AMT_TY': 'sum',
                                'SALE_QTY_TY': 'sum'
                            }).sort_values('SALE_AMT_TY', ascending=False)
                            
                            result += "\n**카테고리별 매출:**\n"
                            for item, row in category_sales.head(5).iterrows():
                                result += f"• {item}: {row['SALE_AMT_TY']:,.0f}원 ({row['SALE_QTY_TY']:,}개)\n"
                        
                        return result
                    else:
                        return f"{brand_name.upper()}의 마지막 일자 데이터가 없습니다."
                else:
                    month_name = f"{target_month}월" if target_month else "해당 기간"
                    return f"{brand_name.upper()}의 {month_name} 데이터가 없습니다."
                
                # 일반적인 전체 기간 분석
                total_sales = brand_sales['SALE_AMT_TY'].sum()
                total_qty = brand_sales['SALE_QTY_TY'].sum()
                avg_sales = brand_sales['SALE_AMT_TY'].mean()
                
                result = f"**{brand_name.upper()} 매출 분석:**\n"
                result += f"• 총 매출액: {total_sales:,.0f}원\n"
                result += f"• 총 판매량: {total_qty:,}개\n"
                result += f"• 평균 매출액: {avg_sales:,.0f}원\n"
                
                # 카테고리별 분석
                if "카테고리" in question_lower or "상품" in question_lower:
                    category_sales = brand_sales.groupby('ITEM_NM').agg({
                        'SALE_AMT_TY': 'sum',
                        'SALE_QTY_TY': 'sum'
                    }).sort_values('SALE_AMT_TY', ascending=False)
                    
                    result += "\n**카테고리별 매출:**\n"
                    for item, row in category_sales.head(5).iterrows():
                        result += f"• {item}: {row['SALE_AMT_TY']:,.0f}원 ({row['SALE_QTY_TY']:,}개)\n"
                
                return result
    
    # 전체 매출 분석
    if "전체" in question_lower or "총" in question_lower:
        total_sales = sales_df['SALE_AMT_TY'].sum()
        total_qty = sales_df['SALE_QTY_TY'].sum()
        return f"전체 매출: {total_sales:,.0f}원 ({total_qty:,}개)"
    
    return "매출 분석을 위해 구체적인 브랜드명을 포함해주세요."

def analyze_execution_data(execution_df, question_lower):
    """집행 데이터 분석"""
    if execution_df.empty:
        return "집행 데이터가 없습니다."
    
    # 메트릭별 분석
    metrics = {
        '노출수': '노출수',
        '좋아요': '좋아요',
        '댓글': '댓글수',
        '조회수': '조회수'
    }
    
    for metric_name, column in metrics.items():
        if metric_name in question_lower and column in execution_df.columns:
            if "최고" in question_lower or "최대" in question_lower:
                max_value = execution_df[column].max()
                max_row = execution_df[execution_df[column] == max_value].iloc[0]
                return f"최고 {metric_name}: {max_row['인플루언서']} ({max_value:,})"
            elif "평균" in question_lower:
                avg_value = execution_df[column].mean()
                return f"평균 {metric_name}: {avg_value:,.0f}"
            elif "총" in question_lower or "합계" in question_lower:
                total_value = execution_df[column].sum()
                return f"총 {metric_name}: {total_value:,}"
    
    return "집행 데이터 분석을 위해 구체적인 지표를 포함해주세요."

def analyze_assignment_data(assignment_df, influencer_df, question_lower):
    """배정 데이터 분석"""
    if assignment_df.empty:
        return "배정 데이터가 없습니다."
    
    # 브랜드별 배정 분석
    if "브랜드" in question_lower:
        brand_assignments = assignment_df.groupby('brand')['assigned_qty'].sum()
        result = "**브랜드별 배정 현황:**\n"
        for brand, qty in brand_assignments.items():
            result += f"• {brand}: {qty}건\n"
        return result
    
    # 월별 배정 분석
    if "월별" in question_lower or "월" in question_lower:
        monthly_assignments = assignment_df.groupby('month')['assigned_qty'].sum()
        result = "**월별 배정 현황:**\n"
        for month, qty in monthly_assignments.items():
            result += f"• {month}월: {qty}건\n"
        return result
    
    return "배정 데이터 분석을 위해 구체적인 조건을 포함해주세요."

def analyze_trends(execution_df, sales_df, question_lower):
    """트렌드 분석"""
    results = []
    
    # 집행 데이터 트렌드
    if not execution_df.empty and '날짜' in execution_df.columns:
        execution_df['날짜'] = pd.to_datetime(execution_df['날짜'])
        daily_metrics = execution_df.groupby(execution_df['날짜'].dt.date).agg({
            '노출수': 'sum',
            '좋아요': 'sum',
            '댓글수': 'sum',
            '조회수': 'sum'
        })
        
        # 최근 7일 트렌드
        recent_7days = daily_metrics.tail(7)
        if len(recent_7days) >= 2:
            exposure_trend = (recent_7days['노출수'].iloc[-1] - recent_7days['노출수'].iloc[0]) / recent_7days['노출수'].iloc[0] * 100
            results.append(f"최근 7일 노출수 변화: {exposure_trend:+.1f}%")
    
    # 매출 데이터 트렌드
    if not sales_df.empty and 'DT' in sales_df.columns:
        sales_df['DT'] = pd.to_datetime(sales_df['DT'])
        daily_sales = sales_df.groupby(sales_df['DT'].dt.date)['SALE_AMT_TY'].sum()
        
        if len(daily_sales) >= 2:
            sales_trend = (daily_sales.iloc[-1] - daily_sales.iloc[0]) / daily_sales.iloc[0] * 100
            results.append(f"최근 매출 변화: {sales_trend:+.1f}%")
    
    return "\n".join(results) if results else "트렌드 분석을 위한 충분한 데이터가 없습니다."

def analyze_comparison(execution_df, sales_df, influencer_df, question_lower):
    """비교 분석"""
    if "브랜드" in question_lower:
        # 브랜드별 성과 비교
        brand_performance = {}
        
        if not execution_df.empty and '브랜드' in execution_df.columns:
            for brand in execution_df['브랜드'].unique():
                brand_data = execution_df[execution_df['브랜드'] == brand]
                brand_performance[brand] = {
                    '노출수': brand_data['노출수'].sum(),
                    '좋아요': brand_data['좋아요'].sum(),
                    '댓글수': brand_data['댓글수'].sum(),
                    '조회수': brand_data['조회수'].sum()
                }
        
        result = "**브랜드별 성과 비교:**\n"
        for brand, metrics in brand_performance.items():
            result += f"**{brand}:**\n"
            for metric, value in metrics.items():
                result += f"  • {metric}: {value:,}\n"
        
        return result
    
    return "비교 분석을 위해 구체적인 비교 대상을 포함해주세요."

def analyze_simple_questions(question, execution_df, influencer_df, sales_df):
    """기존 단순 질문 처리 (fallback)"""
    question_lower = question.lower()
    
    # 인플루언서 데이터 관련 질문 처리
    if not influencer_df.empty:
        # 브랜드별 계약수 질문 (동적 처리)
        brand_mapping = {
            'dv': 'dv_qty',
            'mlb': 'mlb_qty', 
            'dx': 'dx_qty',
            'st': 'st_qty'
        }
        
        for brand, column in brand_mapping.items():
            if brand in question_lower and ("계약수" in question or "계약" in question):
                if "총" in question or "합계" in question or "몇개" in question:
                    total = int(influencer_df[column].sum())
                    return f"{brand.upper()} 총 계약수: {total}건"
                elif "인플루언서" in question or "명" in question:
                    count = len(influencer_df[influencer_df[column] > 0])
                    return f"{brand.upper()} 계약이 있는 인플루언서: {count}명"
        
        # 전체 계약수 관련 질문
        if "전체" in question and ("계약수" in question or "계약" in question):
            if "총" in question or "합계" in question:
                total_contracts = int(influencer_df['total_qty'].sum())
                return f"전체 총 계약수: {total_contracts}건"
            elif "인플루언서" in question or "명" in question:
                total_count = len(influencer_df)
                return f"전체 인플루언서 수: {total_count}명"
    
    # execution 데이터 관련 질문 처리 (일반화)
    if not execution_df.empty:
        # 메트릭 매핑
        metric_mapping = {
            '노출수': '노출수',
            '좋아요': '좋아요', 
            '댓글': '댓글수',
            '조회수': '조회수'
        }
        
        for metric_name, column in metric_mapping.items():
            if metric_name in question_lower and column in execution_df.columns:
                if "최고" in question_lower or "최대" in question_lower or "높은" in question_lower:
                    if "일자" in question_lower or "날짜" in question_lower or "언제" in question_lower:
                        max_value = execution_df[column].max()
                        max_date = execution_df[execution_df[column] == max_value]['날짜'].iloc[0]
                        return f"{metric_name}이 가장 높은 날짜: {max_date} ({max_value:,}개)"
                    else:
                        max_value = execution_df[column].max()
                        max_influencer = execution_df[execution_df[column] == max_value]['인플루언서'].iloc[0]
                        return f"최고 {metric_name}: {max_value:,} (인플루언서: {max_influencer})"
                elif "평균" in question_lower:
                    avg_value = execution_df[column].mean()
                    return f"평균 {metric_name}: {avg_value:,.0f}"
                elif "총" in question_lower or "합계" in question_lower:
                    total_value = execution_df[column].sum()
                    return f"총 {metric_name}: {total_value:,}"
    
    # 매출 관련 질문 - 일반적이고 유연한 처리
    if "매출" in question_lower or "판매" in question_lower:
        if sales_df.empty:
            return "매출 데이터가 없습니다. '📈 데이터 업로드 관리' → '💰 매출 데이터 관리'에서 Snowflake에서 데이터를 불러와주세요."
        
        # 브랜드 매핑
        brand_mapping = {
            'mlb': 'M',
            'dx': 'X', 
            'dv': 'V',
            'st': 'ST'
        }
        
        # 질문에서 브랜드 추출
        detected_brand = None
        brand_code = None
        for brand, code in brand_mapping.items():
            if brand in question_lower:
                detected_brand = brand.upper()
                brand_code = code
                break
        
        # 월 추출 (1월~12월)
        month_mapping = {
            '1월': 1, '2월': 2, '3월': 3, '4월': 4, '5월': 5, '6월': 6,
            '7월': 7, '8월': 8, '9월': 9, '10월': 10, '11월': 11, '12월': 12
        }
        
        detected_month = None
        month_num = None
        for month, num in month_mapping.items():
            if month in question:
                detected_month = month
                month_num = num
                break
        
        # 특정 월의 브랜드 매출 처리
        if detected_month and detected_brand and brand_code:
            if 'BRD_CD' in sales_df.columns:
                brand_sales = sales_df[sales_df['BRD_CD'] == brand_code]
                if not brand_sales.empty:
                    # 날짜 컬럼을 datetime으로 변환
                    brand_sales['DT'] = pd.to_datetime(brand_sales['DT'])
                    
                    # 해당 월 데이터 필터링
                    month_sales = brand_sales[brand_sales['DT'].dt.month == month_num]
                    
                    if not month_sales.empty:
                        # 카테고리별 매출
                        if "카테고리" in question_lower or "별" in question_lower:
                            result = f"{detected_brand} {detected_month} 카테고리별 매출:\n"
                            category_sales = month_sales.groupby('ITEM_NM').agg({
                                'SALE_AMT_TY': 'sum',
                                'SALE_QTY_TY': 'sum'
                            }).sort_values('SALE_AMT_TY', ascending=False)
                            
                            for item_nm, row in category_sales.iterrows():
                                sale_amt = row['SALE_AMT_TY']
                                sale_qty = row['SALE_QTY_TY']
                                result += f"• {item_nm}: {sale_amt:,.0f}원 ({sale_qty:,}개)\n"
                            return result.strip()
                        else:
                            # 총 매출
                            total_amt = month_sales['SALE_AMT_TY'].sum()
                            total_qty = month_sales['SALE_QTY_TY'].sum()
                            return f"{detected_brand} {detected_month} 총매출: {total_amt:,.0f}원 ({total_qty:,}개)"
                    else:
                        return f"{detected_brand} {detected_month} 데이터가 없습니다."
                else:
                    return f"{detected_brand} 매출 데이터가 없습니다."
            else:
                return "브랜드 코드 컬럼을 찾을 수 없습니다."
        
        # 최신 날짜의 특정 브랜드 매출 (일반화)
        elif ("최신" in question_lower or "최신날짜" in question_lower) and detected_brand and brand_code:
            if 'BRD_CD' in sales_df.columns:
                brand_sales = sales_df[sales_df['BRD_CD'] == brand_code]
                if not brand_sales.empty:
                    # 최신 날짜 찾기
                    latest_date = brand_sales['DT'].max()
                    latest_data = brand_sales[brand_sales['DT'] == latest_date]
                    
                    if not latest_data.empty:
                        # 총매출 vs 카테고리별 매출 구분
                        if "총매출" in question_lower or "총" in question_lower:
                            total_amt = latest_data['SALE_AMT_TY'].sum()
                            total_qty = latest_data['SALE_QTY_TY'].sum()
                            return f"{detected_brand} 최신 날짜 ({latest_date}) 총매출: {total_amt:,.0f}원 ({total_qty:,}개)"
                        else:
                            result = f"{detected_brand} 최신 날짜 ({latest_date}) 카테고리별 매출:\n"
                            for _, row in latest_data.iterrows():
                                item_nm = row.get('ITEM_NM', '알 수 없음')
                                sale_amt = row.get('SALE_AMT_TY', 0)
                                sale_qty = row.get('SALE_QTY_TY', 0)
                                result += f"• {item_nm}: {sale_amt:,.0f}원 ({sale_qty:,}개)\n"
                            return result.strip()
                    else:
                        return f"{detected_brand} 최신 날짜 ({latest_date}) 데이터가 없습니다."
                else:
                    return f"{detected_brand} 매출 데이터가 없습니다."
            else:
                return "브랜드 코드 컬럼을 찾을 수 없습니다."
        
        elif "총" in question_lower or "합계" in question_lower:
            total_sales = sales_df['SALE_AMT_TY'].sum() if 'SALE_AMT_TY' in sales_df.columns else 0
            return f"총 매출액: {total_sales:,.0f}원"
        elif "카테고리" in question_lower or "상품" in question_lower:
            if "순위" in question_lower or "랭킹" in question_lower or "상위" in question_lower:
                if 'ITEM_NM' in sales_df.columns and 'SALE_AMT_TY' in sales_df.columns:
                    # 브랜드 필터링 (DX 브랜드만)
                    if "DX" in question:
                        if 'BRD_CD' in sales_df.columns:
                            sales_df = sales_df[sales_df['BRD_CD'] == 'X']
                    
                    category_sales = sales_df.groupby('ITEM_NM')['SALE_AMT_TY'].sum().sort_values(ascending=False)
                    if len(category_sales) > 0:
                        top_5 = category_sales.head(5)
                        result = "상위 카테고리 순위:\n"
                        for i, (category, sales) in enumerate(top_5.items(), 1):
                            result += f"{i}. {category}: {sales:,.0f}원\n"
                        return result.strip()
                    else:
                        return "카테고리 데이터가 없습니다."
                else:
                    return "카테고리별 매출 데이터를 찾을 수 없습니다."
    
    # 일반적인 데이터 분석 능력
    if not execution_df.empty:
        # 데이터 기본 정보
        if "데이터" in question_lower and ("개수" in question_lower or "건수" in question_lower or "몇개" in question_lower):
            return f"집행 데이터 총 {len(execution_df)}건"
        
        # 날짜 범위
        if "날짜" in question_lower and ("범위" in question_lower or "기간" in question_lower):
            if '날짜' in execution_df.columns:
                min_date = execution_df['날짜'].min()
                max_date = execution_df['날짜'].max()
                return f"데이터 기간: {min_date} ~ {max_date}"
        
        # 인플루언서 수
        if "인플루언서" in question_lower and ("몇명" in question_lower or "수" in question_lower):
            if '인플루언서' in execution_df.columns:
                unique_influencers = execution_df['인플루언서'].nunique()
                return f"총 {unique_influencers}명의 인플루언서"
    
    # 매출 데이터 기본 정보
    if not sales_df.empty:
        if "매출" in question_lower and ("데이터" in question_lower or "건수" in question_lower):
            return f"매출 데이터 총 {len(sales_df)}건"
        
        if "브랜드" in question_lower and ("몇개" in question_lower or "수" in question_lower):
            if 'BRD_CD' in sales_df.columns:
                unique_brands = sales_df['BRD_CD'].nunique()
                return f"총 {unique_brands}개 브랜드"
    
    # 기본 응답
    return "질문을 이해하지 못했습니다. 더 구체적으로 질문해주세요."

# =============================================================================
# 검색량분석 관련 함수들
# =============================================================================

def save_search_query(query):
    """검색량 쿼리를 파일로 저장"""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(SEARCH_QUERY_FILE, 'w', encoding='utf-8') as f:
            f.write(query)
        return True
    except Exception as e:
        st.error(f"검색량 쿼리 저장 실패: {str(e)}")
        return False

def load_search_query():
    """저장된 검색량 쿼리 파일에서 로드"""
    try:
        if os.path.exists(SEARCH_QUERY_FILE):
            with open(SEARCH_QUERY_FILE, 'r', encoding='utf-8') as f:
                return f.read()
    except Exception as e:
        st.warning(f"검색량 쿼리 로드 실패: {str(e)}")
    return None

def save_sales_query(query):
    """매출 쿼리를 파일로 저장"""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(SALES_QUERY_FILE, 'w', encoding='utf-8') as f:
            f.write(query)
        return True
    except Exception as e:
        st.error(f"매출 쿼리 저장 실패: {str(e)}")
        return False

def load_sales_query():
    """저장된 매출 쿼리 파일에서 로드"""
    try:
        if os.path.exists(SALES_QUERY_FILE):
            with open(SALES_QUERY_FILE, 'r', encoding='utf-8') as f:
                return f.read()
    except Exception as e:
        st.warning(f"매출 쿼리 로드 실패: {str(e)}")
    return None

def get_default_search_query():
    """기본 검색량 쿼리 반환"""
    # 1. 세션에 저장된 쿼리가 있으면 사용
    if hasattr(st.session_state, 'default_search_query') and st.session_state.default_search_query:
        return st.session_state.default_search_query
    
    # 2. 파일에서 저장된 쿼리 로드
    saved_query = load_search_query()
    if saved_query:
        st.session_state.default_search_query = saved_query
        return saved_query
    
    # 하드코딩된 기본 쿼리 (사용자 설정 쿼리로 업데이트)
    return """
    WITH current_day AS (
        SELECT
            w.SRCH_DT::DATE AS START_DT,                 -- 데일리 출력(START=END=해당일)
            w.SRCH_DT::DATE AS END_DT,
            w.KWD_NM        AS search_keyword,
            SUM(w.SRCH_CNT) AS SRCH_CNT_TY,
            m.BRD_CD        AS brand_code,
            m.ADULT_KIDS,
            m.CAT_NM        AS category,
            m.SUB_CAT_NM    AS sub_category,
            m.KWD_TYPE      AS keyword_type,
            m.COMP_TYPE,
            m.COMP_BRD_NM   AS comp_brand_name,
            YEAR(w.SRCH_DT)             AS ty_year,
            WEEKOFYEAR(w.SRCH_DT)       AS ty_week,
            DAYOFWEEKISO(w.SRCH_DT)     AS ty_dow
        FROM PRCS.DB_SRCH_KWD_NAVER_D w
        JOIN PRCS.DB_SRCH_KWD_NAVER_MST m
          ON w.KWD_NM = m.KWD_NM
        WHERE m.COMP_TYPE = '자사'
          AND m.BRD_CD IN ('M','X','V','ST')
          AND w.SRCH_DT >= '2024-09-01'
        GROUP BY
            w.SRCH_DT, w.KWD_NM,
            m.BRD_CD, m.ADULT_KIDS, m.CAT_NM, m.SUB_CAT_NM,
            m.KWD_TYPE, m.COMP_TYPE, m.COMP_BRD_NM
    ),
    previous_year_day AS (
        SELECT
            w2.SRCH_DT::DATE AS ly_dt,
            w2.KWD_NM        AS search_keyword,
            SUM(w2.SRCH_CNT) AS SRCH_CNT_LY_RAW,
            m2.BRD_CD        AS brand_code,
            m2.ADULT_KIDS,
            m2.CAT_NM        AS category,
            m2.SUB_CAT_NM    AS sub_category,
            m2.KWD_TYPE      AS keyword_type,
            m2.COMP_TYPE,
            m2.COMP_BRD_NM   AS comp_brand_name,
            YEAR(w2.SRCH_DT)         AS ly_year,
            WEEKOFYEAR(w2.SRCH_DT)   AS ly_week,
            DAYOFWEEKISO(w2.SRCH_DT) AS ly_dow
        FROM PRCS.DB_SRCH_KWD_NAVER_D w2
        JOIN PRCS.DB_SRCH_KWD_NAVER_MST m2
          ON w2.KWD_NM = m2.KWD_NM
        WHERE m2.COMP_TYPE = '자사'
          AND m2.BRD_CD IN ('M','X','V','ST')
          AND w2.SRCH_DT >= '2023-09-01'
        GROUP BY
            w2.SRCH_DT, w2.KWD_NM,
            m2.BRD_CD, m2.ADULT_KIDS, m2.CAT_NM, m2.SUB_CAT_NM,
            m2.KWD_TYPE, m2.COMP_TYPE, m2.COMP_BRD_NM
    )
    SELECT
        c.START_DT,
        c.END_DT,
        c.search_keyword,
        c.SRCH_CNT_TY,
        COALESCE(p.SRCH_CNT_LY_RAW, 0) AS SRCH_CNT_LY,
        c.brand_code,
        c.ADULT_KIDS,
        c.category,
        c.sub_category,
        c.keyword_type,
        c.COMP_TYPE,
        c.comp_brand_name
    FROM current_day c
    LEFT JOIN previous_year_day p
      ON  c.search_keyword = p.search_keyword
      AND c.brand_code     = p.brand_code
      /* 같은 주차 + 같은 요일 기준으로 전년 매칭 */
      AND p.ly_year = c.ty_year - 1
      AND p.ly_week = c.ty_week
      AND p.ly_dow  = c.ty_dow
    ORDER BY c.START_DT DESC, c.SRCH_CNT_TY DESC;
    """

def load_snowflake_search_data(start_date=None, end_date=None):
    """Snowflake에서 검색량 데이터 불러오기"""
    try:
        # 날짜 기본값 설정 (2024년 9월 2일부터 최신까지)
        if not start_date:
            start_date = '2024-09-02'  # 2024년 9월 2일부터
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')  # 최신 날짜까지
        
        # 날짜 형식 확인 및 변환
        if isinstance(start_date, datetime):
            start_date = start_date.strftime('%Y-%m-%d')
        if isinstance(end_date, datetime):
            end_date = end_date.strftime('%Y-%m-%d')
        
        # 매번 새로운 연결 생성 (캐시 제거됨)
        conn = get_snowflake_connection()
        if not conn:
            return pd.DataFrame()
        
        # 쿼리 관리에 저장된 쿼리 사용 (대시보드는 업로드 관리 쿼리만 사용)
        query = None
        if hasattr(st.session_state, 'custom_search_query') and st.session_state.custom_search_query:
            query = st.session_state.custom_search_query
        elif hasattr(st.session_state, 'default_search_query') and st.session_state.default_search_query:
            query = st.session_state.default_search_query

        # 관리 쿼리가 없는 경우: 실행하지 않고 안내
        if not query:
            st.info("검색량 데이터가 없습니다. '데이터 업로드 관리'에서 Snowflake에서 데이터를 불러와주세요.")
            return pd.DataFrame()
        
        try:
            # 매번 새로운 연결 생성 (캐시 제거로 인해 항상 새 연결)
            cursor = conn.cursor()
            cursor.execute(query)
            data = cursor.fetchall()
            
            # 컬럼명 가져오기
            columns = [desc[0] for desc in cursor.description]
            
            cursor.close()
            conn.close()
            
        except Exception as e:
            # 연결 정리
            try:
                if conn and not conn.is_closed():
                    conn.close()
            except:
                pass
            
            st.error(f"Snowflake 검색량 데이터 로딩 실패: {str(e)}")
            return pd.DataFrame()
        
        if data:
            df = pd.DataFrame(data, columns=columns)
            return df
        else:
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"Snowflake 검색량 데이터 로딩 실패: {str(e)}")
        return pd.DataFrame()

def save_search_data(df):
    """검색량 데이터를 CSV 파일로 저장"""
    try:
        df.to_csv(SEARCH_FILE, index=False, encoding='utf-8-sig')
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"검색량 데이터 저장 실패: {str(e)}")
        return False

def load_search_data():
    """로컬 검색량 데이터 불러오기"""
    try:
        if os.path.exists(SEARCH_FILE):
            df = pd.read_csv(SEARCH_FILE, encoding='utf-8-sig')
            return df
        else:
            return pd.DataFrame()
    except Exception as e:
        st.error(f"검색량 데이터 로딩 실패: {str(e)}")
        return pd.DataFrame()

def render_sales_dashboard_tab():
    """매출대시보드 탭 렌더링"""
    st.markdown("# 💰 매출대시보드")
    
    # 데이터 로드
    sales_df = load_sales_data()
    
    if sales_df.empty:
        st.warning("매출 데이터가 없습니다.")
        return
    
    st.markdown("## 📊 일별 매출 분석")
    
    # 브랜드, 아이템, 시즌, 기간 필터 추가 (대시보드와 동일한 구조)
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("#### 🏷️ 브랜드")
        # 브랜드 매핑 설정
        brand_mapping = {
            'MLB': 'M',
            'DX': 'X', 
            'DV': 'V',
            'ST': 'ST'
        }
        
        # 매출 데이터에서 브랜드 목록 가져오기 (MLB 우선순위)
        if 'BRD_CD' in sales_df.columns:
            available_brands = []
            for brand_code in sales_df['BRD_CD'].unique():
                for display_name, code in brand_mapping.items():
                    if code == brand_code:
                        available_brands.append(display_name)
                        break
            
            # MLB를 첫 번째로 정렬
            if 'MLB' in available_brands:
                available_brands.remove('MLB')
                available_brands = ['MLB'] + available_brands
        else:
            available_brands = ['전체']
            
        selected_brand = st.selectbox(
            "분석할 브랜드를 선택하세요", 
            options=["전체"] + available_brands,
            index=0,
            key="sales_brand_filter"
        )
    
    with col2:
        st.markdown("#### 📦 아이템")
        # 매출 데이터에서 아이템 목록 가져오기
        if 'ITEM' in sales_df.columns:
            available_items = sales_df['ITEM'].unique().tolist()
        else:
            available_items = []
            
        selected_item = st.selectbox(
            "분석할 아이템을 선택하세요", 
            options=["전체"] + available_items,
            index=0,
            key="sales_item_filter"
        )
    
    with col3:
        st.markdown("#### 🌟 시즌")
        # 시즌 필터 옵션
        season_options = ["전체", "24FW", "25SS", "25FW"]
        selected_season = st.selectbox(
            "분석할 시즌을 선택하세요",
            options=season_options,
            index=0,
            key="sales_season_filter"
        )
    
    with col4:
        st.markdown("#### 📅 기간 선택")
        # 날짜 범위 설정
        if not sales_df.empty and 'DT' in sales_df.columns:
            try:
                sales_df['DT'] = pd.to_datetime(sales_df['DT'], errors='coerce')
                valid_dates = sales_df['DT'].dropna()
                
                if len(valid_dates) > 0:
                    # 시즌 선택에 따른 날짜 범위 설정
                    if selected_season == "24FW":
                        season_min = pd.to_datetime('2024-09-01')
                        season_max = pd.to_datetime('2025-02-28')
                    elif selected_season == "25SS":
                        season_min = pd.to_datetime('2025-03-01')
                        season_max = pd.to_datetime('2025-08-31')
                    elif selected_season == "25FW":
                        season_min = pd.to_datetime('2025-09-01')
                        season_max = pd.to_datetime('2026-02-28')
                    else:
                        # 전체 선택 시 모든 데이터 범위 사용
                        season_min = valid_dates.min()
                        season_max = valid_dates.max()
                    
                    # 시즌 범위와 실제 데이터 범위의 교집합
                    min_date = max(season_min, valid_dates.min())
                    max_date = min(season_max, valid_dates.max())
                    
                    # 날짜 슬라이더
                    date_range = st.slider(
                        "분석할 기간을 선택하세요",
                        min_value=min_date.date(),
                        max_value=max_date.date(),
                        value=(min_date.date(), max_date.date()),
                        format="YYYY-MM-DD",
                        key="sales_date_range_slider"
                    )
                else:
                    st.warning("유효한 날짜 데이터를 찾을 수 없습니다.")
                    date_range = None
            except Exception as e:
                st.error(f"날짜 처리 중 오류가 발생했습니다: {e}")
                date_range = None
        else:
            st.warning("날짜 데이터가 없습니다.")
            date_range = None
    
    # 매출 데이터 필터링 및 그래프 생성
    if not sales_df.empty:
        # 필터링된 매출 데이터 준비
        filtered_sales_df = sales_df.copy()
        
        # 브랜드 필터 적용
        if selected_brand != "전체":
            brand_code = brand_mapping.get(selected_brand, selected_brand)
            if 'BRD_CD' in filtered_sales_df.columns:
                filtered_sales_df = filtered_sales_df[filtered_sales_df['BRD_CD'] == brand_code]
        
        # 아이템 필터 적용
        if selected_item != "전체":
            filtered_sales_df = filtered_sales_df[filtered_sales_df['ITEM'] == selected_item]
        
        # 시즌 필터 적용
        if selected_season != "전체":
            if '시즌' in filtered_sales_df.columns:
                filtered_sales_df = filtered_sales_df[filtered_sales_df['시즌'] == selected_season]
        
        # 날짜 필터 적용
        if date_range:
            filtered_sales_df['DT'] = pd.to_datetime(filtered_sales_df['DT'])
            filtered_sales_df = filtered_sales_df[
                (filtered_sales_df['DT'] >= pd.to_datetime(date_range[0])) &
                (filtered_sales_df['DT'] <= pd.to_datetime(date_range[1]))
            ]
        
        # 일별 매출 데이터 집계
        if not filtered_sales_df.empty:
            daily_sales = filtered_sales_df.groupby('DT').agg({
                'SALE_AMT_TY': 'sum',  # 당해 매출액
                'SALE_AMT_LY': 'sum'   # 전년 매출액
            }).reset_index()
            
            daily_sales['DT'] = pd.to_datetime(daily_sales['DT'])
            daily_sales = daily_sales.sort_values('DT')
            
            # 매출액 그래프 생성
            import plotly.graph_objects as go
            
            fig = go.Figure()
            
            # 당해 매출액 라인
            fig.add_trace(go.Scatter(
                x=daily_sales['DT'],
                y=daily_sales['SALE_AMT_TY'],
                mode='lines+markers',
                name='당해 매출액',
                line=dict(color='#1f77b4', width=3),
                marker=dict(size=6)
            ))
            
            # 전년 매출액 라인 (YoY 비교용)
            if 'SALE_AMT_LY' in daily_sales.columns and not daily_sales['SALE_AMT_LY'].isna().all():
                fig.add_trace(go.Scatter(
                    x=daily_sales['DT'],
                    y=daily_sales['SALE_AMT_LY'],
                    mode='lines+markers',
                    name='전년 매출액',
                    line=dict(color='#ff7f0e', width=2, dash='dash'),
                    marker=dict(size=5)
                ))
            
            # 레이아웃 설정
            fig.update_layout(
                title="일별 매출액 트렌드 (당해 vs 전년)",
                xaxis_title="날짜",
                yaxis_title="매출액 (원)",
                xaxis=dict(
                    type='date',
                    showgrid=True,
                    gridcolor='lightgray'
                ),
                yaxis=dict(
                    showgrid=True,
                    gridcolor='lightgray',
                    tickformat=',.0f'
                ),
                hovermode='x unified',
                                                    legend=dict(
                                                        orientation='h',
                                                        x=0,
                                                        y=1.1,
                                                        xanchor='left',
                                                        yanchor='bottom',
                                                        bgcolor='rgba(255,255,255,0.8)',
                                                        bordercolor='gray',
                                                        borderwidth=1
                                                    ),
                height=500,
                template='plotly_white'
            )
            
            # 호버 템플릿 설정
            fig.update_traces(
                hovertemplate='<b>%{fullData.name}</b><br>' +
                             '날짜: %{x}<br>' +
                             '매출액: %{y:,.0f}원<br>' +
                             '<extra></extra>'
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # 매출 통계 요약
            st.markdown("### 📈 매출 통계 요약")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_current = daily_sales['SALE_AMT_TY'].sum()
                st.metric(
                    label="**선택 기간 매출액**",
                    value=f"{total_current:,.0f}원"
                )
            
            with col2:
                if 'SALE_AMT_LY' in daily_sales.columns and not daily_sales['SALE_AMT_LY'].isna().all():
                    total_previous = daily_sales['SALE_AMT_LY'].sum()
                    st.metric(
                        label="**전년 동일기간 매출액**",
                        value=f"{total_previous:,.0f}원"
                    )
                else:
                    st.metric(
                        label="**전년 동일기간 매출액**",
                        value="데이터 없음"
                    )
            
            with col3:
                if 'SALE_AMT_LY' in daily_sales.columns and not daily_sales['SALE_AMT_LY'].isna().all():
                    total_previous = daily_sales['SALE_AMT_LY'].sum()
                    if total_previous > 0:
                        yoy_change = ((total_current - total_previous) / total_previous) * 100
                        st.metric(
                            label="**전년 동일기간 대비 증감률**",
                            value=f"{yoy_change:+.1f}%"
                        )
                    else:
                        st.metric(
                            label="**전년 동일기간 대비 증감률**",
                            value="계산 불가"
                        )
                else:
                    st.metric(
                        label="**전년 동일기간 대비 증감률**",
                        value="데이터 없음"
                    )
            
            with col4:
                avg_daily = daily_sales['SALE_AMT_TY'].mean()
                st.metric(
                    label="**일평균 매출액**",
                    value=f"{avg_daily:,.0f}원"
                )
        else:
            st.warning("선택한 필터 조건에 해당하는 매출 데이터가 없습니다.")
    else:
        st.warning("매출 데이터를 불러올 수 없습니다.")

def render_search_analysis_tab():
    """검색량분석 탭 렌더링"""
    st.markdown("# 🔍 검색량분석")
    st.markdown("""
    <style>
    div[data-testid="stMetricLabel"] {
        font-size: 0.7em; /* 라벨 글자 크기 줄임 */
    }
    div[data-testid="stMetricValue"] {
        font-size: 0.9em; /* 값 글자 크기 줄임 */
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 검색량 데이터 로드 (저장된 데이터 우선 사용)
    search_df = pd.DataFrame()
    
    # 1. 먼저 세션에 저장된 데이터 확인
    if hasattr(st.session_state, 'search_data') and not st.session_state.search_data.empty:
        search_df = st.session_state.search_data
    # 2. 세션에 없으면 로컬 CSV에서 불러오기 (새로고침 시 세션 초기화 대응)
    elif os.path.exists(SEARCH_FILE):
        file_df = load_search_data()
        if not file_df.empty:
            st.session_state.search_data = file_df
            search_df = file_df
    # 3. 그래도 없으면 데이터 없음 안내 (직접 Snowflake 로드는 사용하지 않음)
    else:
        search_df = pd.DataFrame()
    
    if not search_df.empty:
        # 데이터 필터링 (기본값 사용)
        filtered_df = search_df
        
        if not filtered_df.empty:
            
            # 브랜드 옵션 정의 (실제 데이터에서 동적으로 가져오기)
            brand_columns = ['BRD_CD', 'brand_code', 'BRAND_CODE', 'brand_cd']
            brand_col = None
            for col in brand_columns:
                if col in filtered_df.columns:
                    brand_col = col
                    break
            
            if brand_col:
                # 실제 데이터에서 고유한 브랜드 코드 추출
                unique_brands = sorted(filtered_df[brand_col].dropna().unique())
                
                # 지정된 브랜드만 필터링 (M, X, V, ST)
                allowed_brands = ['M', 'X', 'V', 'ST']
                filtered_brands = [brand for brand in unique_brands if brand in allowed_brands]
                brand_options = ['전체'] + filtered_brands
            else:
                # 브랜드 컬럼이 없으면 기본값 사용
                brand_options = ['전체', 'V']
            
            # 가장 최근 주차 날짜 계산
            if not filtered_df.empty and 'START_DT' in filtered_df.columns:
                # 가장 최근 날짜 찾기
                latest_date = pd.to_datetime(filtered_df['START_DT']).max()
                # 해당 날짜의 주차 시작일 계산 (월요일)
                latest_week_start = latest_date - pd.Timedelta(days=latest_date.weekday())
                default_date = latest_week_start.date()
            else:
                default_date = datetime(2024, 9, 2).date()
            
            dashboard_start_date = st.date_input(
                "대시보드 시작 날짜 (해당 주의 월요일 데이터 표시)",
                value=default_date,  # 가장 최근 주차 시작일
                key="dashboard_start_date",
                help="선택한 날짜가 포함된 주의 월요일 데이터를 표시합니다."
            )
            
            # 검색어 순위 대시보드 표시
            col_title, col_download = st.columns([3, 1])
            with col_title:
                st.markdown("#### 📋 검색어 순위 대시보드")
            with col_download:
                # 주차별 검색량 엑셀 다운로드
                if st.button("📥 주차별 검색량 다운로드", use_container_width=True, key="weekly_search_download"):
                    # 현재 주차가 속한 월의 주차별 검색량 데이터 생성
                    selected_date = pd.to_datetime(dashboard_start_date)
                    current_month = selected_date.month
                    current_year = selected_date.year
                    
                    # 해당 월의 첫날과 마지막날 계산
                    month_start = pd.Timestamp(current_year, current_month, 1)
                    if current_month == 12:
                        month_end = pd.Timestamp(current_year + 1, 1, 1) - pd.Timedelta(days=1)
                    else:
                        month_end = pd.Timestamp(current_year, current_month + 1, 1) - pd.Timedelta(days=1)
                    
                    # START_DT와 END_DT 기준으로 해당 월에 포함되는 데이터 필터링
                    filtered_df_copy = filtered_df.copy()
                    filtered_df_copy['START_DT_dt'] = pd.to_datetime(filtered_df_copy['START_DT'])
                    
                    # END_DT 컬럼이 있으면 사용, 없으면 START_DT를 END_DT로 사용
                    if 'END_DT' in filtered_df_copy.columns:
                        filtered_df_copy['END_DT_dt'] = pd.to_datetime(filtered_df_copy['END_DT'])
                    else:
                        filtered_df_copy['END_DT_dt'] = filtered_df_copy['START_DT_dt']
                    
                    # START_DT 또는 END_DT가 해당 월에 포함되면 포함
                    # 즉, 주차의 일부라도 해당 월에 포함되면 주차 전체를 포함
                    month_data = filtered_df_copy[
                        (
                            (filtered_df_copy['START_DT_dt'] >= month_start) & 
                            (filtered_df_copy['START_DT_dt'] <= month_end)
                        ) | (
                            (filtered_df_copy['END_DT_dt'] >= month_start) & 
                            (filtered_df_copy['END_DT_dt'] <= month_end)
                        )
                    ].copy()
                    
                    # 주차 단위로 포함하기 위해, 해당 월과 겹치는 주차의 모든 날짜를 포함해야 함
                    # 주차 시작일(월요일) 계산 - ISO 기준 (월요일=0, 일요일=6)
                    filtered_df_copy['week_start'] = filtered_df_copy['START_DT_dt'].apply(
                        lambda x: x - pd.Timedelta(days=x.weekday())
                    )
                    
                    # 해당 월에 포함되는 데이터들의 주차 찾기
                    month_overlap_data = filtered_df_copy[
                        (
                            (filtered_df_copy['START_DT_dt'] >= month_start) & 
                            (filtered_df_copy['START_DT_dt'] <= month_end)
                        ) | (
                            (filtered_df_copy['END_DT_dt'] >= month_start) & 
                            (filtered_df_copy['END_DT_dt'] <= month_end)
                        )
                    ]
                    
                    # 해당 월과 겹치는 모든 주차 찾기 (주차 시작일 기준)
                    valid_weeks = month_overlap_data['week_start'].unique()
                    
                    # 원본 데이터에서 해당 주차들의 모든 데이터 가져오기 (주차 전체 포함)
                    month_data = filtered_df_copy[
                        filtered_df_copy['week_start'].isin(valid_weeks)
                    ].copy()
                    
                    # 임시 컬럼 정리 (week_start는 나중에 사용하므로 유지)
                    month_data = month_data.drop(columns=['START_DT_dt', 'END_DT_dt'], errors='ignore')
                    
                    if not month_data.empty:
                        # 컬럼명 동적 확인
                        keyword_cols = ['SEARCH_KEYWORD', 'search_keyword', '검색어']
                        keyword_col = None
                        for col in keyword_cols:
                            if col in month_data.columns:
                                keyword_col = col
                                break
                        
                        if keyword_col is None:
                            st.error("검색어 컬럼을 찾을 수 없습니다.")
                        else:
                            # 브랜드 컬럼 확인
                            brand_cols = ['BRAND_CODE', 'brand_code', 'BRD_CD', 'brand_cd']
                            brand_col = None
                            for col in brand_cols:
                                if col in month_data.columns:
                                    brand_col = col
                                    break
                            
                            # 주차별로 그룹화 (월요일 기준)
                            month_data['week_start'] = pd.to_datetime(month_data['START_DT']).apply(
                                lambda x: x - pd.Timedelta(days=x.weekday())
                            )
                            # 주차 끝날짜 (일요일) 계산
                            month_data['week_end'] = month_data['week_start'] + pd.Timedelta(days=6)
                            
                            # 그룹화할 컬럼 목록
                            groupby_cols = ['week_start', 'week_end', keyword_col]
                            if brand_col:
                                groupby_cols.append(brand_col)
                            
                            # 주차별, 검색어별, 브랜드별 검색량 집계
                            # 먼저 주차별로 실제 포함된 날짜 확인 (디버깅용)
                            if len(month_data) > 0:
                                # 주차별로 포함된 START_DT 목록 확인
                                week_dates_check = month_data.groupby('week_start')['START_DT'].unique().reset_index()
                                week_dates_check['week_end'] = week_dates_check['week_start'] + pd.Timedelta(days=6)
                            
                            weekly_summary = month_data.groupby(groupby_cols).agg({
                                'SRCH_CNT_TY': 'sum',
                                'SRCH_CNT_LY': 'sum'
                            }).reset_index()
                            
                            # 주차별 실제 포함된 날짜 범위 확인 (디버깅 메시지)
                            if len(weekly_summary) > 0:
                                check_df = month_data.groupby('week_start').agg({
                                    'START_DT': ['min', 'max']
                                }).reset_index()
                                check_df.columns = ['week_start', 'min_date', 'max_date']
                                # 임시로 확인 메시지 표시 (디버깅용, 필요시 제거)
                                # st.info(f"주차별 데이터 범위: {len(check_df)}개 주차 포함")
                            
                            # 전년 비교 계산
                            weekly_summary['전년비교'] = weekly_summary.apply(
                                lambda row: f"{((row['SRCH_CNT_TY'] - row['SRCH_CNT_LY']) / row['SRCH_CNT_LY'] * 100):+.1f}%" 
                                if row['SRCH_CNT_LY'] > 0 else "신규", axis=1
                            )
                            
                            # 컬럼명 변경
                            rename_dict = {
                                'week_start': 'START_DT',
                                'week_end': 'END_DT',
                                keyword_col: '검색어',
                                'SRCH_CNT_TY': '기간검색량',
                                'SRCH_CNT_LY': '전년검색량'
                            }
                            
                            # 브랜드 컬럼명 매핑
                            brand_mapping = {'M': 'MLB', 'X': 'DX', 'V': 'DV', 'ST': 'ST'}
                            if brand_col:
                                rename_dict[brand_col] = '브랜드'
                                # 브랜드 코드를 브랜드명으로 변환
                                weekly_summary[brand_col] = weekly_summary[brand_col].map(brand_mapping).fillna(weekly_summary[brand_col])
                            
                            export_df = weekly_summary.rename(columns=rename_dict)
                            
                            # 전년검색량을 정수로 포맷팅 (없는 경우 0)
                            export_df['전년검색량'] = export_df['전년검색량'].fillna(0).astype(int)
                            
                            # 컬럼 순서 설정
                            if brand_col:
                                export_df = export_df[['START_DT', 'END_DT', '브랜드', '검색어', '기간검색량', '전년검색량', '전년비교']]
                            else:
                                export_df = export_df[['START_DT', 'END_DT', '검색어', '기간검색량', '전년검색량', '전년비교']]
                            
                            # 주차 날짜를 datetime 형식으로 유지 (시간은 00:00:00으로 설정)
                            export_df['START_DT'] = pd.to_datetime(export_df['START_DT'])
                            export_df['END_DT'] = pd.to_datetime(export_df['END_DT'])
                            
                            # 주차 순서대로 정렬
                            export_df = export_df.sort_values(['START_DT', '기간검색량'], ascending=[True, False])
                            
                            # 엑셀 다운로드
                            from io import BytesIO
                            from datetime import datetime
                            
                            output = BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl', date_format='yyyy-mm-dd') as writer:
                                export_df.to_excel(writer, sheet_name='주차별검색량', index=False)
                                
                                # 날짜 컬럼 포맷 설정 (시간 제거)
                                workbook = writer.book
                                worksheet = writer.sheets['주차별검색량']
                                
                                # 날짜 컬럼 찾아서 포맷 적용
                                for col_idx, col_name in enumerate(export_df.columns, start=1):
                                    if col_name in ['START_DT', 'END_DT']:
                                        # 헤더를 제외한 모든 행에 날짜 포맷 적용
                                        for row_idx in range(2, len(export_df) + 2):
                                            cell = worksheet.cell(row=row_idx, column=col_idx)
                                            if isinstance(cell.value, datetime):
                                                cell.number_format = 'YYYY-MM-DD'
                            
                            output.seek(0)
                            
                            st.download_button(
                                label="💾 엑셀 파일 다운로드",
                                data=output.getvalue(),
                                file_name=f"주차별검색량_{current_year}{current_month:02d}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="weekly_search_excel_download"
                            )
                            st.success(f"✅ {current_year}년 {current_month}월 주차별 검색량 데이터를 준비했습니다!")
                    else:
                        st.warning("해당 월의 데이터가 없습니다.")
            
            # 로데이터 기반으로 대시보드 데이터 생성
            display_df = filtered_df.copy()
            
            # 1. 검색량 합계 계산 (키워드별로 그룹화)
            if 'SRCH_CNT_TY' in display_df.columns and 'SEARCH_KEYWORD' in display_df.columns:
                # 키워드별 검색량 합계
                keyword_summary = display_df.groupby('SEARCH_KEYWORD').agg({
                    'SRCH_CNT_TY': 'sum',
                    'SRCH_CNT_LY': 'sum',
                    'BRAND_CODE': 'first',
                    'CATEGORY': 'first',
                    'SUB_CATEGORY': 'first',
                    'KEYWORD_TYPE': 'first'
                }).reset_index()
                
                # 2. YoY 증감률 계산
                keyword_summary['YOY_CHANGE_PCT'] = keyword_summary.apply(
                    lambda row: ((row['SRCH_CNT_TY'] - row['SRCH_CNT_LY']) / row['SRCH_CNT_LY'] * 100) 
                    if row['SRCH_CNT_LY'] > 0 else 0, axis=1
                )
                
                # 3. 순위 계산 (검색량 기준 내림차순)
                keyword_summary = keyword_summary.sort_values('SRCH_CNT_TY', ascending=False)
                keyword_summary['RANK'] = range(1, len(keyword_summary) + 1)
                
                # 4. 컬럼명 매핑
                display_df = keyword_summary.rename(columns={
                    'RANK': '순위',
                    'SEARCH_KEYWORD': '검색어',
                    'YOY_CHANGE_PCT': '전년대비 증감',
                    'SRCH_CNT_TY': '기간검색량',
                    'BRAND_CODE': '브랜드코드',
                    'CATEGORY': '카테고리',
                    'SUB_CATEGORY': '서브카테고리',
                    'KEYWORD_TYPE': '키워드타입'
                })
                
                # 5. 전년대비 증감 포맷팅
                display_df['전년대비 증감'] = display_df['전년대비 증감'].apply(
                    lambda x: f"{x:+.1f}%" if pd.notna(x) else "0.0%"
                )
            else:
                # 컬럼이 없는 경우 기본값으로 빈 데이터프레임 생성
                display_df = pd.DataFrame(columns=['순위', '검색어', '기간검색량'])
            
            # 주차별 데이터 처리
            if not display_df.empty:
                # 선택된 시작날짜의 주차 계산
                selected_date = pd.to_datetime(dashboard_start_date)
                selected_week_start = selected_date - pd.Timedelta(days=selected_date.weekday())
                selected_week_end = selected_week_start + pd.Timedelta(days=6)
                
                # 전주차 계산
                prev_week_start = selected_week_start - pd.Timedelta(days=7)
                prev_week_end = prev_week_start + pd.Timedelta(days=6)
                
                # 전년도 동일주차 계산 (정확한 주차 매칭)
                # 현재 주차의 연도와 주차 번호를 구해서 전년도 동일 주차 찾기
                current_year = selected_week_start.year
                current_week_num = selected_week_start.isocalendar()[1]  # ISO 주차 번호
                
                # 전년도 동일 주차의 월요일 계산
                prev_year = current_year - 1
                prev_year_week_start = pd.Timestamp.fromisocalendar(prev_year, current_week_num, 1)
                prev_year_week_end = prev_year_week_start + pd.Timedelta(days=6)
                
                # 데이터가 월요일 기준이므로, 정확한 매칭을 위해 월요일만 필터링
                current_week_monday = selected_week_start
                prev_week_monday = prev_week_start
                prev_year_week_monday = prev_year_week_start
                
                # 원본 데이터에서 주차별 필터링 (주차 전체 데이터 집계)
                current_week_end = current_week_monday + pd.Timedelta(days=6)
                prev_week_end = prev_week_monday + pd.Timedelta(days=6)
                prev_year_week_end = prev_year_week_monday + pd.Timedelta(days=6)
                
                current_week_data = filtered_df[
                    (pd.to_datetime(filtered_df['START_DT']) >= pd.Timestamp(current_week_monday)) &
                    (pd.to_datetime(filtered_df['START_DT']) <= pd.Timestamp(current_week_end))
                ].copy()
                
                prev_week_data = filtered_df[
                    (pd.to_datetime(filtered_df['START_DT']) >= pd.Timestamp(prev_week_monday)) &
                    (pd.to_datetime(filtered_df['START_DT']) <= pd.Timestamp(prev_week_end))
                ].copy()
                
                prev_year_week_data = filtered_df[
                    (pd.to_datetime(filtered_df['START_DT']) >= pd.Timestamp(prev_year_week_monday)) &
                    (pd.to_datetime(filtered_df['START_DT']) <= pd.Timestamp(prev_year_week_end))
                ].copy()
                
                
                # 브랜드별로 이미 나누어서 표시하므로 전체 데이터 사용
                
                # 주차별 검색어 순위 계산
                def calculate_weekly_ranking(week_data, brand_code=None, rising_keywords=None, falling_keywords=None):
                    if week_data.empty:
                        return pd.DataFrame(columns=['순위', '검색어', '기간검색량'])
                    
                    # 브랜드별 특정 인물명이 포함된 키워드 제외
                    if brand_code:
                        excluded_keywords = []
                        if brand_code == 'V':  # DV 브랜드
                            excluded_keywords = ['김지원']
                        elif brand_code == 'ST':  # ST 브랜드
                            excluded_keywords = ['박지현']
                        elif brand_code == 'X':  # DX 브랜드
                            excluded_keywords = ['고윤정', '변우석']
                        
                        # 제외할 키워드가 포함된 검색어 필터링
                        if excluded_keywords:
                            for excluded_name in excluded_keywords:
                                week_data = week_data[~week_data['SEARCH_KEYWORD'].str.contains(excluded_name, na=False)]
                    
                    # 키워드별 검색량 합계
                    keyword_summary = week_data.groupby('SEARCH_KEYWORD').agg({
                        'SRCH_CNT_TY': 'sum',
                        'SRCH_CNT_LY': 'sum'
                    }).reset_index()
                    
                    # YoY 증감률 계산
                    keyword_summary['YOY_CHANGE_PCT'] = keyword_summary.apply(
                        lambda row: ((row['SRCH_CNT_TY'] - row['SRCH_CNT_LY']) / row['SRCH_CNT_LY'] * 100) 
                        if row['SRCH_CNT_LY'] > 0 else 0, axis=1
                    )
                    
                    # 순위 계산
                    keyword_summary = keyword_summary.sort_values('SRCH_CNT_TY', ascending=False)
                    keyword_summary['RANK'] = range(1, len(keyword_summary) + 1)
                    
                    # 기간검색량만 표시 (전년대비 증감률 제거)
                    result_df = keyword_summary.copy()
                    result_df['기간검색량'] = result_df['SRCH_CNT_TY'].apply(lambda x: f"{x:,.0f}")
                    
                    # 상승/하락 키워드에 아이콘 추가 (키워드명 오른쪽에 표시)
                    if rising_keywords:
                        # 4단계 이상 상승 키워드 (초록 동그라미 두 개)
                        major_rising_keywords = [item['검색어'] for item in rising_keywords if item['순위변화'] >= 4]
                        # 2-3단계 상승 키워드 (초록 동그라미 한 개)
                        minor_rising_keywords = [item['검색어'] for item in rising_keywords if 2 <= item['순위변화'] < 4]
                        
                        def add_rising_icon(x):
                            if x in major_rising_keywords:
                                return f"{x} 🟢🟢"
                            elif x in minor_rising_keywords:
                                return f"{x} 🟢"
                            else:
                                return x
                        
                        result_df['검색어'] = result_df['SEARCH_KEYWORD'].apply(add_rising_icon)
                    else:
                        result_df['검색어'] = result_df['SEARCH_KEYWORD']
                    
                    # 하락 키워드에 회색 동그라미 추가 (4단계 이상 하락)
                    if falling_keywords:
                        falling_keyword_list = [item['검색어'] for item in falling_keywords]
                        result_df['검색어'] = result_df['검색어'].apply(
                            lambda x: f"{x} 🔴" if any(keyword in x for keyword in falling_keyword_list) else x
                        )
                    
                    # 컬럼명 매핑
                    result_df = result_df.rename(columns={
                        'RANK': '순위'
                    })
                    
                    # SRCH_CNT_LY 컬럼 유지
                    result_df = result_df[['순위', '검색어', '기간검색량', 'SRCH_CNT_TY', 'SRCH_CNT_LY']]
                    
                    return result_df.head(20)
                
                
                # 현재 주차, 전주차, 전년도 동일주차 순위 계산
                current_week_ranking = calculate_weekly_ranking(current_week_data)
                prev_week_ranking = calculate_weekly_ranking(prev_week_data)
                prev_year_week_ranking = calculate_weekly_ranking(prev_year_week_data)
                
                # 브랜드별 대시보드 표시
                brands = ['M', 'X', 'V', 'ST']
                brand_names = ['MLB', 'DX', 'DV', 'ST']
                
                for i, (brand_code, brand_name) in enumerate(zip(brands, brand_names)):
                    st.markdown(f"### 🏷️ {brand_name} 브랜드 검색어 순위")
                    
                    # 브랜드별 데이터 필터링
                    current_brand_data = current_week_data[current_week_data['BRAND_CODE'] == brand_code] if not current_week_data.empty else pd.DataFrame()
                    prev_brand_data = prev_week_data[prev_week_data['BRAND_CODE'] == brand_code] if not prev_week_data.empty else pd.DataFrame()
                    prev_year_brand_data = prev_year_week_data[prev_year_week_data['BRAND_CODE'] == brand_code] if not prev_year_week_data.empty else pd.DataFrame()
                    
                    # 전전주 데이터 계산
                    prev_prev_week_start = selected_week_start - pd.Timedelta(days=14)
                    prev_prev_week_end = selected_week_start - pd.Timedelta(days=8)
                    
                    # START_DT 컬럼을 datetime으로 변환
                    if not search_df.empty and 'START_DT' in search_df.columns:
                        search_df_copy = search_df.copy()
                        search_df_copy['START_DT'] = pd.to_datetime(search_df_copy['START_DT'])
                        prev_prev_week_data = search_df_copy[
                            (search_df_copy['START_DT'] >= prev_prev_week_start) & 
                            (search_df_copy['START_DT'] <= prev_prev_week_end)
                        ]
                    else:
                        prev_prev_week_data = pd.DataFrame()
                    prev_prev_brand_data = prev_prev_week_data[prev_prev_week_data['BRAND_CODE'] == brand_code] if not prev_prev_week_data.empty else pd.DataFrame()
                    
                    # 브랜드별 순위 계산 (브랜드 코드 전달)
                    current_brand_ranking = calculate_weekly_ranking(current_brand_data, brand_code)
                    prev_brand_ranking = calculate_weekly_ranking(prev_brand_data, brand_code)
                    prev_prev_brand_ranking = calculate_weekly_ranking(prev_prev_brand_data, brand_code)
                    prev_year_brand_ranking = calculate_weekly_ranking(prev_year_brand_data, brand_code)

                    # 전체(상위 20 제한 없이) 랭크 사전 생성 함수
                    def build_full_rank_dicts(week_df):
                        if week_df.empty:
                            return {}, {}
                        base = week_df.copy()
                        # 키워드별 집계
                        aggregated = base.groupby('SEARCH_KEYWORD', as_index=False)['SRCH_CNT_TY'].sum()
                        aggregated = aggregated.sort_values('SRCH_CNT_TY', ascending=False).reset_index(drop=True)
                        aggregated['순위'] = aggregated.index + 1
                        keyword_to_rank = dict(zip(aggregated['SEARCH_KEYWORD'], aggregated['순위']))
                        rank_to_keyword = dict(zip(aggregated['순위'], aggregated['SEARCH_KEYWORD']))
                        return keyword_to_rank, rank_to_keyword

                    # 전전주/전주 사전: 키워드→순위, 순위→키워드 모두 준비 (전체 랭크)
                    prev_prev_kw_to_rank, prev_prev_rank_to_kw = build_full_rank_dicts(prev_prev_brand_data)
                    prev_kw_to_rank, prev_rank_to_kw = build_full_rank_dicts(prev_brand_data)
                    
                    # 상승 키워드 계산 (현재 주차 기준) - 전체 순위 사전 사용
                    ranking_changes = []
                    if not current_brand_ranking.empty:
                        current_dict = dict(zip(current_brand_ranking['검색어'], current_brand_ranking['순위']))
                        
                        for keyword in current_dict:
                            current_rank = current_dict[keyword]
                            # 전주 전체 순위에서 해당 키워드의 순위 찾기
                            prev_rank = prev_kw_to_rank.get(keyword)
                            
                            if prev_rank is not None:
                                rank_change = prev_rank - current_rank  # 양수면 상승, 음수면 하락
                                ranking_changes.append({
                                    '검색어': keyword,
                                    '현재순위': current_rank,
                                    '전주차순위': prev_rank,
                                    '순위변화': rank_change
                                })
                    
                    # 2단계 이상 상승 키워드만 필터링
                    rising_keywords = [item for item in ranking_changes if item['순위변화'] >= 2]
                    rising_keywords.sort(key=lambda x: x['순위변화'], reverse=True)
                    
                    # 4단계 이상 하락 키워드 필터링
                    falling_keywords = [item for item in ranking_changes if item['순위변화'] <= -4]
                    falling_keywords.sort(key=lambda x: x['순위변화'])
                    
                    # 상승/하락 키워드가 있는 경우 현재 주차 테이블에 아이콘 표시
                    current_brand_ranking_with_icons = calculate_weekly_ranking(current_brand_data, brand_code, rising_keywords, falling_keywords)
                    
                    # 현재 주차 테이블에 전주대비 증감 컬럼 추가
                    current_brand_total = current_brand_data['SRCH_CNT_TY'].sum() if not current_brand_data.empty else 0
                    st.markdown(f"#### 📊 {selected_week_start.strftime('%Y/%m/%d')}~{selected_week_end.strftime('%m/%d')} 주차 (이번주차)")
                    st.markdown(f"<small>전체 검색량: {current_brand_total:,} (🟢🟢 4단계 이상 상승 | 🟢 2-3단계 상승 | 🔴 4단계 이상 하락)</small>", unsafe_allow_html=True)
                    
                    if not current_brand_ranking_with_icons.empty:
                        # 3개 주차 데이터 매칭 및 검색어 컬럼 확장
                        current_ranking_with_three_weeks = current_brand_ranking_with_icons.copy()
                        
                        # 전전주, 전주 데이터에서 검색어 정보 가져오기
                        # build_full_rank_dicts 결과 사용
                        prev_prev_dict = prev_prev_rank_to_kw  # 순위 → 검색어
                        prev_dict = prev_rank_to_kw            # 순위 → 검색어
                        
                        def get_keyword_by_rank(rank_dict, target_rank):
                            """순위에 해당하는 키워드 찾기 (양방향 dict 지원)"""
                            if not rank_dict:
                                return "-"
                            # case 1: {keyword -> rank}
                            sample_key = next(iter(rank_dict.keys()))
                            if isinstance(rank_dict[sample_key], (int, float)):
                                for keyword, rank in rank_dict.items():
                                    if rank == target_rank:
                                        return keyword
                                return "-"
                            # case 2: {rank -> keyword}
                            if target_rank in rank_dict:
                                return rank_dict.get(target_rank, "-")
                            return "-"
                        
                        def get_rank_by_keyword(rank_dict, keyword):
                            """키워드에 해당하는 순위 찾기 (양방향 dict 지원)"""
                            if not rank_dict:
                                return None
                            sample_key = next(iter(rank_dict.keys()))
                            # case 1: {keyword -> rank}
                            if isinstance(rank_dict[sample_key], (int, float)):
                                return rank_dict.get(keyword)
                            # case 2: {rank -> keyword}
                            for r, k in rank_dict.items():
                                if k == keyword:
                                    return r
                            return None
                        
                        def get_rank_display(rank_dict, keyword):
                            """키워드에 해당하는 순위 표시. 없으면 X 반환 (양방향 dict 지원)"""
                            actual_rank = get_rank_by_keyword(rank_dict, keyword)
                            if actual_rank is None:
                                return "X"
                            return f"{actual_rank}위"
                        
                        def get_clean_keyword(keyword_with_icons):
                            """아이콘이 포함된 키워드에서 원본 키워드 추출"""
                            if '🟢🟢' in keyword_with_icons:
                                return keyword_with_icons.replace('🟢🟢', '').strip()
                            elif '🟢' in keyword_with_icons:
                                return keyword_with_icons.replace('🟢', '').strip()
                            elif '🔴' in keyword_with_icons:
                                return keyword_with_icons.replace('🔴', '').strip()
                            else:
                                return keyword_with_icons
                        
                        # 3개 주차 검색어를 각각 별도 컬럼으로 생성
                        def create_prev_prev_keyword(row):
                            current_rank = row['순위']
                            return get_keyword_by_rank(prev_prev_dict, current_rank)
                        
                        def create_prev_keyword(row):
                            current_rank = row['순위']
                            return get_keyword_by_rank(prev_dict, current_rank)
                        
                        def create_current_keyword(row):
                            # 원본 검색어에 아이콘이 이미 포함되어 있으므로 그대로 반환
                            return row['검색어']
                        
                        # 3개 주차 검색어 컬럼 생성
                        current_ranking_with_three_weeks['검색어(전전주)'] = current_ranking_with_three_weeks.apply(create_prev_prev_keyword, axis=1)
                        current_ranking_with_three_weeks['검색어(전주)'] = current_ranking_with_three_weeks.apply(create_prev_keyword, axis=1)
                        current_ranking_with_three_weeks['⭐ 검색어(이번주)'] = current_ranking_with_three_weeks.apply(create_current_keyword, axis=1)
                        
                        # 3주간 순위 변동 계산
                        def calculate_three_week_rank_change(row):
                            current_rank = row['순위']
                            current_keyword = get_clean_keyword(row['검색어'])
                            
                            # 전전주 순위 표시 (키워드→순위 사전으로 정확한 실제 순위 표시)
                            prev_prev_rank_display = get_rank_display(prev_prev_kw_to_rank, current_keyword)
                            
                            # 전주 순위 표시
                            prev_rank_display = get_rank_display(prev_kw_to_rank, current_keyword)
                            
                            # 현재 주차 순위
                            current_rank_str = f"{current_rank}위"
                            
                            # 순위 변동 문자열 생성
                            return f"{prev_prev_rank_display}->{prev_rank_display}->{current_rank_str}"
                        
                        current_ranking_with_three_weeks['순위 변동'] = current_ranking_with_three_weeks.apply(calculate_three_week_rank_change, axis=1)
                        
                        # 전년동일주차 대비 증감률 계산
                        def calculate_yoy_change(row):
                            current_volume = row['SRCH_CNT_TY']
                            current_keyword = get_clean_keyword(row['검색어'])
                            
                            # 현재 데이터에서 SRCH_CNT_LY 컬럼 사용
                            if 'SRCH_CNT_LY' in row and pd.notna(row['SRCH_CNT_LY']) and row['SRCH_CNT_LY'] > 0:
                                prev_year_volume = row['SRCH_CNT_LY']
                                yoy_change_pct = ((current_volume - prev_year_volume) / prev_year_volume) * 100
                                
                                # 화살표 아이콘 추가
                                if yoy_change_pct > 0:
                                    arrow = "▲"
                                elif yoy_change_pct < 0:
                                    arrow = "▼"
                                else:
                                    arrow = "→"
                                
                                return f"{current_volume:,.0f} ({arrow}{yoy_change_pct:+.1f}%)"
                            else:
                                return f"{current_volume:,.0f} (신규)"
                        
                        current_ranking_with_three_weeks['기간검색량'] = current_ranking_with_three_weeks.apply(calculate_yoy_change, axis=1)
                        
                        # 테이블과 그래프를 나란히 표시
                        col_table, col_chart = st.columns([1.3, 0.7])
                        
                        with col_table:
                            # 테이블 표시 (3개 주차 검색어 컬럼 포함)
                            # 20개 행 모두 표시 (확실히 20개만 표시)
                            display_df = current_ranking_with_three_weeks[['순위', '검색어(전전주)', '검색어(전주)', '⭐ 검색어(이번주)', '기간검색량', '순위 변동']].head(20)
                            
                            # Streamlit의 dataframe은 height 파라미터로 스크롤 가능한 영역 생성
                            # 20개 행을 모두 보기 위해 충분한 높이 설정 (각 행 약 50px, 20행 = 1000px + 헤더)
                            st.dataframe(
                                display_df,
                                width='stretch',
                                hide_index=True,
                                height=1100  # 20개 행 모두 표시 가능한 높이
                            )
                        
                        with col_chart:
                            # 3주치 일별 매출 그래프 생성
                            
                            # 매출 데이터 로드 (Snowflake 데이터 사용)
                            sales_df = load_snowflake_sales_data()
                            
                            if not sales_df.empty and 'BRD_CD' in sales_df.columns:
                                # 브랜드별 매출 데이터 필터링
                                brand_sales = sales_df[sales_df['BRD_CD'] == brand_code].copy()
                                
                                # 필터 영역 (아이템과 검색카테고리 나란히 배치)
                                if not brand_sales.empty:
                                    col_item, col_category = st.columns(2)
                                    
                                    with col_item:
                                        # 사용 가능한 아이템 목록 가져오기 (ITEM 컬럼 우선 사용)
                                        available_items = ['전체']
                                        if 'ITEM' in brand_sales.columns:
                                            items = brand_sales['ITEM'].dropna().unique().tolist()
                                            available_items.extend(sorted(items))
                                        elif 'item' in brand_sales.columns:
                                            items = brand_sales['item'].dropna().unique().tolist()
                                            available_items.extend(sorted(items))
                                        elif 'ITEM_CD' in brand_sales.columns:
                                            items = brand_sales['ITEM_CD'].dropna().unique().tolist()
                                            available_items.extend(sorted(items))
                                        elif 'ITEM_NM' in brand_sales.columns:
                                            items = brand_sales['ITEM_NM'].dropna().unique().tolist()
                                            available_items.extend(sorted(items))
                                        
                                        # 아이템 다중선택 필터
                                        selected_items = st.multiselect(
                                            "📦 아이템",
                                            options=available_items,
                                            default=['전체'] if '전체' in available_items else available_items[:1],
                                            help="분석할 아이템을 선택하세요"
                                        )
                                    
                                    with col_category:
                                        # 검색어 데이터에서 SUB_CATEGORY 목록 가져오기
                                        if hasattr(st.session_state, 'search_data') and not st.session_state.search_data.empty:
                                            search_df = st.session_state.search_data
                                        else:
                                            search_df = load_snowflake_search_data()
                                        available_categories = ['전체']
                                        
                                        if not search_df.empty and 'SUB_CATEGORY' in search_df.columns:
                                            # 현재 브랜드의 서브카테고리만 필터링
                                            brand_search_data = search_df[search_df['BRAND_CODE'] == brand_code]
                                            if not brand_search_data.empty:
                                                categories = brand_search_data['SUB_CATEGORY'].dropna().unique().tolist()
                                                available_categories.extend(sorted(categories))
                                        
                                        # 검색카테고리 다중선택 필터
                                        selected_categories = st.multiselect(
                                            "🔍 검색카테고리",
                                            options=available_categories,
                                            default=['전체'] if '전체' in available_categories else available_categories[:1],
                                            help="분석할 검색카테고리를 선택하세요"
                                        )
                                    
                                    # 아이템 필터링 적용 (ITEM 컬럼 우선 사용)
                                    if selected_items and '전체' not in selected_items:
                                        if 'ITEM' in brand_sales.columns:
                                            brand_sales = brand_sales[brand_sales['ITEM'].isin(selected_items)]
                                        elif 'item' in brand_sales.columns:
                                            brand_sales = brand_sales[brand_sales['item'].isin(selected_items)]
                                        elif 'ITEM_CD' in brand_sales.columns:
                                            brand_sales = brand_sales[brand_sales['ITEM_CD'].isin(selected_items)]
                                        elif 'ITEM_NM' in brand_sales.columns:
                                            brand_sales = brand_sales[brand_sales['ITEM_NM'].isin(selected_items)]
                                    
                                    # 검색카테고리 필터링 적용 및 검색량 표시
                                    if selected_categories and '전체' not in selected_categories:
                                        # 검색어 데이터에서 선택된 카테고리의 키워드들 가져오기
                                        if not search_df.empty and 'SUB_CATEGORY' in search_df.columns:
                                            brand_search_data = search_df[search_df['BRAND_CODE'] == brand_code]
                                            if not brand_search_data.empty:
                                                # 선택된 카테고리에 해당하는 키워드들 필터링
                                                filtered_keywords = brand_search_data[
                                                    brand_search_data['SUB_CATEGORY'].isin(selected_categories)
                                                ]['SEARCH_KEYWORD'].unique().tolist()
                                                
                                                # 선택된 카테고리의 검색량 계산
                                                category_search_data = brand_search_data[
                                                    brand_search_data['SUB_CATEGORY'].isin(selected_categories)
                                                ]
                                                
                                                # 이번주 검색량 계산 (날짜 타입 변환)
                                                category_search_data_copy = category_search_data.copy()
                                                category_search_data_copy['START_DT'] = pd.to_datetime(category_search_data_copy['START_DT'])
                                                
                                                current_week_search = category_search_data_copy[
                                                    (category_search_data_copy['START_DT'] >= pd.Timestamp(selected_week_start)) & 
                                                    (category_search_data_copy['START_DT'] <= pd.Timestamp(selected_week_end))
                                                ]
                                                
                                                if not current_week_search.empty:
                                                    current_week_search_volume = current_week_search['SRCH_CNT_TY'].sum()
                                                    
                                                    # 전년 동일주 검색량 계산 (SRCH_CNT_LY 사용)
                                                    prev_year_week_search_volume = current_week_search['SRCH_CNT_LY'].sum()
                                                    
                                                    # 전년 대비 증감률 계산
                                                    if prev_year_week_search_volume > 0:
                                                        search_yoy_change = ((current_week_search_volume - prev_year_week_search_volume) / prev_year_week_search_volume) * 100
                                                        search_yoy_display = f" (전년 대비 {search_yoy_change:+.1f}%)"
                                                    else:
                                                        search_yoy_display = " (전년 데이터 없음)"
                                                    
                                                    # 검색량은 매출액과 함께 표시하므로 여기서는 제거
                                                else:
                                                    st.info(f"선택된 검색카테고리 ({', '.join(selected_categories)})의 이번주 검색량 데이터가 없습니다.")
                                
                                if not brand_sales.empty and 'DT' in brand_sales.columns:
                                    # 날짜 변환
                                    brand_sales['DT'] = pd.to_datetime(brand_sales['DT'])
                                    
                                    # 3주치 데이터 필터링 (전전주, 전주, 이번주)
                                    three_weeks_start = prev_prev_week_start
                                    three_weeks_end = selected_week_end
                                    
                                    three_weeks_sales = brand_sales[
                                        (brand_sales['DT'] >= three_weeks_start) & 
                                        (brand_sales['DT'] <= three_weeks_end)
                                    ]
                                    
                                    if not three_weeks_sales.empty:
                                        import plotly.graph_objects as go
                                        
                                        # 여러 아이템 선택 시 각각 개별 그래프 표시
                                        if selected_items and '전체' not in selected_items and len(selected_items) > 1:
                                            # 각 아이템별로 개별 그래프 생성
                                            for i, item in enumerate(selected_items):
                                                # 해당 아이템의 데이터만 필터링
                                                item_data = three_weeks_sales.copy()
                                                if 'ITEM' in item_data.columns:
                                                    item_data = item_data[item_data['ITEM'] == item]
                                                elif 'item' in item_data.columns:
                                                    item_data = item_data[item_data['item'] == item]
                                                elif 'ITEM_CD' in item_data.columns:
                                                    item_data = item_data[item_data['ITEM_CD'] == item]
                                                elif 'ITEM_NM' in item_data.columns:
                                                    item_data = item_data[item_data['ITEM_NM'] == item]
                                                
                                                if not item_data.empty:
                                                    # 일별 매출 집계
                                                    daily_sales = item_data.groupby('DT').agg({
                                                        'SALE_AMT_TY': 'sum',
                                                        'SALE_AMT_LY': 'sum'
                                                    }).reset_index()
                                                    
                                                    daily_sales = daily_sales.sort_values('DT')
                                                    
                                                    # 그래프 생성
                                                    fig = go.Figure()
                                                    
                                                    # 당해 매출액 라인
                                                    fig.add_trace(go.Scatter(
                                                        x=daily_sales['DT'],
                                                        y=daily_sales['SALE_AMT_TY'],
                                                        mode='lines+markers',
                                                        name='당해 매출액',
                                                        line=dict(color='#1f77b4', width=2),
                                                        marker=dict(size=4)
                                                    ))
                                                    
                                                    # 전년 매출액 라인 (YoY 비교용)
                                                    if 'SALE_AMT_LY' in daily_sales.columns and not daily_sales['SALE_AMT_LY'].isna().all():
                                                        fig.add_trace(go.Scatter(
                                                            x=daily_sales['DT'],
                                                            y=daily_sales['SALE_AMT_LY'],
                                                            mode='lines+markers',
                                                            name='전년 매출액',
                                                            line=dict(color='#ff7f0e', width=2, dash='dash'),
                                                            marker=dict(size=4)
                                                        ))
                                                    
                                                    # 일별 검색량 막대그래프 추가
                                                    # 해당 아이템의 검색량 데이터 필터링
                                                    item_search_data = brand_search_data.copy()
                                                    if 'ITEM' in item_search_data.columns:
                                                        item_search_data = item_search_data[item_search_data['ITEM'] == item]
                                                    elif 'item' in item_search_data.columns:
                                                        item_search_data = item_search_data[item_search_data['item'] == item]
                                                    elif 'ITEM_CD' in item_search_data.columns:
                                                        item_search_data = item_search_data[item_search_data['ITEM_CD'] == item]
                                                    elif 'ITEM_NM' in item_search_data.columns:
                                                        item_search_data = item_search_data[item_search_data['ITEM_NM'] == item]
                                                    
                                                    # 카테고리 필터링
                                                    if selected_categories and '전체' not in selected_categories:
                                                        # 선택된 카테고리만 필터링
                                                        category_search_data = item_search_data[
                                                            item_search_data['SUB_CATEGORY'].isin(selected_categories)
                                                        ]
                                                        search_label = f'검색량 ({",".join(selected_categories)})'
                                                    else:
                                                        # 전체 검색량
                                                        category_search_data = item_search_data
                                                        search_label = '전체 검색량'
                                                    
                                                    if not category_search_data.empty:
                                                        # 일별 검색량 집계
                                                        daily_search = category_search_data.groupby('START_DT').agg({
                                                            'SRCH_CNT_TY': 'sum'
                                                        }).reset_index()
                                                        
                                                        # 매출 그래프와 같은 날짜 범위로 필터링
                                                        daily_search = daily_search[
                                                            (pd.to_datetime(daily_search['START_DT']) >= pd.Timestamp(three_weeks_start)) &
                                                            (pd.to_datetime(daily_search['START_DT']) <= pd.Timestamp(three_weeks_end))
                                                        ]
                                                        
                                                        if not daily_search.empty:
                                                            # 검색량 막대그래프 추가
                                                            fig.add_trace(go.Bar(
                                                                x=pd.to_datetime(daily_search['START_DT']),
                                                                y=daily_search['SRCH_CNT_TY'],
                                                                name=search_label,
                                                                marker_color='rgba(128, 128, 128, 0.3)',
                                                                opacity=0.6,
                                                                yaxis='y2'
                                                            ))
                                                    
                                                    # 레이아웃 설정
                                                    fig.update_layout(
                                                        title=f"{brand_name} 일별 매출/검색량",
                                                        xaxis_title="날짜",
                                                        yaxis_title="매출액 (원)",
                                                        xaxis=dict(
                                                            type='date',
                                                            showgrid=True,
                                                            gridcolor='lightgray'
                                                        ),
                                                        yaxis=dict(
                                                            showgrid=True,
                                                            gridcolor='lightgray',
                                                            tickformat=',.0f'
                                                        ),
                                                        yaxis2=dict(
                                                            title="검색량 (회)",
                                                            overlaying="y",
                                                            side="right",
                                                            showgrid=False,
                                                            tickformat=',.0f'
                                                        ),
                                                        hovermode='x unified',
                                                        legend=dict(
                                                            orientation='h',
                                                            x=1,
                                                            y=1.1,
                                                            xanchor='right',
                                                            yanchor='bottom',
                                                            bgcolor='rgba(255,255,255,0.8)',
                                                            bordercolor='rgba(255,255,255,0)',
                                                            borderwidth=0
                                                        ),
                                                        height=300,
                                                        margin=dict(t=100),
                                                        template='plotly_white'
                                                    )
                                                    
                                                    # 호버 템플릿 설정
                                                    fig.update_traces(
                                                        hovertemplate='<b>%{fullData.name}</b><br>' +
                                                                     '날짜: %{x}<br>' +
                                                                     '매출액: %{y:,.0f}원<br>' +
                                                                     '<extra></extra>'
                                                    )
                                                    
                                                    # 검색량 막대그래프 호버 템플릿 별도 설정
                                                    fig.update_traces(
                                                        selector=dict(type='bar'),
                                                        hovertemplate='<b>%{fullData.name}</b><br>' +
                                                                     '날짜: %{x}<br>' +
                                                                     '검색량: %{y:,.0f}회<br>' +
                                                                     '<extra></extra>'
                                                    )
                                                    
                                                    st.plotly_chart(fig, use_container_width=True)
                                                    
                                                    # 개별 아이템 매출 통계
                                                    total_current = daily_sales['SALE_AMT_TY'].sum()
                                                    
                                                    # 이번주(10/20~10/26) 매출액 계산
                                                    current_week_start = selected_week_start
                                                    current_week_end = selected_week_end
                                                    current_week_sales = item_data[
                                                        (item_data['DT'] >= current_week_start) & 
                                                        (item_data['DT'] <= current_week_end)
                                                    ]
                                                    
                                                    if not current_week_sales.empty:
                                                        current_week_total = current_week_sales['SALE_AMT_TY'].sum()
                                                        current_week_prev_year = current_week_sales['SALE_AMT_LY'].sum() if 'SALE_AMT_LY' in current_week_sales.columns else 0
                                                        
                                                        # 전년 대비 증감률 계산
                                                        if current_week_prev_year > 0:
                                                            yoy_change = ((current_week_total - current_week_prev_year) / current_week_prev_year) * 100
                                                            yoy_display = f" (전년 대비 {yoy_change:+.1f}%)"
                                                        else:
                                                            yoy_display = " (전년 데이터 없음)"
                                                        
                                                    # 이번주 매출액과 검색량을 나란히 표시
                                                    col_sales, col_search = st.columns(2)
                                                    
                                                    with col_sales:
                                                        # 날짜 범위 포맷팅
                                                        date_range = f"{selected_week_start.strftime('%m/%d')}~{selected_week_end.strftime('%m/%d')}"
                                                        st.metric(
                                                            label=f"**{date_range} 매출액 ({item})**",
                                                            value=f"{current_week_total:,.0f}원{yoy_display}"
                                                        )
                                                    
                                                    with col_search:
                                                        if selected_categories and '전체' not in selected_categories:
                                                            # 선택된 카테고리 검색량 표시
                                                            st.metric(
                                                                label=f"**{date_range} 검색량 ({','.join(selected_categories)})**",
                                                                value=f"{current_week_search_volume:,.0f}회{search_yoy_display}"
                                                            )
                                                        else:
                                                            # 전체 검색량 계산
                                                            total_current_week_search = brand_search_data[
                                                                (pd.to_datetime(brand_search_data['START_DT']) >= pd.Timestamp(current_week_start)) &
                                                                (pd.to_datetime(brand_search_data['START_DT']) <= pd.Timestamp(current_week_end))
                                                            ]
                                                            total_current_week_search_volume = total_current_week_search['SRCH_CNT_TY'].sum() if not total_current_week_search.empty else 0
                                                            
                                                            # 전년 동일주 전체 검색량 계산 (SRCH_CNT_LY 사용)
                                                            total_prev_year_week_search_volume = brand_search_data['SRCH_CNT_LY'].sum() if not brand_search_data.empty else 0
                                                            
                                                            # 전년 대비 증감률 계산
                                                            total_search_yoy_display = ""
                                                            if total_prev_year_week_search_volume > 0:
                                                                total_search_yoy_change = ((total_current_week_search_volume - total_prev_year_week_search_volume) / total_prev_year_week_search_volume) * 100
                                                                total_search_yoy_display = f" (전년 대비 {total_search_yoy_change:+.1f}%)"
                                                            elif total_current_week_search_volume > 0:
                                                                total_search_yoy_display = " (신규)"
                                                            else:
                                                                total_search_yoy_display = " (-)"
                                                            
                                                            st.metric(
                                                                label=f"**{date_range} 전체 검색량**",
                                                                value=f"{total_current_week_search_volume:,.0f}회{total_search_yoy_display}"
                                                            )
                                                    
                                                    # 아이템 간 구분선
                                                    if i < len(selected_items) - 1:
                                                        st.markdown("---")
                                        else:
                                            # 단일 아이템 또는 전체 선택 시 기존 로직
                                            # 일별 매출 집계
                                            daily_sales = three_weeks_sales.groupby('DT').agg({
                                                'SALE_AMT_TY': 'sum',
                                                'SALE_AMT_LY': 'sum'
                                            }).reset_index()
                                            
                                            daily_sales = daily_sales.sort_values('DT')
                                            
                                            # 그래프 생성
                                            fig = go.Figure()
                                            
                                            # 당해 매출액 라인
                                            fig.add_trace(go.Scatter(
                                                x=daily_sales['DT'],
                                                y=daily_sales['SALE_AMT_TY'],
                                                mode='lines+markers',
                                                name='당해 매출액',
                                                line=dict(color='#1f77b4', width=2),
                                                marker=dict(size=4)
                                            ))
                                            
                                            # 전년 매출액 라인 (YoY 비교용)
                                            if 'SALE_AMT_LY' in daily_sales.columns and not daily_sales['SALE_AMT_LY'].isna().all():
                                                fig.add_trace(go.Scatter(
                                                    x=daily_sales['DT'],
                                                    y=daily_sales['SALE_AMT_LY'],
                                                    mode='lines+markers',
                                                    name='전년 매출액',
                                                    line=dict(color='#ff7f0e', width=2, dash='dash'),
                                                    marker=dict(size=4)
                                                ))
                                            
                                            # 일별 검색량 막대그래프 추가
                                            # 카테고리 필터링
                                            if selected_categories and '전체' not in selected_categories:
                                                # 선택된 카테고리만 필터링
                                                category_search_data = brand_search_data[
                                                    brand_search_data['SUB_CATEGORY'].isin(selected_categories)
                                                ]
                                                search_label = f'검색량 ({",".join(selected_categories)})'
                                            else:
                                                # 전체 검색량
                                                category_search_data = brand_search_data
                                                search_label = '전체 검색량'
                                            
                                            if not category_search_data.empty:
                                                # 일별 검색량 집계
                                                daily_search = category_search_data.groupby('START_DT').agg({
                                                    'SRCH_CNT_TY': 'sum'
                                                }).reset_index()
                                                
                                                # 매출 그래프와 같은 날짜 범위로 필터링
                                                daily_search = daily_search[
                                                    (pd.to_datetime(daily_search['START_DT']) >= pd.Timestamp(three_weeks_start)) &
                                                    (pd.to_datetime(daily_search['START_DT']) <= pd.Timestamp(three_weeks_end))
                                                ]
                                                
                                                if not daily_search.empty:
                                                    # 검색량 막대그래프 추가
                                                    fig.add_trace(go.Bar(
                                                        x=pd.to_datetime(daily_search['START_DT']),
                                                        y=daily_search['SRCH_CNT_TY'],
                                                        name=search_label,
                                                        marker_color='rgba(128, 128, 128, 0.3)',
                                                        opacity=0.6,
                                                        yaxis='y2'
                                                    ))
                                            
                                            # 그래프 제목 설정
                                            title = f"{brand_name} 일별 매출/검색량"
                                            
                                            # 레이아웃 설정
                                            fig.update_layout(
                                                title=title,
                                                xaxis_title="날짜",
                                                yaxis_title="매출액 (원)",
                                                xaxis=dict(
                                                    type='date',
                                                    showgrid=True,
                                                    gridcolor='lightgray'
                                                ),
                                                yaxis=dict(
                                                    showgrid=True,
                                                    gridcolor='lightgray',
                                                    tickformat=',.0f'
                                                ),
                                                yaxis2=dict(
                                                    title="검색량 (회)",
                                                    overlaying="y",
                                                    side="right",
                                                    showgrid=False,
                                                    tickformat=',.0f'
                                                ),
                                                hovermode='x unified',
                                                legend=dict(
                                                    orientation='h',
                                                    x=1,
                                                    y=1.1,
                                                    xanchor='right',
                                                    yanchor='bottom',
                                                    bgcolor='rgba(255,255,255,0.8)',
                                                    bordercolor='rgba(255,255,255,0)',
                                                    borderwidth=0
                                                ),
                                                height=400,
                                                margin=dict(t=100),
                                                template='plotly_white'
                                            )
                                            
                                            # 호버 템플릿 설정
                                            fig.update_traces(
                                                hovertemplate='<b>%{fullData.name}</b><br>' +
                                                             '날짜: %{x}<br>' +
                                                             '매출액: %{y:,.0f}원<br>' +
                                                             '<extra></extra>'
                                            )
                                            
                                            # 검색량 막대그래프 호버 템플릿 별도 설정
                                            fig.update_traces(
                                                selector=dict(type='bar'),
                                                hovertemplate='<b>%{fullData.name}</b><br>' +
                                                             '날짜: %{x}<br>' +
                                                             '검색량: %{y:,.0f}회<br>' +
                                                             '<extra></extra>'
                                            )
                                            
                                            st.plotly_chart(fig, use_container_width=True)
                                            
                                            # 매출 통계 요약 (선택된 아이템 반영)
                                            total_current = daily_sales['SALE_AMT_TY'].sum()
                                            
                                            # 이번주(10/20~10/26) 매출액 계산
                                            current_week_start = selected_week_start
                                            current_week_end = selected_week_end
                                            current_week_sales = three_weeks_sales[
                                                (three_weeks_sales['DT'] >= current_week_start) & 
                                                (three_weeks_sales['DT'] <= current_week_end)
                                            ]
                                            
                                            if not current_week_sales.empty:
                                                current_week_total = current_week_sales['SALE_AMT_TY'].sum()
                                                current_week_prev_year = current_week_sales['SALE_AMT_LY'].sum() if 'SALE_AMT_LY' in current_week_sales.columns else 0
                                                
                                                # 전년 대비 증감률 계산
                                                if current_week_prev_year > 0:
                                                    yoy_change = ((current_week_total - current_week_prev_year) / current_week_prev_year) * 100
                                                    yoy_display = f" (전년 대비 {yoy_change:+.1f}%)"
                                                else:
                                                    yoy_display = " (전년 데이터 없음)"
                                                
                                                # 이번주 매출액과 검색량을 나란히 표시
                                                col_sales, col_search = st.columns(2)
                                                
                                                with col_sales:
                                                    # 날짜 범위 포맷팅
                                                    date_range = f"{selected_week_start.strftime('%m/%d')}~{selected_week_end.strftime('%m/%d')}"
                                                    
                                                    # 이번주 매출액 표시
                                                    if selected_items and '전체' not in selected_items:
                                                        if len(selected_items) == 1:
                                                            label = f"{date_range} 매출액 ({selected_items[0]})"
                                                        else:
                                                            label = f"{date_range} 매출액 (선택 아이템)"
                                                    else:
                                                        label = f"{date_range} 매출액"
                                                    
                                                    st.metric(
                                                        label=f"**{label}**",
                                                        value=f"{current_week_total:,.0f}원{yoy_display}"
                                                    )
                                                
                                                with col_search:
                                                    if selected_categories and '전체' not in selected_categories:
                                                        # 선택된 카테고리 검색량 표시
                                                        st.metric(
                                                            label=f"**{date_range} 검색량 ({','.join(selected_categories)})**",
                                                            value=f"{current_week_search_volume:,.0f}회{search_yoy_display}"
                                                        )
                                                    else:
                                                        # 전체 검색량 계산
                                                        total_current_week_search = brand_search_data[
                                                            (pd.to_datetime(brand_search_data['START_DT']) >= pd.Timestamp(selected_week_start)) &
                                                            (pd.to_datetime(brand_search_data['START_DT']) <= pd.Timestamp(selected_week_end))
                                                        ]
                                                        total_current_week_search_volume = total_current_week_search['SRCH_CNT_TY'].sum() if not total_current_week_search.empty else 0
                                                        
                                                        # 전년 동일주 전체 검색량 계산 (SRCH_CNT_LY 사용)
                                                        total_prev_year_week_search_volume = brand_search_data['SRCH_CNT_LY'].sum() if not brand_search_data.empty else 0
                                                        
                                                        # 전년 대비 증감률 계산
                                                        total_search_yoy_display = ""
                                                        if total_prev_year_week_search_volume > 0:
                                                            total_search_yoy_change = ((total_current_week_search_volume - total_prev_year_week_search_volume) / total_prev_year_week_search_volume) * 100
                                                            total_search_yoy_display = f" (전년 대비 {total_search_yoy_change:+.1f}%)"
                                                        elif total_current_week_search_volume > 0:
                                                            total_search_yoy_display = " (신규)"
                                                        else:
                                                            total_search_yoy_display = " (-)"
                                                        
                                                        st.metric(
                                                            label=f"**{date_range} 전체 검색량**",
                                                            value=f"{total_current_week_search_volume:,.0f}회{total_search_yoy_display}"
                                                        )
                                    else:
                                        st.info("해당 기간의 매출 데이터가 없습니다.")
                                else:
                                    st.info("매출 데이터에 날짜 정보가 없습니다.")
                            else:
                                st.info("매출 데이터가 없습니다.")
                    else:
                        st.warning("해당 주차 데이터가 없습니다.")
                    
                    # 전주차 대비 랭킹 상승 키워드 분석
                    if not current_brand_ranking.empty and not prev_brand_ranking.empty:
                        st.markdown("#### 🔍 전주차 대비 랭킹 상승 키워드 분석")
                        
                        # 현재 주차와 전주차 데이터를 키워드로 매칭
                        current_dict = dict(zip(current_brand_ranking['검색어'], current_brand_ranking['순위']))
                        prev_dict = dict(zip(prev_brand_ranking['검색어'], prev_brand_ranking['순위']))
                        
                        ranking_changes = []
                        for keyword in current_dict.keys():
                            if keyword in prev_dict:
                                current_rank = current_dict[keyword]
                                prev_rank = prev_dict[keyword]
                                rank_change = prev_rank - current_rank  # 양수면 상승, 음수면 하락
                                ranking_changes.append({
                                    '검색어': keyword,
                                    '현재순위': current_rank,
                                    '전주차순위': prev_rank,
                                    '순위변화': rank_change,
                                    '현재검색량': current_brand_ranking[current_brand_ranking['검색어'] == keyword]['SRCH_CNT_TY'].iloc[0] if not current_brand_ranking[current_brand_ranking['검색어'] == keyword].empty else 0
                                })
                        
                        # 순위 상승 키워드만 필터링 (2단계 이상 상승, 상승폭이 큰 순으로 정렬)
                        rising_keywords = [item for item in ranking_changes if item['순위변화'] >= 2]
                        rising_keywords.sort(key=lambda x: x['순위변화'], reverse=True)
                        
                        if rising_keywords:
                            # 상위 5개 상승 키워드 표시
                            st.markdown("**📈 주요 상승 키워드 (상위 5개)**")
                            for idx, item in enumerate(rising_keywords[:5], 1):
                                st.markdown(f"**{idx}.** {item['검색어']} - {item['전주차순위']}위 → {item['현재순위']}위 (▲{item['순위변화']}단계 상승)")
                            
                        else:
                            st.info("전주차 대비 순위 상승 키워드가 없습니다.")
                    
                    # 브랜드 간 구분선
                    if i < len(brands) - 1:
                        st.markdown("---")
    else:
        st.info("검색량 데이터가 없습니다. '데이터 업로드 관리' 메뉴에서 먼저 데이터를 업로드하세요.")

# =============================================================================
# 메인 함수
# =============================================================================

def main():
    """메인 함수"""
    # 페이지 설정
    st.set_page_config(
        page_title="마케팅 분석 시스템",
        page_icon="👥",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # 사이드바
    st.sidebar.markdown("## 📊 마케팅 분석 시스템")
    
    # 메뉴 섹션
    st.sidebar.markdown("### ★ 메뉴")
    
    # 메뉴 버튼들
    if st.sidebar.button("👥 F&F CREW LIST", use_container_width=True):
        st.session_state.selected_menu = "👥 F&F CREW LIST"
    
    
    if st.sidebar.button("📈 데이터 업로드 관리", use_container_width=True):
        st.session_state.selected_menu = "📈 데이터 업로드 관리"
    
    if st.sidebar.button("📊 대시보드", use_container_width=True):
        st.session_state.selected_menu = "📊 대시보드"
    
    if st.sidebar.button("💰 매출대시보드", use_container_width=True):
        st.session_state.selected_menu = "💰 매출대시보드"
    
    if st.sidebar.button("🔍 검색량분석", use_container_width=True):
        st.session_state.selected_menu = "🔍 검색량분석"
    
    # 데이터 문의 기능 (사이드바 하단)
    st.sidebar.markdown("---")
    st.sidebar.markdown("### 💬 데이터 문의")
    question = st.sidebar.text_area("데이터에 대해 궁금한 점을 물어보세요:", height=100)
    if st.sidebar.button("🔍 분석"):
        if question:
            execution_df = load_execution_data()
            influencer_df = load_influencer_data()
            answer = analyze_data_question(question, execution_df, influencer_df)
            st.sidebar.success("분석 완료!")
            st.sidebar.markdown(f"**답변:** {answer}")
    
    # 기본 메뉴 설정
    if 'selected_menu' not in st.session_state:
        st.session_state.selected_menu = "👥 F&F CREW LIST"
    
    selected_menu = st.session_state.selected_menu
    
    # 데이터 로드
    df = load_influencer_data()
    
    # 메뉴별 렌더링
    if selected_menu == "👥 F&F CREW LIST":
        render_influencer_tab(df)
    elif selected_menu == "📈 데이터 업로드 관리":
        render_execution_data_management_tab()
    elif selected_menu == "📊 대시보드":
        render_dashboard_tab()
    elif selected_menu == "💰 매출대시보드":
        render_sales_dashboard_tab()
    elif selected_menu == "🔍 검색량분석":
        render_search_analysis_tab()

def get_season_options(df):
    """시즌 옵션 생성"""
    return ["25FW", "26SS"]

def filter_by_season(influencer_summary, df, months):
    """시즌별 필터링"""
    # 월별 배정 데이터에서 해당 월들의 데이터만 필터링
    return influencer_summary

def add_brand_details(influencer_summary, df, selected_brand_filter):
    """브랜드별 상세 정보 추가"""
    if selected_brand_filter == "전체":
        # 전체 브랜드 정보 추가
        for brand in BRANDS:
            qty_col = f"{brand.lower()}_qty"
            if qty_col in df.columns:
                influencer_summary[f"{brand}_계약수"] = df[qty_col]
    else:
        # 선택된 브랜드 정보만 추가
        qty_col = f"{selected_brand_filter.lower()}_qty"
        if qty_col in df.columns:
            influencer_summary[f"{selected_brand_filter}_계약수"] = df[qty_col]


def reorder_columns(influencer_summary):
    """컬럼 순서 조정: 브랜드별 계약수 컬럼들을 월별 컬럼들 앞으로 이동"""
    # 기본 컬럼들 (순서 유지)
    basic_columns = ["contract_id", "sns_id", "name", "follower", "unit_fee", "sec_usage", "sec_period", "전체_계약수"]
    
    # 브랜드별 계약수 컬럼들
    brand_contract_columns = [col for col in influencer_summary.columns if col.endswith("_계약수")]
    
    
    # 월별 컬럼들 (1월~12월)
    monthly_columns = [col for col in influencer_summary.columns if col in ["1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]]
    
    # 기타 컬럼들
    other_columns = [col for col in influencer_summary.columns if col not in basic_columns + brand_contract_columns + monthly_columns]
    
    # 최종 컬럼 순서 (중복 제거)
    final_columns = []
    for col in basic_columns + brand_contract_columns + monthly_columns + other_columns:
        if col in influencer_summary.columns and col not in final_columns:
            final_columns.append(col)
    
    return influencer_summary[final_columns]

def add_monthly_columns(influencer_summary, df, selected_brand_filter, selected_season_filter=None):
    """월별 컬럼 추가 - CREW 배정관리의 배정 데이터 반영"""
    # 시즌에 따른 월별 컬럼 설정
    if selected_season_filter == "25FW":
        months = ["9월", "10월", "11월", "12월", "1월", "2월"]
    elif selected_season_filter == "26SS":
        months = ["3월", "4월", "5월", "6월", "7월", "8월"]
    else:
        # 기본값: 25FW 시즌
        months = ["9월", "10월", "11월", "12월", "1월", "2월"]
    
    for month in months:
        influencer_summary[month] = ""
    
    # CREW 배정관리의 배정 데이터 로드 및 반영
    try:
        if os.path.exists(ASSIGNMENT_FILE):
            assignment_data = pd.read_csv(ASSIGNMENT_FILE, encoding="utf-8")
            if not assignment_data.empty and 'contract_id' in assignment_data.columns:
                # 배정 데이터를 인플루언서별, 월별로 그룹화
                assignment_pivot = assignment_data.groupby(['contract_id', '배정월'])['브랜드'].apply(lambda x: ', '.join(sorted(x.unique(), key=lambda b: ["MLB", "DX", "DV", "ST"].index(b) if b in ["MLB", "DX", "DV", "ST"] else 999))).reset_index()
                
                # 인플루언서별로 월별 배정 정보 매핑
                for _, row in influencer_summary.iterrows():
                    influencer_id = row["contract_id"]
                    influencer_assignments = assignment_pivot[assignment_pivot['contract_id'] == influencer_id]
                    
                    for _, assignment in influencer_assignments.iterrows():
                        month = assignment['배정월']
                        brands = assignment['브랜드']
                        if month in months:
                            influencer_summary.loc[influencer_summary['contract_id'] == influencer_id, month] = brands
    except Exception as e:
        print(f"월별 컬럼 추가 중 오류: {e}")

def render_influencer_table(influencer_summary, selected_brand_filter, selected_season_filter, influencer_count=None):
    """인플루언서 테이블 렌더링"""
    if influencer_count:
        st.markdown(f"**총 {influencer_count}명의 인플루언서**")
    
    # 컬럼 설정
    column_config = {
        "contract_id": st.column_config.NumberColumn("contract_id", width="small"),
        "SNS ID": st.column_config.TextColumn("SNS ID", width="medium"),
        "이름": st.column_config.TextColumn("이름", width="medium"),
        "FLW": st.column_config.NumberColumn("팔로워", format="%d", width="medium"),
        "1회계약단가": st.column_config.NumberColumn("1회계약단가", format="%d원", width="medium"),
        "2차활용": st.column_config.TextColumn("2차활용", width="small"),
        "2차기간": st.column_config.TextColumn("2차기간", width="small"),
        "최종상태": st.column_config.TextColumn("최종상태", width="small")
    }
    
    # 월별 컬럼 설정
    for month in MONTHS:
        column_config[month] = st.column_config.TextColumn(month, width="small")
    
    st.dataframe(
        influencer_summary,
        use_container_width=True,
        column_config=column_config,
        hide_index=True,
        height=600  # 테이블 높이를 600px로 설정
    )

def render_influencer_tab(df):
    """F&F CREW LIST 탭 렌더링"""
    st.markdown("# 👥 F&F CREW LIST")
    
    if df.empty:
        st.warning("인플루언서 데이터가 없습니다.")
        return
    
    # 필터 섹션
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**🏷️ 브랜드**")
        selected_brand_filter = st.selectbox("브랜드를 선택하세요", BRAND_OPTIONS, key="crew_brand_filter", label_visibility="collapsed")
    
    with col2:
        st.markdown("**📅 시즌**")
        season_options = get_season_options(df)
        selected_season_filter = st.selectbox("시즌을 선택하세요", season_options, key="crew_season_filter", label_visibility="collapsed", index=0)
    
    # 시즌 필터 적용 (항상 선택된 시즌으로 필터링)
    if 'contract_sesn' in df.columns:
        filtered_df = df[df['contract_sesn'] == selected_season_filter]
    else:
        filtered_df = df  # 시즌 컬럼이 없으면 전체 데이터 사용
    
    # 브랜드 필터 적용
    if selected_brand_filter != '전체':
        qty_col = f"{selected_brand_filter.lower()}_qty"
        if qty_col in filtered_df.columns:
            filtered_df = filtered_df[filtered_df[qty_col] > 0]
        else:
            filtered_df = pd.DataFrame()  # 해당 브랜드 계약수 컬럼이 없으면 빈 데이터프레임
    
    # 통계 계산
    total_influencers = len(filtered_df)
    
    # 총 계약수 계산
    total_contracts = 0
    if not filtered_df.empty:
        if selected_brand_filter != '전체':
            # 특정 브랜드 선택 시 해당 브랜드 계약수만 계산
            qty_col = f"{selected_brand_filter.lower()}_qty"
            if qty_col in filtered_df.columns:
                total_contracts = int(filtered_df[qty_col].sum())
        else:
            # 전체 선택 시 모든 브랜드 계약수 합산
            qty_cols = [f"{brand.lower()}_qty" for brand in BRANDS]
            available_qty_cols = [col for col in qty_cols if col in filtered_df.columns]
            if available_qty_cols:
                total_contracts = int(filtered_df[available_qty_cols].sum().sum())
    
    # NaN 값 처리
    if 'follower' in filtered_df.columns and not filtered_df['follower'].isna().all():
        avg_followers = int(filtered_df['follower'].mean()) if not pd.isna(filtered_df['follower'].mean()) else 0
    else:
        avg_followers = 0
        
    if 'unit_fee' in filtered_df.columns and not filtered_df['unit_fee'].isna().all():
        avg_contract_fee = int(filtered_df['unit_fee'].mean()) if not pd.isna(filtered_df['unit_fee'].mean()) else 0
    else:
        avg_contract_fee = 0
    
    # 통계 표시 (제목 없이)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("총 인플루언서 수", f"{total_influencers:,}명")
    with col2:
        st.metric("총 계약수", f"{total_contracts:,}건")
    with col3:
        st.metric("평균 팔로워 수", f"{avg_followers:,}")
    with col4:
        st.metric("평균 계약단가", f"{avg_contract_fee:,}원")
    
    st.markdown("---")
    
    # 인플루언서 데이터 준비 (필터링된 데이터 사용)
    influencer_summary = prepare_influencer_summary(filtered_df, selected_brand_filter, selected_season_filter)
    
    if influencer_summary.empty:
        st.warning("선택된 조건에 맞는 인플루언서 데이터가 없습니다.")
        return
    
    # 데이터 표시
    render_influencer_table(influencer_summary, selected_brand_filter, selected_season_filter, influencer_count=len(influencer_summary))
    
    # CREW 배정 관리 섹션 추가
    st.markdown("---")
    st.markdown("## 📋 CREW 배정 관리")
    
    # 월별 배정수량 표 생성
    st.markdown("### 📊 월별 배정수량 관리")
    
    # 시즌 선택
    season_options = ["25FW", "26SS"]
    selected_season = st.selectbox("시즌", season_options, key="assignment_season")
    
    # 시즌에 따른 월 설정
    if selected_season == "25FW":
        months = ["9월", "10월", "11월", "12월", "1월", "2월"]
    else:  # 26SS
        months = ["3월", "4월", "5월", "6월", "7월", "8월"]
    
    # 기존 목표 데이터 로드
    targets_df = load_monthly_targets()
    
    # 브랜드별 월별 배정수량 표 생성
    brands = ["MLB", "DX", "DV", "ST"]
    
    # 시즌별 월 가져오기
    months = SEASON_MONTHS.get(selected_season, ["9월", "10월", "11월", "12월", "1월", "2월"])
    
    # 현재 시즌/월에 맞는 데이터 필터링 또는 초기화
    if not targets_df.empty:
        targets_df_filtered = targets_df[targets_df['season'] == selected_season]
        
        # 현재 시즌의 모든 월이 포함되도록 데이터프레임 재구성
        current_targets_data = []
        for month in months:
            for brand in brands:
                # 월을 숫자로 변환 (9월 -> 9, 10월 -> 10, ...)
                month_num = int(month.replace('월', ''))
                
                # 기존 데이터에서 해당 월과 브랜드의 목표 수량 찾기
                existing_target = targets_df_filtered[
                    (targets_df_filtered['month'] == month_num) & 
                    (targets_df_filtered['brand'] == brand)
                ]
                if not existing_target.empty:
                    current_targets_data.append({
                        'month': month,
                        '브랜드': brand,
                        '목표수량': int(existing_target['target_quantity'].iloc[0])
                    })
                else:
                    current_targets_data.append({
                        'month': month,
                        '브랜드': brand,
                        '목표수량': 0
                    })
        
        current_targets_df = pd.DataFrame(current_targets_data)
        
        # 피벗 테이블 형태로 변환
        pivot_df = current_targets_df.pivot_table(
            index='month', columns='브랜드', values='목표수량'
        ).reindex(months, axis=0).reindex(brands, axis=1).fillna(0).astype(int)
    else:
        # 초기 데이터프레임 생성
        pivot_data = []
        for month in months:
            row = {"월": month}
            for brand in brands:
                row[brand] = 0
            pivot_data.append(row)
        
        pivot_df = pd.DataFrame(pivot_data)
        pivot_df = pivot_df.set_index("월")
    
    # 편집 가능한 데이터프레임
    st.markdown("#### 배정수량 입력")
    edited_df = st.data_editor(
        pivot_df,
        use_container_width=True,
        key="monthly_targets_editor",
        hide_index=False,
        column_config={
            col: st.column_config.NumberColumn(col, min_value=0) 
            for col in pivot_df.columns
        }
    )
    
    # 버튼들
    col1, col2, col3, col_spacer, col4 = st.columns([0.2, 0.2, 0.1, 0.1, 0.4])
    
    with col1:
        if st.button("💾 배정수량 저장", type="secondary", use_container_width=True):
            # edited_df를 원래 형식으로 변환하여 저장
            saved_targets = edited_df.stack().reset_index()
            saved_targets.columns = ['month', 'brand', 'target_quantity']
            
            # 월을 숫자로 변환 (9월 -> 9, 10월 -> 10, ...)
            saved_targets['month'] = saved_targets['month'].str.replace('월', '').astype(int)
            saved_targets['season'] = selected_season  # 시즌 정보 추가
            
            # 기존 데이터에서 현재 시즌 데이터만 업데이트
            if not targets_df.empty:
                other_season_targets = targets_df[targets_df['season'] != selected_season]
                updated_targets_df = pd.concat([other_season_targets, saved_targets], ignore_index=True)
            else:
                updated_targets_df = saved_targets
            
            save_monthly_targets(updated_targets_df)
            st.success("✅ 배정수량이 저장되었습니다.")
            st.rerun()
    
    with col2:
        if st.button("🚀 자동배정실행", type="secondary", use_container_width=True):
            # 자동배정 실행
            execute_monthly_automatic_assignment_from_table(edited_df, df)
    
    with col3:
        pass  # 빈 공간
    
    with col4:
        pass  # 빈 공간
    
    # 인플루언서별 월별 배정 내역 (메인 섹션)
    st.markdown("---")
    st.subheader("👥 인플루언서별 월별 배정 내역")
    
    # 배정 이력이 있는 경우 분석
    if os.path.exists('data/assignment_history.csv'):
        try:
            assignment_df = pd.read_csv('data/assignment_history.csv', encoding="utf-8")
            
            if not assignment_df.empty:
                # 시즌 필터링 적용 (맨 위의 시즌 필터 사용)
                selected_season = st.session_state.get('assignment_season', '25FW')
                
                if '시즌' in assignment_df.columns:
                    original_count = len(assignment_df)
                    assignment_df = assignment_df[assignment_df['시즌'] == selected_season]
                    filtered_count = len(assignment_df)
                    st.info(f"📊 총 {filtered_count}건의 배정이 완료되었습니다")
                    
                    # 시즌에 배정 데이터가 없으면 빈 테이블 표시
                    if assignment_df.empty:
                        st.info("📊 아직 배정된 데이터가 없습니다")
                        return
                else:
                    # 시즌 컬럼이 없으면 기존 데이터는 25FW로 간주
                    if selected_season != '25FW':
                        st.info("📊 아직 배정된 데이터가 없습니다")
                        return
                
                # 상태 컬럼이 있으면 배정완료 상태만 필터링, 없으면 전체 사용
                if '상태' in assignment_df.columns:
                    completed_assignments = assignment_df[assignment_df['상태'] == '📋 배정완료']
                else:
                    completed_assignments = assignment_df
                
                # 인플루언서별 월별 배정 내역 표시
                if not completed_assignments.empty:
                    # 인플루언서별 월별 배정 내역 생성
                    assignment_summary = create_assignment_summary(completed_assignments, selected_season, df)
                    if not assignment_summary.empty:
                        st.dataframe(assignment_summary, use_container_width=True, height=400, hide_index=True)
                        
                        # 엑셀 다운로드 및 배정초기화 버튼
                        col1, col2, col3, col_spacer, col4 = st.columns([0.2, 0.2, 0.1, 0.1, 0.4])
                        
                        with col1:
                            excel_buffer = create_excel_with_two_sheets(assignment_summary, completed_assignments)
                            
                            st.download_button(
                                label="📊 엑셀 다운로드",
                                data=excel_buffer,
                                file_name=f"assignment_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="secondary",
                                use_container_width=True
                            )
                        
                        with col2:
                            if st.button("🗑️ 배정초기화", type="secondary", use_container_width=True):
                                # 배정 이력만 초기화 (배정수량 입력 데이터는 보존)
                                if os.path.exists("data/assignment_history.csv"):
                                    os.remove("data/assignment_history.csv")
                                
                                st.success("✅ 배정 이력이 초기화되었습니다.")
                                st.rerun()
                    else:
                        st.info("📊 아직 배정된 데이터가 없습니다")
                else:
                    st.info("📊 아직 배정된 데이터가 없습니다")
                
            else:
                st.info("📊 아직 배정된 데이터가 없습니다")
        except Exception as e:
            st.warning(f"배정 이력 분석 중 오류가 발생했습니다: {e}")
    else:
        st.info("📊 아직 배정된 데이터가 없습니다")
    
    # 배정현황모니터링
    st.markdown("---")
    st.subheader("📊 배정현황모니터링")
    
    # 그래프 통계 섹션
    st.markdown("#### 📈 그래프 통계")
    
    # 배정 이력이 있는 경우 그래프 표시
    if os.path.exists('data/assignment_history.csv'):
        try:
            assignment_df = pd.read_csv('data/assignment_history.csv', encoding="utf-8")
            
            if not assignment_df.empty:
                # 시즌 필터링 적용
                selected_season = st.session_state.get('assignment_season', '25FW')
                
                if '시즌' in assignment_df.columns:
                    assignment_df = assignment_df[assignment_df['시즌'] == selected_season]
                
                # 상태 컬럼이 있으면 배정완료 상태만 필터링
                if '상태' in assignment_df.columns:
                    assignment_df = assignment_df[assignment_df['상태'] == '📋 배정완료']
                
                if not assignment_df.empty:
                    # 브랜드별 배정 현황
                    brand_counts = assignment_df['브랜드'].value_counts()
                    
                    # 그래프 생성
                    fig = px.pie(
                        values=brand_counts.values,
                        names=brand_counts.index,
                        title="브랜드별 배정 현황"
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    
                    # 월별 배정 현황
                    if '배정월' in assignment_df.columns:
                        month_counts = assignment_df['배정월'].value_counts()
                        
                        # 월 순서 정렬
                        month_order = ["9월", "10월", "11월", "12월", "1월", "2월"] if selected_season == "25FW" else ["3월", "4월", "5월", "6월", "7월", "8월"]
                        month_counts = month_counts.reindex([m for m in month_order if m in month_counts.index])
                        
                        fig2 = px.bar(
                            x=month_counts.index,
                            y=month_counts.values,
                            title="월별 배정 현황"
                        )
                        st.plotly_chart(fig2, use_container_width=True)
                else:
                    st.info("📊 아직 배정된 데이터가 없습니다")
            else:
                st.info("📊 아직 배정된 데이터가 없습니다")
        except Exception as e:
            st.warning(f"그래프 생성 중 오류가 발생했습니다: {e}")
    else:
        st.info("📊 아직 배정된 데이터가 없습니다")
    

def prepare_influencer_summary(df, selected_brand_filter, selected_season_filter):
    """인플루언서 요약 데이터 준비"""
    if df.empty:
        return pd.DataFrame()
    
    # 기본 컬럼 선택
    required_columns = ["contract_id", "sns_id", "name", "follower", "unit_fee", "sec_usage", "sec_period"]
    available_columns = [col for col in required_columns if col in df.columns]
    
    if not available_columns:
        st.error("필요한 컬럼이 데이터에 없습니다.")
        return pd.DataFrame()
    
    influencer_summary = df[available_columns].copy()
    
    # 전체 계약수 계산
    qty_cols = [f"{brand.lower()}_qty" for brand in BRANDS]
    available_qty_cols = [col for col in qty_cols if col in df.columns]
    if available_qty_cols:
        influencer_summary["전체_계약수"] = df[available_qty_cols].sum(axis=1)
    
    # 시즌 필터 적용 (이미 filtered_df로 필터링됨)
    if selected_season_filter != "전체":
        # 시즌별 월 컬럼 추가
        months = SEASON_MONTHS.get(selected_season_filter, ["9월", "10월", "11월", "12월", "1월", "2월"])
        for month in months:
            influencer_summary[month] = ""
    
    # 브랜드 필터 적용
    if selected_brand_filter != "전체":
        qty_col = f"{selected_brand_filter.lower()}_qty"
        if qty_col in df.columns:
            brand_filter_mask = df[qty_col] > 0
            influencer_summary = influencer_summary[brand_filter_mask]
    
    # 브랜드별 상세 정보 추가
    add_brand_details(influencer_summary, df, selected_brand_filter)
    
    
    # 월별 컬럼 추가 (컬럼명 변경 전에) - 시즌 필터에 따라 추가
    add_monthly_columns(influencer_summary, df, selected_brand_filter, selected_season_filter)
    
    # 컬럼 순서 조정: 브랜드별 계약수 컬럼들을 월별 컬럼들 앞으로 이동
    influencer_summary = reorder_columns(influencer_summary)
    
    # 컬럼명 변경
    influencer_summary = influencer_summary.rename(columns={
        "contract_id": "contract_id", "sns_id": "SNS ID", "name": "이름", "follower": "FLW", "unit_fee": "1회계약단가", 
        "sec_usage": "2차활용", "sec_period": "2차기간"
    })
    
    return influencer_summary

def render_assignment_tab():
    """CREW 배정 관리 탭 렌더링"""
    st.markdown("# 📋 CREW 배정 관리")
    
    # 인플루언서 데이터 로드
    df = load_influencer_data()
    if df.empty:
        st.warning("인플루언서 데이터가 없습니다.")
        return
    
    # 월별 배정수량 표 생성
    st.markdown("## 📊 월별 배정수량 관리")
    
    # 시즌 선택
    season_options = ["25FW", "26SS"]
    selected_season = st.selectbox("시즌", season_options, key="assignment_season")
    
    # 시즌에 따른 월 설정
    if selected_season == "25FW":
        months = ["9월", "10월", "11월", "12월", "1월", "2월"]
    else:  # 26SS
        months = ["3월", "4월", "5월", "6월", "7월", "8월"]
    
    # 기존 목표 데이터 로드
    targets_df = load_monthly_targets()
    
    # 브랜드별 월별 배정수량 표 생성
    brands = ["MLB", "DX", "DV", "ST"]
    
    # 시즌별 월 가져오기
    months = SEASON_MONTHS.get(selected_season, ["9월", "10월", "11월", "12월", "1월", "2월"])
    
    # 현재 시즌/월에 맞는 데이터 필터링 또는 초기화
    if not targets_df.empty:
        targets_df_filtered = targets_df[targets_df['season'] == selected_season]
        
        # 현재 시즌의 모든 월이 포함되도록 데이터프레임 재구성
        current_targets_data = []
        for month in months:
            for brand in brands:
                # 월을 숫자로 변환 (9월 -> 9, 10월 -> 10, ...)
                month_num = int(month.replace('월', ''))
                
                # 기존 데이터에서 해당 월과 브랜드의 목표 수량 찾기
                existing_target = targets_df_filtered[
                    (targets_df_filtered['month'] == month_num) & 
                    (targets_df_filtered['brand'] == brand)
                ]
                if not existing_target.empty:
                    current_targets_data.append({
                        'month': month,
                        '브랜드': brand,
                        '목표수량': int(existing_target['target_quantity'].iloc[0])
                    })
                else:
                    current_targets_data.append({
                        'month': month,
                        '브랜드': brand,
                        '목표수량': 0
                    })
        
        current_targets_df = pd.DataFrame(current_targets_data)
        
        # 피벗 테이블 형태로 변환
        pivot_df = current_targets_df.pivot_table(
            index='month', columns='브랜드', values='목표수량'
        ).reindex(months, axis=0).reindex(brands, axis=1).fillna(0).astype(int)
    else:
        # 초기 데이터프레임 생성
        pivot_data = []
        for month in months:
            row = {"월": month}
            for brand in brands:
                row[brand] = 0
            pivot_data.append(row)
        
        pivot_df = pd.DataFrame(pivot_data)
        pivot_df = pivot_df.set_index("월")
    
    # 편집 가능한 데이터프레임
    st.markdown("### 배정수량 입력")
    edited_df = st.data_editor(
        pivot_df,
        use_container_width=True,
        key="monthly_targets_editor",
        hide_index=False,
        column_config={
            col: st.column_config.NumberColumn(col, min_value=0) 
            for col in pivot_df.columns
        }
    )
    
    # 버튼들
    col1, col2, col3, col_spacer, col4 = st.columns([0.2, 0.2, 0.1, 0.1, 0.4])
    
    with col1:
        if st.button("💾 배정수량 저장", type="secondary", use_container_width=True):
            # edited_df를 원래 형식으로 변환하여 저장
            saved_targets = edited_df.stack().reset_index()
            saved_targets.columns = ['month', 'brand', 'target_quantity']
            
            # 월을 숫자로 변환 (9월 -> 9, 10월 -> 10, ...)
            saved_targets['month'] = saved_targets['month'].str.replace('월', '').astype(int)
            saved_targets['season'] = selected_season  # 시즌 정보 추가
            
            # 기존 데이터에서 현재 시즌 데이터만 업데이트
            if not targets_df.empty:
                other_season_targets = targets_df[targets_df['season'] != selected_season]
                updated_targets_df = pd.concat([other_season_targets, saved_targets], ignore_index=True)
            else:
                updated_targets_df = saved_targets
            
            save_monthly_targets(updated_targets_df)
            st.success("✅ 배정수량이 저장되었습니다.")
            st.rerun()
    
    with col2:
        if st.button("🚀 자동배정실행", type="secondary", use_container_width=True):
            # 자동배정 실행
            execute_monthly_automatic_assignment_from_table(edited_df, df)
    
    with col3:
        pass  # 빈 공간
    
    with col4:
        pass  # 빈 공간
    
    # 인플루언서별 월별 배정 내역 (메인 섹션)
    st.markdown("---")
    st.subheader("👥 인플루언서별 월별 배정 내역")
    
    # 배정 이력이 있는 경우 분석
    if os.path.exists('data/assignment_history.csv'):
        try:
            assignment_df = pd.read_csv('data/assignment_history.csv', encoding="utf-8")
            
            if not assignment_df.empty:
                # 시즌 필터링 적용 (맨 위의 시즌 필터 사용)
                selected_season = st.session_state.get('assignment_season', '25FW')
                
                if '시즌' in assignment_df.columns:
                    original_count = len(assignment_df)
                    assignment_df = assignment_df[assignment_df['시즌'] == selected_season]
                    filtered_count = len(assignment_df)
                    st.info(f"📊 총 {filtered_count}건의 배정이 완료되었습니다")
                    
                    # 시즌에 배정 데이터가 없으면 빈 테이블 표시
                    if assignment_df.empty:
                        st.info("📊 아직 배정된 데이터가 없습니다")
                        return
                else:
                    # 시즌 컬럼이 없으면 기존 데이터는 25FW로 간주
                    if selected_season != '25FW':
                        st.info("📊 아직 배정된 데이터가 없습니다")
                        return
                
                # 상태 컬럼이 있으면 배정완료 상태만 필터링, 없으면 전체 사용
                if '상태' in assignment_df.columns:
                    completed_assignments = assignment_df[assignment_df['상태'] == '📋 배정완료']
                else:
                    completed_assignments = assignment_df
                
                # 인플루언서별 월별 배정 내역 표시
                if not completed_assignments.empty:
                    # 인플루언서별 월별 배정 내역 생성
                    assignment_summary = create_assignment_summary(completed_assignments, selected_season, df)
                    if not assignment_summary.empty:
                        st.dataframe(assignment_summary, use_container_width=True, height=400, hide_index=True)
                        
                        # 엑셀 다운로드 및 배정초기화 버튼
                        col1, col2, col3, col_spacer, col4 = st.columns([0.2, 0.2, 0.1, 0.1, 0.4])
                        
                        with col1:
                            excel_buffer = create_excel_with_two_sheets(assignment_summary, completed_assignments)
                            
                            st.download_button(
                                label="📊 엑셀 다운로드",
                                data=excel_buffer,
                                file_name=f"assignment_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="secondary",
                                use_container_width=True
                            )
                        
                        with col2:
                            if st.button("🗑️ 배정초기화", type="secondary", use_container_width=True):
                                # 배정 이력만 초기화 (배정수량 입력 데이터는 보존)
                                if os.path.exists("data/assignment_history.csv"):
                                    os.remove("data/assignment_history.csv")
                                
                                st.success("✅ 배정 이력이 초기화되었습니다.")
                                st.rerun()
                    else:
                        st.info("📊 아직 배정된 데이터가 없습니다")
                else:
                    st.info("📊 아직 배정된 데이터가 없습니다")
                
            else:
                st.info("📊 아직 배정된 데이터가 없습니다")
        except Exception as e:
            st.warning(f"배정 이력 분석 중 오류가 발생했습니다: {e}")
    else:
        st.info("📊 아직 배정된 데이터가 없습니다")
    
    # 배정현황모니터링
    st.markdown("---")
    st.subheader("📊 배정현황모니터링")
    
    # 그래프 통계 섹션
    st.markdown("#### 📈 그래프 통계")
    
    # 배정 이력이 있는 경우 그래프 표시
    if os.path.exists('data/assignment_history.csv'):
        try:
            assignment_df = pd.read_csv('data/assignment_history.csv', encoding="utf-8")
            
            if not assignment_df.empty and '배정월' in assignment_df.columns and '브랜드' in assignment_df.columns:
                # 시즌 필터링 적용 (맨 위의 시즌 필터 사용)
                if '시즌' in assignment_df.columns:
                    # 맨 위의 시즌 필터 값 가져오기
                    selected_season = st.session_state.get('assignment_season', '25FW')
                    assignment_df = assignment_df[assignment_df['시즌'] == selected_season]
                    
                    # 시즌에 배정 데이터가 없으면 빈 그래프 표시
                    if assignment_df.empty:
                        return
                
                # 상태 컬럼이 있으면 배정완료 상태만 필터링, 없으면 전체 사용
                if '상태' in assignment_df.columns:
                    completed_assignments = assignment_df[assignment_df['상태'] == '📋 배정완료']
                else:
                    completed_assignments = assignment_df
                
                if not completed_assignments.empty:
                    # 2개 컬럼으로 그래프 나란히 배치
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # 브랜드별 배정 현황 그래프
                        brand_counts = completed_assignments['브랜드'].value_counts()
                        
                        # 막대 그래프 생성
                        import plotly.express as px
                        fig = px.bar(
                            x=brand_counts.index, 
                            y=brand_counts.values,
                            title="브랜드별 배정 현황",
                            labels={'x': '브랜드', 'y': '배정 수량'},
                            color_discrete_sequence=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
                        )
                        fig.update_layout(
                            showlegend=False,
                            height=400,
                            title_x=0.5,
                            margin=dict(t=50, b=50, l=50, r=50)  # 상단 여백 증가
                        )
                        # 막대 위에 숫자 레이블 추가
                        fig.update_traces(
                            text=brand_counts.values,
                            textposition='outside'
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    
                    with col2:
                        # 월별 배정 현황 그래프
                        month_counts = completed_assignments['배정월'].value_counts()
                        month_order = ['9월', '10월', '11월', '12월', '1월', '2월']
                        month_counts = month_counts.reindex(month_order, fill_value=0)
                        
                        fig2 = px.bar(
                            x=month_counts.index, 
                            y=month_counts.values,
                            title="월별 배정 현황",
                            labels={'x': '월', 'y': '배정 수량'},
                            color_discrete_sequence=['#2ca02c', '#ff7f0e', '#1f77b4', '#d62728', '#9467bd', '#8c564b']
                        )
                        fig2.update_layout(
                            showlegend=False,
                            height=400,
                            title_x=0.5,
                            margin=dict(t=50, b=50, l=50, r=50)  # 상단 여백 증가
                        )
                        # 막대 위에 숫자 레이블 추가
                        fig2.update_traces(
                            text=month_counts.values,
                            textposition='outside'
                        )
                        st.plotly_chart(fig2, use_container_width=True)
                else:
                    st.info("배정된 데이터가 없습니다.")
            else:
                st.info("배정 이력에 필요한 컬럼이 없습니다.")
        except Exception as e:
            st.warning(f"그래프 생성 중 오류가 발생했습니다: {e}")
    else:
        st.info("아직 배정 이력이 없습니다.")
    
    # 배정요청수량 vs 배정수량 비교 테이블 (맨 아래에 표시)
    st.markdown("#### 📈 배정요청수량 vs 배정수량 비교")
    
    # 배정 이력이 있는 경우에만 비교 테이블 표시
    if os.path.exists('data/assignment_history.csv'):
        try:
            assignment_df = pd.read_csv('data/assignment_history.csv', encoding="utf-8")
            
            if not assignment_df.empty and '배정월' in assignment_df.columns and '브랜드' in assignment_df.columns:
                # 시즌 필터링 적용 (맨 위의 시즌 필터 사용)
                if '시즌' in assignment_df.columns:
                    # 맨 위의 시즌 필터 값 가져오기
                    selected_season = st.session_state.get('assignment_season', '25FW')
                    assignment_df = assignment_df[assignment_df['시즌'] == selected_season]
                    
                    # 시즌에 배정 데이터가 없으면 빈 비교 테이블 표시
                    if assignment_df.empty:
                        st.info(f"📊 {selected_season} 시즌에는 아직 배정 데이터가 없습니다.")
                        return
                
                # 요청수량 (edited_df에서)
                request_data = []
                for brand in brands:
                    for month in months:
                        target = int(edited_df.loc[month, brand])
                        request_data.append({
                            '브랜드': brand,
                            '월': month,
                            '요청수량': target
                        })
                
                request_df = pd.DataFrame(request_data)
                
                # 상태 컬럼이 있으면 배정완료 상태만 필터링, 없으면 전체 사용
                if '상태' in assignment_df.columns:
                    completed_assignments = assignment_df[assignment_df['상태'] == '📋 배정완료']
                else:
                    completed_assignments = assignment_df
                
                # 브랜드별, 월별 배정수량 집계
                actual_data = []
                for brand in brands:
                    for month in months:
                        brand_month_assignments = completed_assignments[
                            (completed_assignments['브랜드'] == brand) & 
                            (completed_assignments['배정월'] == month)
                        ]
                        actual_count = len(brand_month_assignments)
                        actual_data.append({
                            '브랜드': brand,
                            '월': month,
                            '배정수량': actual_count
                        })
                
                actual_df = pd.DataFrame(actual_data)
                
                # 요청수량과 배정수량 병합
                comparison_df = request_df.merge(actual_df, on=['브랜드', '월'], how='left')
                comparison_df['배정수량'] = comparison_df['배정수량'].fillna(0).astype(int)
                
                # 달성률 계산
                comparison_df['달성률'] = (comparison_df['배정수량'] / comparison_df['요청수량'] * 100).round(1)
                comparison_df['달성률'] = comparison_df['달성률'].replace([float('inf'), -float('inf')], 0)
                comparison_df['달성률'] = comparison_df['달성률'].fillna(0)
                
                # 상태 표시
                comparison_df['상태'] = comparison_df.apply(
                    lambda row: "✅ 완료" if row['배정수량'] >= row['요청수량'] else "❌ 미달성", 
                    axis=1
                )
                
                # 표시
                st.dataframe(comparison_df, use_container_width=True)
            else:
                st.info("배정 이력에 필요한 컬럼이 없습니다.")
        except Exception as e:
            st.warning(f"배정 비교 분석 중 오류가 발생했습니다: {e}")
    else:
        st.info("아직 배정 이력이 없습니다.")

def create_excel_with_two_sheets(assignment_summary, completed_assignments):
    """엑셀 파일을 두 개의 시트로 생성"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # 워크북 생성
    wb = Workbook()
    
    # 시트1: 인플루언서별 월별 배정 내역 (테이블 형태)
    ws1 = wb.active
    ws1.title = "인플루언서별 월별 배정내역"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 시트1 데이터 작성
    if not assignment_summary.empty:
        # 헤더 작성
        headers = assignment_summary.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        for row_idx, row in assignment_summary.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws1.cell(row=row_idx+2, column=col_idx, value=value)
        
        # 열 너비 자동 조정
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
    
    # 시트2: 개별 배정 내역 (로우 형태)
    ws2 = wb.create_sheet("개별 배정내역")
    
    # 시트2 데이터 작성
    if not completed_assignments.empty:
        # 필요한 컬럼만 선택
        display_columns = ['contract_id', 'SNS ID', '이름', '브랜드', '배정월', '배정일']
        available_columns = [col for col in display_columns if col in completed_assignments.columns]
        
        if available_columns:
            row_data = completed_assignments[available_columns].copy()
            
            # 헤더 작성
            for col_idx, header in enumerate(available_columns, 1):
                cell = ws2.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 데이터 작성
            for row_idx, row in row_data.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws2.cell(row=row_idx+2, column=col_idx, value=value)
            
            # 열 너비 자동 조정
            for column in ws2.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws2.column_dimensions[column_letter].width = adjusted_width
    
    # 메모리에 엑셀 파일 저장
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def create_assignment_summary(completed_assignments, selected_season=None, df=None):
    """인플루언서별 월별 배정 내역 생성"""
    if completed_assignments.empty:
        return pd.DataFrame()
    
    # 필요한 컬럼 확인
    required_columns = ['이름', '브랜드', '배정월']
    if not all(col in completed_assignments.columns for col in required_columns):
        return pd.DataFrame()
    
    # 시즌별 월 가져오기
    months = SEASON_MONTHS.get(selected_season, ["9월", "10월", "11월", "12월", "1월", "2월"])
    
    # 인플루언서별로 그룹화
    summary_data = []
    
    for name in completed_assignments['이름'].unique():
        influencer_data = completed_assignments[completed_assignments['이름'] == name]
        
        # 인플루언서 기본 정보 (첫 번째 행에서 가져오기)
        first_row = influencer_data.iloc[0]
        contract_id = first_row.get('contract_id', '')
        sns_id = first_row.get('SNS ID', '')
        
        # 월별 배정 브랜드 정리 (중복 제거)
        monthly_assignments = {}
        for month in months:
            month_data = influencer_data[influencer_data['배정월'] == month]
            if not month_data.empty:
                brands = month_data['브랜드'].tolist()
                # 중복 제거하고 브랜드 우선순위에 따라 정렬 (MLB, DX, DV, ST)
                unique_brands = list(set(brands))
                brand_order = ['MLB', 'DX', 'DV', 'ST']
                sorted_brands = [brand for brand in brand_order if brand in unique_brands]
                monthly_assignments[month] = ','.join(sorted_brands)
            else:
                monthly_assignments[month] = ''
        
        # 총계약수 계산 (influencer.csv의 실제 계약수)
        total_contracts = 0
        if df is not None and not df.empty:
            # contract_id로 해당 인플루언서 찾기
            if 'contract_id' in df.columns and contract_id:
                influencer_row = df[df['contract_id'] == contract_id]
                if not influencer_row.empty:
                    # 모든 브랜드 계약수 합계
                    qty_cols = [f"{brand.lower()}_qty" for brand in BRANDS]
                    available_qty_cols = [col for col in qty_cols if col in influencer_row.columns]
                    if available_qty_cols:
                        total_contracts = int(influencer_row[available_qty_cols].sum().sum())
        
        # 요약 데이터 생성
        summary_row = {
            'contract_id': contract_id,
            'SNS ID': sns_id,
            '이름': name,
            '총계약수': total_contracts
        }
        
        # 시즌별 월 컬럼 추가
        for month in months:
            summary_row[month] = monthly_assignments[month]
        
        summary_data.append(summary_row)
    
    return pd.DataFrame(summary_data)

def execute_monthly_automatic_assignment_from_table(edited_df, df):
    """월별 배정수량 표에서 자동배정 실행 (최적배정알고리즘 적용)"""
    try:
        import pulp
    except ImportError:
        st.error("❌ PuLP 라이브러리가 설치되지 않았습니다. 'pip install pulp' 명령으로 설치해주세요.")
        return
    
    # 인플루언서 데이터 확인
    if df.empty:
        st.error("❌ 인플루언서 데이터가 없습니다.")
        return
    
    # 현재 선택된 시즌 가져오기
    selected_season = st.session_state.get('assignment_season', '25FW')
    
    # 브랜드별 컬럼 매핑
    brand_columns = {
        'MLB': 'mlb_qty',
        'DX': 'dx_qty', 
        'DV': 'dv_qty',
        'ST': 'st_qty'
    }
    
    # 각 브랜드별로 최적배정 실행
    assignment_data = []
    
    for brand in edited_df.columns:
        # 해당 브랜드의 월별 목표 수량 확인
        brand_targets = edited_df[brand]
        if brand_targets.sum() <= 0:
            continue  # 요청이 안된 브랜드
        
        # 해당 브랜드의 인플루언서 필터링
        brand_column = brand_columns.get(brand)
        if not brand_column or brand_column not in df.columns:
            st.warning(f"⚠️ {brand} 브랜드의 컬럼을 찾을 수 없습니다.")
            continue
        
        brand_influencers = df[df[brand_column] > 0].copy()
        if brand_influencers.empty:
            st.warning(f"⚠️ {brand} 브랜드에 계약수량이 있는 인플루언서가 없습니다.")
            continue
        
        # 최적배정알고리즘 실행
        brand_assignments = execute_optimal_assignment_for_brand(
            brand, brand_influencers, brand_targets, brand_column, selected_season
        )
        assignment_data.extend(brand_assignments)
    
    # 배정 결과 저장
    if assignment_data:
        # DataFrame으로 변환
        assignment_df = pd.DataFrame(assignment_data)
        
        # 기존 배정 이력 로드
        existing_assignments = load_assignment_history()
        
        # 새로운 배정 이력을 기존 이력에 추가
        updated_assignments = pd.concat([existing_assignments, assignment_df], ignore_index=True)
        
        # 중복 제거 (동일 인플루언서-브랜드-월 중복 배정 방지)
        updated_assignments.drop_duplicates(
            subset=['contract_id', '브랜드', '배정월'], 
            keep='last', 
            inplace=True
        )
        
        # CSV로 저장
        save_assignment_history(updated_assignments)
    else:
        st.warning("배정할 데이터가 없습니다.")

def execute_optimal_assignment_for_brand(brand, influencers_df, monthly_targets, brand_column, selected_season):
    """특정 브랜드에 대한 최적배정 실행"""
    try:
        import pulp
        
        # 데이터 준비
        influencer_ids = influencers_df['contract_id'].tolist()
        months = monthly_targets.index.tolist()
        
        # 계약수량 매핑
        contract_qty = {row['contract_id']: row[brand_column] for _, row in influencers_df.iterrows()}
        
        # 월별 목표 수량 매핑
        target_qty = {month: int(monthly_targets[month]) for month in months}
        
        # 최적화 문제 정의
        prob = pulp.LpProblem(f"{brand}_Assignment", pulp.LpMaximize)
        
        # 의사결정 변수 생성
        x = pulp.LpVariable.dicts("assign", 
                                 [(i, j) for i in influencer_ids for j in months], 
                                 0, 1, pulp.LpBinary)
        
        # 목적 함수: 총 배정 수량 최대화
        prob += pulp.lpSum(x[i, j] for i in influencer_ids for j in months)
        
        # 제약 조건 1: 인플루언서별 계약수 제한
        for i in influencer_ids:
            prob += pulp.lpSum(x[i, j] for j in months) <= contract_qty[i]
        
        # 제약 조건 2: 월별 목표 수량 제한
        for j in months:
            if target_qty[j] > 0:
                prob += pulp.lpSum(x[i, j] for i in influencer_ids) == target_qty[j]
        
        # 제약 조건 3: 중복 배정 방지 (안전성 확보)
        # 동일 인플루언서는 동일 브랜드로 월 1회 이상 배정 불가
        for i in influencer_ids:
            for j in months:
                prob += x[i, j] <= 1  # 최대 1회 배정
                prob += x[i, j] >= 0  # 최소 0회 배정
        
        # 최적화 실행
        prob.solve()
        
        # 결과 처리
        assignment_data = []
        if pulp.LpStatus[prob.status] == 'Optimal':
            # 최적해를 배정 데이터로 변환
            for i in influencer_ids:
                for j in months:
                    if pulp.value(x[i, j]) == 1:
                        # 인플루언서 정보 찾기
                        influencer_info = influencers_df[influencers_df['contract_id'] == i].iloc[0]
                        
                        assignment_info = {
                            'contract_id': i,
                            'SNS ID': influencer_info.get('sns_id', ''),
                            '이름': influencer_info.get('name', ''),
                            '브랜드': brand,
                            '배정월': j,
                            '상태': '📋 배정완료',
                            '배정일': pd.Timestamp.now().strftime('%Y-%m-%d'),
                            '시즌': selected_season
                        }
                        assignment_data.append(assignment_info)
        else:
            st.warning(f"⚠️ {brand} 브랜드의 최적해를 찾을 수 없습니다.")
        
        return assignment_data
        
    except Exception as e:
        st.error(f"❌ {brand} 브랜드 최적배정 중 오류 발생: {e}")
        return []

def execute_automatic_assignment(selected_month, selected_season, quantities, df):
    """자동 배정 실행"""
    st.info("자동 배정이 실행되었습니다.")
    st.success("배정이 완료되었습니다.")

def render_manual_assignment_section(selected_month, selected_season):
    """수동 배정 섹션 렌더링"""
    st.markdown("## 👤 수동 배정")
    st.info("수동 배정 기능은 추후 구현 예정입니다.")

def analyze_advanced_statistics(execution_df, sales_df, influencer_df, question_lower):
    """고급 통계 분석"""
    results = []
    
    # 상관관계 분석
    if "상관관계" in question_lower or "연관성" in question_lower:
        if not execution_df.empty:
            # 숫자형 컬럼들만 선택
            numeric_cols = execution_df.select_dtypes(include=[np.number]).columns
            if len(numeric_cols) > 1:
                correlation_matrix = execution_df[numeric_cols].corr()
                
                # 높은 상관관계 찾기
                high_corr = []
                for i in range(len(correlation_matrix.columns)):
                    for j in range(i+1, len(correlation_matrix.columns)):
                        corr_value = correlation_matrix.iloc[i, j]
                        if abs(corr_value) > 0.7:  # 높은 상관관계
                            high_corr.append(f"• {correlation_matrix.columns[i]} ↔ {correlation_matrix.columns[j]}: {corr_value:.3f}")
                
                if high_corr:
                    results.append("**높은 상관관계 발견:**")
                    results.extend(high_corr)
                else:
                    results.append("높은 상관관계를 찾을 수 없습니다.")
    
    # 분포 분석
    if "분포" in question_lower:
        if not execution_df.empty and '노출수' in execution_df.columns:
            exposure_stats = execution_df['노출수'].describe()
            results.append("**노출수 분포:**")
            results.append(f"• 평균: {exposure_stats['mean']:,.0f}")
            results.append(f"• 중앙값: {exposure_stats['50%']:,.0f}")
            results.append(f"• 표준편차: {exposure_stats['std']:,.0f}")
            results.append(f"• 최솟값: {exposure_stats['min']:,.0f}")
            results.append(f"• 최댓값: {exposure_stats['max']:,.0f}")
    
    # 패턴 분석
    if "패턴" in question_lower:
        if not execution_df.empty and '날짜' in execution_df.columns:
            execution_df['날짜'] = pd.to_datetime(execution_df['날짜'])
            execution_df['요일'] = execution_df['날짜'].dt.day_name()
            
            # 요일별 평균 노출수
            daily_pattern = execution_df.groupby('요일')['노출수'].mean().sort_values(ascending=False)
            results.append("**요일별 평균 노출수 패턴:**")
            for day, avg_exposure in daily_pattern.items():
                results.append(f"• {day}: {avg_exposure:,.0f}")
    
    return "\n".join(results) if results else "고급 통계 분석을 위한 충분한 데이터가 없습니다."

def analyze_predictions(execution_df, sales_df, question_lower):
    """예측 분석"""
    results = []
    
    # 집행 데이터 예측
    if not execution_df.empty and '날짜' in execution_df.columns:
        execution_df['날짜'] = pd.to_datetime(execution_df['날짜'])
        daily_metrics = execution_df.groupby('날짜').agg({
            '노출수': 'sum',
            '좋아요': 'sum',
            '댓글수': 'sum',
            '조회수': 'sum'
        }).reset_index()
        
        if len(daily_metrics) >= 7:
            # 최근 7일 평균으로 다음 주 예측
            recent_7days = daily_metrics.tail(7)
            avg_exposure = recent_7days['노출수'].mean()
            avg_likes = recent_7days['좋아요'].mean()
            avg_comments = recent_7days['댓글수'].mean()
            avg_views = recent_7days['조회수'].mean()
            
            results.append("**다음 주 예측 (최근 7일 평균 기준):**")
            results.append(f"• 예상 노출수: {avg_exposure:,.0f}")
            results.append(f"• 예상 좋아요: {avg_likes:,.0f}")
            results.append(f"• 예상 댓글수: {avg_comments:,.0f}")
            results.append(f"• 예상 조회수: {avg_views:,.0f}")
    
    # 매출 데이터 예측
    if not sales_df.empty and 'DT' in sales_df.columns:
        sales_df['DT'] = pd.to_datetime(sales_df['DT'])
        daily_sales = sales_df.groupby('DT')['SALE_AMT_TY'].sum().reset_index()
        
        if len(daily_sales) >= 7:
            recent_sales = daily_sales.tail(7)
            avg_sales = recent_sales['SALE_AMT_TY'].mean()
            results.append(f"• 예상 일일 매출: {avg_sales:,.0f}원")
    
    return "\n".join(results) if results else "예측 분석을 위한 충분한 데이터가 없습니다."

def analyze_insights(execution_df, sales_df, influencer_df, question_lower):
    """인사이트 분석"""
    results = []
    
    # 집행 데이터 인사이트
    if not execution_df.empty:
        # 최고 성과 인플루언서
        if '인플루언서' in execution_df.columns and '노출수' in execution_df.columns:
            top_influencer = execution_df.loc[execution_df['노출수'].idxmax()]
            results.append(f"**핵심 인사이트:**")
            results.append(f"• 최고 성과 인플루언서: {top_influencer['인플루언서']} (노출수: {top_influencer['노출수']:,})")
        
        # 효율성 분석
        if '좋아요' in execution_df.columns and '댓글수' in execution_df.columns and '노출수' in execution_df.columns:
            execution_df['engagement_rate'] = (execution_df['좋아요'] + execution_df['댓글수']) / execution_df['노출수'] * 100
            avg_engagement = execution_df['engagement_rate'].mean()
            results.append(f"• 평균 참여율: {avg_engagement:.2f}%")
        
        # 브랜드별 성과 차이
        if '브랜드' in execution_df.columns:
            brand_performance = execution_df.groupby('브랜드')['노출수'].sum().sort_values(ascending=False)
            results.append("**브랜드별 성과 순위:**")
            for i, (brand, exposure) in enumerate(brand_performance.items(), 1):
                results.append(f"• {i}위 {brand}: {exposure:,}")
    
    # 매출 데이터 인사이트
    if not sales_df.empty and 'BRD_CD' in sales_df.columns:
        brand_sales = sales_df.groupby('BRD_CD')['SALE_AMT_TY'].sum().sort_values(ascending=False)
        brand_mapping = {'M': 'MLB', 'X': 'DX', 'V': 'DV', 'ST': 'ST'}
        
        results.append("**브랜드별 매출 순위:**")
        for i, (brand_code, sales) in enumerate(brand_sales.items(), 1):
            brand_name = brand_mapping.get(brand_code, brand_code)
            results.append(f"• {i}위 {brand_name}: {sales:,.0f}원")
    
    return "\n".join(results) if results else "인사이트 분석을 위한 충분한 데이터가 없습니다."

if __name__ == "__main__":
    main()
